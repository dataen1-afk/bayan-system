"""
Certified Clients Registry Routes (BAC-F6-19)
Certified Clients management and Excel export endpoints
"""
from fastapi import APIRouter, HTTPException, Depends
from fastapi.security import HTTPAuthorizationCredentials
from fastapi.responses import FileResponse, Response
from pydantic import BaseModel, Field
from typing import Optional, List
from datetime import datetime, timezone
import uuid
import logging
from io import BytesIO
from collections import Counter

from auth import get_current_user, security
from dependencies import create_notification, CONTRACTS_DIR
import app_documents_pg as doc_pg

# Try to import openpyxl for Excel export
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Try to import PDF generator
try:
    from certified_clients_generator import generate_certified_client_pdf
    PDF_GENERATOR_AVAILABLE = True
except ImportError:
    PDF_GENERATOR_AVAILABLE = False

router = APIRouter(prefix="/certified-clients", tags=["Certified Clients"])


async def _next_certified_serial() -> int:
    clients = await doc_pg.list_ordered(doc_pg.C_CERTIFIED_CLIENTS, 2000)
    if not clients:
        return 1
    m = max((int(c.get("serial_number") or 0) for c in clients), default=0)
    return m + 1


def _sort_by_serial(clients: list) -> list:
    return sorted(clients, key=lambda c: int(c.get("serial_number") or 0))


# ================= MODELS =================

class CertifiedClientCreate(BaseModel):
    client_name: str
    client_name_ar: str = ""
    address: str = ""
    address_ar: str = ""
    contact_person: str = ""
    contact_number: str = ""
    scope: str = ""
    scope_ar: str = ""
    accreditation: List[str] = []
    ea_code: str = ""
    certificate_number: str = ""
    issue_date: str = ""
    expiry_date: str = ""
    surveillance_1_date: str = ""
    surveillance_2_date: str = ""
    recertification_date: str = ""
    linked_certificate_id: str = ""

class CertifiedClientUpdate(BaseModel):
    client_name: Optional[str] = None
    client_name_ar: Optional[str] = None
    address: Optional[str] = None
    address_ar: Optional[str] = None
    contact_person: Optional[str] = None
    contact_number: Optional[str] = None
    scope: Optional[str] = None
    scope_ar: Optional[str] = None
    accreditation: Optional[List[str]] = None
    ea_code: Optional[str] = None
    certificate_number: Optional[str] = None
    issue_date: Optional[str] = None
    expiry_date: Optional[str] = None
    surveillance_1_date: Optional[str] = None
    surveillance_2_date: Optional[str] = None
    recertification_date: Optional[str] = None
    status: Optional[str] = None
    notes: Optional[str] = None

class CertifiedClient(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    serial_number: int = 0
    client_name: str = ""
    client_name_ar: str = ""
    address: str = ""
    address_ar: str = ""
    contact_person: str = ""
    contact_number: str = ""
    scope: str = ""
    scope_ar: str = ""
    accreditation: List[str] = []
    ea_code: str = ""
    certificate_number: str = ""
    issue_date: str = ""
    expiry_date: str = ""
    surveillance_1_date: str = ""
    surveillance_2_date: str = ""
    recertification_date: str = ""
    status: str = "active"
    notes: str = ""
    linked_certificate_id: str = ""
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: Optional[datetime] = None

# ================= ENDPOINTS =================

@router.get("")
async def get_certified_clients(
    status: Optional[str] = None,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Get all certified clients"""
    await get_current_user(credentials)
    
    clients = await doc_pg.list_ordered(doc_pg.C_CERTIFIED_CLIENTS, 1000)
    if status and status != "all":
        clients = [c for c in clients if c.get("status") == status]
    return _sort_by_serial(clients)

@router.post("")
async def create_certified_client(
    data: CertifiedClientCreate,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Create a new certified client record"""
    await get_current_user(credentials)
    
    next_serial = await _next_certified_serial()

    client = CertifiedClient(
        serial_number=next_serial,
        client_name=data.client_name,
        client_name_ar=data.client_name_ar,
        address=data.address,
        address_ar=data.address_ar,
        contact_person=data.contact_person,
        contact_number=data.contact_number,
        scope=data.scope,
        scope_ar=data.scope_ar,
        accreditation=data.accreditation,
        ea_code=data.ea_code,
        certificate_number=data.certificate_number,
        issue_date=data.issue_date,
        expiry_date=data.expiry_date,
        surveillance_1_date=data.surveillance_1_date,
        surveillance_2_date=data.surveillance_2_date,
        recertification_date=data.recertification_date,
        linked_certificate_id=data.linked_certificate_id,
        status="active"
    )
    
    client_doc = client.model_dump()
    client_doc['created_at'] = client_doc['created_at'].isoformat()
    
    await doc_pg.insert_document(doc_pg.C_CERTIFIED_CLIENTS, client_doc)
    
    await create_notification(
        notification_type="certified_client_added",
        title="Certified Client Added",
        title_ar="تمت إضافة عميل معتمد",
        message=f"New certified client added: {client.client_name}",
        message_ar=f"تمت إضافة عميل معتمد جديد: {client.client_name}",
        related_id=client.id,
        related_type="certified_client"
    )
    
    return {"message": "Certified client created", "id": client.id, "serial_number": next_serial}

@router.get("/stats/overview")
async def get_certified_clients_stats(
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Get statistics overview for certified clients"""
    await get_current_user(credentials)
    
    total = await doc_pg.count_all(doc_pg.C_CERTIFIED_CLIENTS)
    active = await doc_pg.count_status(doc_pg.C_CERTIFIED_CLIENTS, "active")
    suspended = await doc_pg.count_status(doc_pg.C_CERTIFIED_CLIENTS, "suspended")
    expired = await doc_pg.count_status(doc_pg.C_CERTIFIED_CLIENTS, "expired")
    withdrawn = await doc_pg.count_status(doc_pg.C_CERTIFIED_CLIENTS, "withdrawn")

    all_clients = await doc_pg.list_ordered(doc_pg.C_CERTIFIED_CLIENTS, 2000)
    ctr: Counter[str] = Counter()
    for c in all_clients:
        for acc in c.get("accreditation") or []:
            if acc:
                ctr[str(acc)] += 1
    by_standard = [{"_id": k, "count": v} for k, v in ctr.most_common(100)]
    
    return {
        "total": total,
        "active": active,
        "suspended": suspended,
        "expired": expired,
        "withdrawn": withdrawn,
        "by_standard": by_standard
    }

@router.get("/{client_id}")
async def get_certified_client(
    client_id: str,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Get a specific certified client"""
    await get_current_user(credentials)
    
    client = await doc_pg.get_by_doc_id(doc_pg.C_CERTIFIED_CLIENTS, client_id)
    if not client:
        raise HTTPException(status_code=404, detail="Certified client not found")

    return client

@router.put("/{client_id}")
async def update_certified_client(
    client_id: str,
    data: CertifiedClientUpdate,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Update a certified client record"""
    await get_current_user(credentials)
    
    existing = await doc_pg.get_by_doc_id(doc_pg.C_CERTIFIED_CLIENTS, client_id)
    if not existing:
        raise HTTPException(status_code=404, detail="Certified client not found")

    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    merged = {**existing, **update_data}
    await doc_pg.replace_payload(
        doc_pg.C_CERTIFIED_CLIENTS, client_id, merged
    )

    return {"message": "Certified client updated"}

@router.delete("/{client_id}")
async def delete_certified_client(
    client_id: str,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Delete a certified client record"""
    await get_current_user(credentials)
    
    existing = await doc_pg.get_by_doc_id(doc_pg.C_CERTIFIED_CLIENTS, client_id)
    if not existing:
        raise HTTPException(status_code=404, detail="Certified client not found")

    await doc_pg.delete_by_doc_id(doc_pg.C_CERTIFIED_CLIENTS, client_id)
    return {"message": "Certified client deleted"}

@router.post("/sync-from-certificates")
async def sync_certified_clients_from_certificates(
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Sync certified clients from active certificates"""
    await get_current_user(credentials)
    
    all_c = await doc_pg.list_ordered(doc_pg.C_CERTIFICATES, 1000)
    certificates = [c for c in all_c if c.get("status") == "active"]
    
    synced_count = 0
    for cert in certificates:
        existing = await doc_pg.get_by_payload_field(
            doc_pg.C_CERTIFIED_CLIENTS,
            "linked_certificate_id",
            str(cert["id"]),
        )

        if not existing:
            next_serial = await _next_certified_serial()

            new_client = {
                "id": str(uuid.uuid4()),
                "serial_number": next_serial,
                "client_name": cert.get('organization_name', ''),
                "client_name_ar": cert.get('organization_name_ar', ''),
                "scope": cert.get('scope', ''),
                "scope_ar": cert.get('scope_ar', ''),
                "accreditation": cert.get('standards', []),
                "certificate_number": cert.get('certificate_number', ''),
                "issue_date": cert.get('issue_date', ''),
                "expiry_date": cert.get('expiry_date', ''),
                "status": "active",
                "linked_certificate_id": cert['id'],
                "created_at": datetime.now(timezone.utc).isoformat()
            }

            await doc_pg.insert_document(doc_pg.C_CERTIFIED_CLIENTS, new_client)
            synced_count += 1
    
    return {"message": f"Synced {synced_count} new certified clients from certificates"}

@router.get("/export/excel")
async def export_certified_clients_excel(
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Export certified clients list to Excel"""
    await get_current_user(credentials)
    
    if not OPENPYXL_AVAILABLE:
        raise HTTPException(status_code=500, detail="Excel export not available - openpyxl not installed")
    
    all_c = await doc_pg.list_ordered(doc_pg.C_CERTIFIED_CLIENTS, 1000)
    clients = _sort_by_serial([c for c in all_c if c.get("status") == "active"])
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Certified Clients"
    
    # Headers
    headers = ["S.No.", "Client Name", "اسم العميل", "Address", "العنوان", "Scope", "النطاق", 
               "Accreditation", "EA Code", "Certificate No.", "Issue Date", "Expiry Date",
               "Surveillance 1", "Surveillance 2", "Recertification", "Status"]
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data rows
    for row_idx, client in enumerate(clients, 2):
        ws.cell(row=row_idx, column=1, value=client.get('serial_number', row_idx-1))
        ws.cell(row=row_idx, column=2, value=client.get('client_name', ''))
        ws.cell(row=row_idx, column=3, value=client.get('client_name_ar', ''))
        ws.cell(row=row_idx, column=4, value=client.get('address', ''))
        ws.cell(row=row_idx, column=5, value=client.get('address_ar', ''))
        ws.cell(row=row_idx, column=6, value=client.get('scope', ''))
        ws.cell(row=row_idx, column=7, value=client.get('scope_ar', ''))
        ws.cell(row=row_idx, column=8, value=', '.join(client.get('accreditation', [])))
        ws.cell(row=row_idx, column=9, value=client.get('ea_code', ''))
        ws.cell(row=row_idx, column=10, value=client.get('certificate_number', ''))
        ws.cell(row=row_idx, column=11, value=client.get('issue_date', ''))
        ws.cell(row=row_idx, column=12, value=client.get('expiry_date', ''))
        ws.cell(row=row_idx, column=13, value=client.get('surveillance_1_date', ''))
        ws.cell(row=row_idx, column=14, value=client.get('surveillance_2_date', ''))
        ws.cell(row=row_idx, column=15, value=client.get('recertification_date', ''))
        ws.cell(row=row_idx, column=16, value=client.get('status', '').title())
    
    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"certified_clients_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.get("/{client_id}/pdf")
async def get_certified_client_pdf(
    client_id: str,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Generate PDF for a certified client"""
    await get_current_user(credentials)
    
    client = await doc_pg.get_by_doc_id(doc_pg.C_CERTIFIED_CLIENTS, client_id)
    if not client:
        raise HTTPException(status_code=404, detail="Certified client not found")

    if not PDF_GENERATOR_AVAILABLE:
        raise HTTPException(status_code=500, detail="PDF generator not available")
    
    try:
        pdf_path = str(CONTRACTS_DIR / f"certified_client_{client_id[:8]}.pdf")
        generate_certified_client_pdf(client, pdf_path)
        
        return FileResponse(
            pdf_path,
            media_type="application/pdf",
            filename=f"certified_client_{client_id[:8]}.pdf"
        )
    except Exception as e:
        logging.error(f"Error generating Certified Client PDF: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error generating PDF: {str(e)}")
