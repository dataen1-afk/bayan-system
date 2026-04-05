"""
Suspended Clients Registry Routes (BAC-F6-20)
Suspended Clients management and Excel export endpoints
"""
from fastapi import APIRouter, HTTPException, Depends
from fastapi.security import HTTPAuthorizationCredentials
from fastapi.responses import FileResponse, Response
from pydantic import BaseModel, Field
from typing import Optional, List
from datetime import datetime, timezone
import uuid
import logging
from io import BytesIO

from auth import get_current_user, security
from dependencies import create_notification, CONTRACTS_DIR
import app_documents_pg as doc_pg

# Try to import openpyxl for Excel export
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Try to import PDF generator
try:
    from suspended_clients_generator import generate_suspended_client_pdf
    PDF_GENERATOR_AVAILABLE = True
except ImportError:
    PDF_GENERATOR_AVAILABLE = False

router = APIRouter(prefix="/suspended-clients", tags=["Suspended Clients"])


async def _next_suspended_serial() -> int:
    clients = await doc_pg.list_ordered(doc_pg.C_SUSPENDED_CLIENTS, 2000)
    if not clients:
        return 1
    m = max((int(c.get("serial_number") or 0) for c in clients), default=0)
    return m + 1


def _sort_suspended_by_serial(clients: list) -> list:
    return sorted(clients, key=lambda c: int(c.get("serial_number") or 0))


async def _set_certified_status(certified_id: str, status: str) -> None:
    cur = await doc_pg.get_by_doc_id(doc_pg.C_CERTIFIED_CLIENTS, certified_id)
    if not cur:
        return
    cur["status"] = status
    cur["updated_at"] = datetime.now(timezone.utc).isoformat()
    await doc_pg.replace_payload(doc_pg.C_CERTIFIED_CLIENTS, certified_id, cur)


# ================= MODELS =================

class SuspendedClientCreate(BaseModel):
    client_id: str = ""
    client_name: str
    client_name_ar: str = ""
    address: str = ""
    address_ar: str = ""
    registration_date: str = ""
    suspended_on: str = ""
    reason_for_suspension: str = ""
    reason_for_suspension_ar: str = ""
    future_action: str = ""
    remarks: str = ""
    linked_certified_client_id: str = ""

class SuspendedClientUpdate(BaseModel):
    client_id: Optional[str] = None
    client_name: Optional[str] = None
    client_name_ar: Optional[str] = None
    address: Optional[str] = None
    address_ar: Optional[str] = None
    registration_date: Optional[str] = None
    suspended_on: Optional[str] = None
    reason_for_suspension: Optional[str] = None
    reason_for_suspension_ar: Optional[str] = None
    future_action: Optional[str] = None
    remarks: Optional[str] = None
    status: Optional[str] = None
    lifted_on: Optional[str] = None
    lifted_reason: Optional[str] = None

class SuspendedClient(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    serial_number: int = 0
    client_id: str = ""
    client_name: str = ""
    client_name_ar: str = ""
    address: str = ""
    address_ar: str = ""
    registration_date: str = ""
    suspended_on: str = ""
    reason_for_suspension: str = ""
    reason_for_suspension_ar: str = ""
    future_action: str = ""
    remarks: str = ""
    status: str = "suspended"
    lifted_on: str = ""
    lifted_reason: str = ""
    linked_certified_client_id: str = ""
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: Optional[datetime] = None

# ================= ENDPOINTS =================

@router.get("")
async def get_suspended_clients(
    status: Optional[str] = None,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Get all suspended clients"""
    await get_current_user(credentials)
    
    clients = await doc_pg.list_ordered(doc_pg.C_SUSPENDED_CLIENTS, 1000)
    if status and status != "all":
        clients = [c for c in clients if c.get("status") == status]
    return _sort_suspended_by_serial(clients)

@router.post("")
async def create_suspended_client(
    data: SuspendedClientCreate,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Create a new suspended client record"""
    await get_current_user(credentials)
    
    next_serial = await _next_suspended_serial()

    client = SuspendedClient(
        serial_number=next_serial,
        client_id=data.client_id,
        client_name=data.client_name,
        client_name_ar=data.client_name_ar,
        address=data.address,
        address_ar=data.address_ar,
        registration_date=data.registration_date,
        suspended_on=data.suspended_on,
        reason_for_suspension=data.reason_for_suspension,
        reason_for_suspension_ar=data.reason_for_suspension_ar,
        future_action=data.future_action,
        remarks=data.remarks,
        linked_certified_client_id=data.linked_certified_client_id,
        status="suspended"
    )
    
    client_doc = client.model_dump()
    client_doc['created_at'] = client_doc['created_at'].isoformat()
    
    await doc_pg.insert_document(doc_pg.C_SUSPENDED_CLIENTS, client_doc)

    # If linked to a certified client, update its status
    if data.linked_certified_client_id:
        await _set_certified_status(
            data.linked_certified_client_id, "suspended"
        )
    
    await create_notification(
        notification_type="client_suspended",
        title="Client Suspended",
        title_ar="تم إيقاف العميل",
        message=f"Client suspended: {client.client_name}",
        message_ar=f"تم إيقاف العميل: {client.client_name}",
        related_id=client.id,
        related_type="suspended_client"
    )
    
    return {"message": "Suspended client record created", "id": client.id, "serial_number": next_serial}

@router.get("/stats/overview")
async def get_suspended_clients_stats(
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Get statistics overview for suspended clients"""
    await get_current_user(credentials)
    
    total = await doc_pg.count_all(doc_pg.C_SUSPENDED_CLIENTS)
    suspended = await doc_pg.count_status(doc_pg.C_SUSPENDED_CLIENTS, "suspended")
    reinstated = await doc_pg.count_status(doc_pg.C_SUSPENDED_CLIENTS, "reinstated")
    withdrawn = await doc_pg.count_status(doc_pg.C_SUSPENDED_CLIENTS, "withdrawn")

    by_action = await doc_pg.count_group_by_payload_text_field(
        doc_pg.C_SUSPENDED_CLIENTS, "future_action", limit=100
    )
    
    return {
        "total": total,
        "suspended": suspended,
        "reinstated": reinstated,
        "withdrawn": withdrawn,
        "by_future_action": by_action
    }

@router.get("/{client_id}")
async def get_suspended_client(
    client_id: str,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Get a specific suspended client"""
    await get_current_user(credentials)
    
    client = await doc_pg.get_by_doc_id(doc_pg.C_SUSPENDED_CLIENTS, client_id)
    if not client:
        raise HTTPException(status_code=404, detail="Suspended client not found")

    return client

@router.put("/{client_id}")
async def update_suspended_client(
    client_id: str,
    data: SuspendedClientUpdate,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Update a suspended client record"""
    await get_current_user(credentials)
    
    existing = await doc_pg.get_by_doc_id(doc_pg.C_SUSPENDED_CLIENTS, client_id)
    if not existing:
        raise HTTPException(status_code=404, detail="Suspended client not found")

    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    merged = {**existing, **update_data}
    await doc_pg.replace_payload(
        doc_pg.C_SUSPENDED_CLIENTS, client_id, merged
    )

    return {"message": "Suspended client updated"}

@router.delete("/{client_id}")
async def delete_suspended_client(
    client_id: str,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Delete a suspended client record"""
    await get_current_user(credentials)
    
    existing = await doc_pg.get_by_doc_id(doc_pg.C_SUSPENDED_CLIENTS, client_id)
    if not existing:
        raise HTTPException(status_code=404, detail="Suspended client not found")

    await doc_pg.delete_by_doc_id(doc_pg.C_SUSPENDED_CLIENTS, client_id)
    return {"message": "Suspended client deleted"}

@router.post("/{client_id}/lift-suspension")
async def lift_suspension(
    client_id: str,
    data: dict,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Lift suspension and optionally reinstate the client"""
    await get_current_user(credentials)
    
    existing = await doc_pg.get_by_doc_id(doc_pg.C_SUSPENDED_CLIENTS, client_id)
    if not existing:
        raise HTTPException(status_code=404, detail="Suspended client not found")

    new_status = data.get('new_status', 'reinstated')
    if new_status not in ['reinstated', 'withdrawn']:
        raise HTTPException(status_code=400, detail="Invalid status. Use 'reinstated' or 'withdrawn'")

    update_data = {
        "status": new_status,
        "lifted_on": data.get('lifted_on', datetime.now().strftime("%Y-%m-%d")),
        "lifted_reason": data.get('lifted_reason', ''),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }

    merged = {**existing, **update_data}
    await doc_pg.replace_payload(
        doc_pg.C_SUSPENDED_CLIENTS, client_id, merged
    )

    # Update linked certified client status
    if existing.get('linked_certified_client_id'):
        certified_status = "active" if new_status == "reinstated" else "withdrawn"
        await _set_certified_status(
            existing['linked_certified_client_id'], certified_status
        )
    
    status_ar = {'reinstated': 'تم إعادة التفعيل', 'withdrawn': 'تم السحب'}.get(new_status, new_status)
    
    await create_notification(
        notification_type="suspension_lifted",
        title=f"Suspension {new_status.title()}",
        title_ar=f"رفع الإيقاف - {status_ar}",
        message=f"Client {existing.get('client_name', '')} suspension lifted - {new_status}",
        message_ar=f"تم رفع إيقاف العميل {existing.get('client_name', '')} - {status_ar}",
        related_id=client_id,
        related_type="suspended_client"
    )
    
    return {"message": f"Suspension lifted - status changed to {new_status}"}

@router.post("/sync-from-certified")
async def sync_suspended_from_certified(
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Sync suspended clients from certified clients with suspended status"""
    await get_current_user(credentials)
    
    all_cert = await doc_pg.list_ordered(doc_pg.C_CERTIFIED_CLIENTS, 2000)
    suspended_certified = [c for c in all_cert if c.get("status") == "suspended"]

    synced_count = 0
    for cert_client in suspended_certified:
        existing = await doc_pg.get_by_payload_field(
            doc_pg.C_SUSPENDED_CLIENTS,
            "linked_certified_client_id",
            str(cert_client["id"]),
        )

        if not existing:
            next_serial = await _next_suspended_serial()

            new_suspended = {
                "id": str(uuid.uuid4()),
                "serial_number": next_serial,
                "client_id": str(cert_client.get('serial_number', '')),
                "client_name": cert_client.get('client_name', ''),
                "client_name_ar": cert_client.get('client_name_ar', ''),
                "address": cert_client.get('address', ''),
                "address_ar": cert_client.get('address_ar', ''),
                "registration_date": cert_client.get('issue_date', ''),
                "suspended_on": datetime.now().strftime("%Y-%m-%d"),
                "reason_for_suspension": "Synced from certified clients",
                "status": "suspended",
                "linked_certified_client_id": cert_client['id'],
                "created_at": datetime.now(timezone.utc).isoformat()
            }

            await doc_pg.insert_document(doc_pg.C_SUSPENDED_CLIENTS, new_suspended)
            synced_count += 1
    
    return {"message": f"Synced {synced_count} suspended clients from certified clients"}

@router.get("/export/excel")
async def export_suspended_clients_excel(
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Export suspended clients list to Excel"""
    await get_current_user(credentials)
    
    if not OPENPYXL_AVAILABLE:
        raise HTTPException(status_code=500, detail="Excel export not available - openpyxl not installed")
    
    clients = _sort_suspended_by_serial(
        await doc_pg.list_ordered(doc_pg.C_SUSPENDED_CLIENTS, 1000)
    )
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Suspended Clients"
    
    # Headers
    headers = ["Sr. No.", "Client ID", "Client Name", "اسم العميل", "Address", "العنوان",
               "Registration Date", "Suspended On", "Reason", "سبب الإيقاف", 
               "Future Action", "Remarks", "Status", "Lifted On", "Lifted Reason"]
    
    header_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data rows
    for row_idx, client in enumerate(clients, 2):
        ws.cell(row=row_idx, column=1, value=client.get('serial_number', row_idx-1))
        ws.cell(row=row_idx, column=2, value=client.get('client_id', ''))
        ws.cell(row=row_idx, column=3, value=client.get('client_name', ''))
        ws.cell(row=row_idx, column=4, value=client.get('client_name_ar', ''))
        ws.cell(row=row_idx, column=5, value=client.get('address', ''))
        ws.cell(row=row_idx, column=6, value=client.get('address_ar', ''))
        ws.cell(row=row_idx, column=7, value=client.get('registration_date', ''))
        ws.cell(row=row_idx, column=8, value=client.get('suspended_on', ''))
        ws.cell(row=row_idx, column=9, value=client.get('reason_for_suspension', ''))
        ws.cell(row=row_idx, column=10, value=client.get('reason_for_suspension_ar', ''))
        ws.cell(row=row_idx, column=11, value=client.get('future_action', ''))
        ws.cell(row=row_idx, column=12, value=client.get('remarks', ''))
        ws.cell(row=row_idx, column=13, value=client.get('status', '').title())
        ws.cell(row=row_idx, column=14, value=client.get('lifted_on', ''))
        ws.cell(row=row_idx, column=15, value=client.get('lifted_reason', ''))
    
    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"suspended_clients_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.get("/{client_id}/pdf")
async def get_suspended_client_pdf(
    client_id: str,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Generate PDF for a suspended client"""
    await get_current_user(credentials)
    
    client = await doc_pg.get_by_doc_id(doc_pg.C_SUSPENDED_CLIENTS, client_id)
    if not client:
        raise HTTPException(status_code=404, detail="Suspended client not found")

    if not PDF_GENERATOR_AVAILABLE:
        raise HTTPException(status_code=500, detail="PDF generator not available")
    
    try:
        pdf_path = str(CONTRACTS_DIR / f"suspended_client_{client_id[:8]}.pdf")
        generate_suspended_client_pdf(client, pdf_path)
        
        return FileResponse(
            pdf_path,
            media_type="application/pdf",
            filename=f"suspended_client_{client_id[:8]}.pdf"
        )
    except Exception as e:
        logging.error(f"Error generating Suspended Client PDF: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error generating PDF: {str(e)}")
