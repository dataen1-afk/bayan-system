from fastapi import FastAPI, APIRouter, HTTPException, Depends, status
from fastapi.responses import FileResponse, Response, RedirectResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, EmailStr
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone, timedelta
from passlib.context import CryptContext
import jwt
from reportlab.lib.pagesizes import letter, A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import inch, cm
from reportlab.lib.colors import HexColor
import aiosmtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import qrcode
from io import BytesIO
import base64

# Import audit calculator
from audit_calculator import calculate_total_audit_time

# Import PDF generators
from pdf_generator import generate_contract_pdf
from bilingual_pdf_generator import generate_bilingual_contract_pdf
from certificate_generator import generate_certificate_pdf, get_qr_code_base64
from grant_agreement_generator import generate_grant_agreement_pdf
from contract_review_generator import generate_contract_review_pdf
from audit_program_generator import generate_audit_program_pdf
from job_order_generator import generate_job_order_pdf
from stage1_audit_plan_generator import generate_stage1_audit_plan_pdf
from stage2_audit_plan_generator import generate_stage2_audit_plan_pdf
from opening_closing_meeting_generator import generate_opening_closing_meeting_pdf
from stage1_audit_report_generator import generate_stage1_audit_report_pdf
from stage2_audit_report_generator import generate_stage2_audit_report_pdf
from auditor_notes_generator import generate_auditor_notes_pdf
from nonconformity_report_generator import generate_nonconformity_report_pdf
from certificate_data_generator import generate_certificate_data_pdf
from technical_review_generator import generate_technical_review_pdf
from customer_feedback_generator import generate_customer_feedback_pdf
from pre_transfer_review_generator import generate_pre_transfer_review_pdf
from certified_clients_generator import generate_certified_client_pdf, generate_certified_clients_list_pdf

# Import modular routers
from routes.auth import router as auth_router
from routes.notifications import router as notifications_router
from routes.sites import router as sites_router
from routes.contacts import router as contacts_router
from routes.documents import router as documents_router

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Create the main app without a prefix
app = FastAPI()

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# Password hashing
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")

# JWT settings
JWT_SECRET = os.environ.get('JWT_SECRET', 'your-secret-key-change-in-production')
JWT_ALGORITHM = "HS256"
JWT_EXPIRATION_HOURS = 24

# Security
security = HTTPBearer()

# Ensure contracts directory exists
CONTRACTS_DIR = ROOT_DIR / "contracts"
CONTRACTS_DIR.mkdir(exist_ok=True)

# ================= MODELS =================

class UserRole:
    ADMIN = "admin"
    CLIENT = "client"

class UserRegister(BaseModel):
    name: str
    email: EmailStr
    password: str
    role: str = Field(default=UserRole.CLIENT)

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class User(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    email: EmailStr
    role: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class TokenResponse(BaseModel):
    token: str
    user: User

class FormField(BaseModel):
    label: str
    type: str  # text, textarea, number, email, etc.
    required: bool = True
    options: Optional[List[str]] = None

class FormCreate(BaseModel):
    client_id: str
    fields: List[FormField]

class Form(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    client_id: str
    fields: List[FormField]
    responses: Optional[Dict[str, Any]] = None
    status: str = "pending"  # pending, submitted
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class FormSubmission(BaseModel):
    responses: Dict[str, Any]

# ================= QUOTATION/PROPOSAL MODELS =================

class AuditDuration(BaseModel):
    stage_1: float = 0
    stage_2: float = 0
    surveillance_1: float = 0
    surveillance_2: float = 0
    recertification: float = 0

class ServiceFees(BaseModel):
    initial_certification: float = 0
    surveillance_1: float = 0
    surveillance_2: float = 0
    recertification: float = 0
    currency: str = "SAR"

class ProposalCreate(BaseModel):
    application_form_id: str
    # Organization details (pre-filled from application)
    organization_name: str
    organization_address: str
    organization_phone: str
    contact_person: str
    contact_position: str
    contact_email: EmailStr
    # Service details
    standards: List[str]
    scope: str
    total_employees: int
    number_of_sites: int
    # Audit duration
    audit_duration: AuditDuration
    # Pricing
    service_fees: ServiceFees
    # Additional notes
    notes: str = ""
    # Validity
    validity_days: int = 30
    # First Party (Bayan) authorized signatory
    issuer_name: str = "Abdullah Al-Rashid"  # Default authorized signatory
    issuer_designation: str = "General Manager"  # Default job title
    issuer_signature: str = ""  # Base64 encoded signature image
    issuer_stamp: str = ""  # Base64 encoded stamp image

class Proposal(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    access_token: str = Field(default_factory=lambda: str(uuid.uuid4()))  # For client access
    application_form_id: str
    # Organization details
    organization_name: str
    organization_address: str
    organization_phone: str
    contact_person: str
    contact_position: str
    contact_email: EmailStr
    # Service details
    standards: List[str]
    scope: str
    total_employees: int
    number_of_sites: int
    # Audit duration
    audit_duration: AuditDuration
    # Pricing
    service_fees: ServiceFees
    total_amount: float = 0
    # Additional
    notes: str = ""
    validity_days: int = 30
    # Status
    status: str = "pending"  # pending, accepted, rejected, expired
    # First Party (Bayan) Signatures
    issuer_name: str = ""
    issuer_designation: str = ""
    issuer_signature: str = ""  # Base64 encoded signature image
    issuer_stamp: str = ""  # Base64 encoded stamp image
    issued_date: Optional[datetime] = None
    # Client response
    client_response_date: Optional[datetime] = None
    client_signatory_name: str = ""
    client_signatory_designation: str = ""
    rejection_reason: str = ""
    # Timestamps
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ProposalResponse(BaseModel):
    status: str  # accepted or rejected
    signatory_name: str = ""
    signatory_designation: str = ""
    rejection_reason: str = ""

class ModificationRequest(BaseModel):
    comment: str
    requested_changes: str = ""

class PublicProposalResponse(BaseModel):
    id: str
    organization_name: str
    organization_address: str = ""
    contact_person: str
    contact_email: str
    standards: List[str]
    scope: str
    total_employees: int
    number_of_sites: int
    audit_duration: AuditDuration
    service_fees: ServiceFees
    total_amount: float
    notes: str
    validity_days: int
    status: str
    issuer_name: str
    issuer_designation: str
    issued_date: Optional[datetime]
    modification_comment: str = ""
    modification_requested_changes: str = ""

# ================= CERTIFICATION AGREEMENT MODELS =================

class AgreementAcknowledgements(BaseModel):
    certificationRules: bool = False
    publicDirectory: bool = False
    certificationCommunication: bool = False
    surveillanceSchedule: bool = False
    nonconformityResolution: bool = False
    feesAndPayment: bool = False

class CertificationAgreementSubmit(BaseModel):
    organization_name: str
    organization_address: str
    selected_standards: List[str]
    other_standard: str = ""
    scope_of_services: str
    sites: List[str]
    signatory_name: str
    signatory_position: str
    signatory_date: str
    acknowledgements: AgreementAcknowledgements
    signature_image: str  # Base64 encoded signature image (required)
    stamp_image: str  # Base64 encoded stamp image (required)

class CertificationAgreement(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    proposal_id: str
    proposal_access_token: str
    organization_name: str
    organization_address: str
    selected_standards: List[str]
    other_standard: str = ""
    scope_of_services: str
    sites: List[str]
    signatory_name: str
    signatory_position: str
    signatory_date: str
    acknowledgements: Dict[str, bool] = {}
    signature_image: Optional[str] = None  # Base64 encoded signature image
    stamp_image: Optional[str] = None  # Base64 encoded stamp image
    status: str = "submitted"  # submitted, contract_generated
    contract_pdf_path: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# ================= INVOICE & PAYMENT MODELS =================

class InvoiceItem(BaseModel):
    description: str
    description_ar: str = ""
    quantity: int = 1
    unit_price: float
    total: float = 0

class Invoice(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    invoice_number: str = ""  # Auto-generated: INV-2026-001
    contract_id: str  # Links to proposal/agreement
    organization_name: str
    organization_address: str = ""
    contact_email: str = ""
    contact_phone: str = ""
    # Invoice details
    items: List[InvoiceItem] = []
    subtotal: float = 0
    tax_rate: float = 15  # VAT percentage (Saudi Arabia: 15%)
    tax_amount: float = 0
    total_amount: float = 0
    currency: str = "SAR"
    # Payment terms
    issue_date: str = ""  # When invoice was created
    due_date: str = ""  # Payment due date
    payment_terms: str = "net_30"  # net_15, net_30, net_60, due_on_receipt
    # Status tracking
    status: str = "draft"  # draft, sent, viewed, paid, overdue, cancelled
    paid_amount: float = 0
    paid_date: Optional[str] = None
    payment_method: str = ""  # bank_transfer, stripe, cash, cheque
    payment_reference: str = ""  # Transaction ID or reference
    # Stripe integration
    stripe_invoice_id: str = ""
    stripe_payment_intent_id: str = ""
    stripe_payment_url: str = ""
    # Notes
    notes: str = ""
    internal_notes: str = ""
    # Timestamps
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: Optional[datetime] = None
    sent_at: Optional[datetime] = None

class InvoiceCreate(BaseModel):
    contract_id: str
    items: List[InvoiceItem] = []
    tax_rate: float = 15
    due_date: str = ""
    payment_terms: str = "net_30"
    notes: str = ""

class PaymentRecord(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    invoice_id: str
    amount: float
    payment_date: str
    payment_method: str  # bank_transfer, stripe, cash, cheque
    reference: str = ""  # Transaction ID
    notes: str = ""
    recorded_by: str = ""  # Admin who recorded the payment
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# ================= CERTIFICATE MODELS =================

class Certificate(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    certificate_number: str = ""  # Auto-generated: CERT-YYYY-XXXX
    contract_id: str  # Reference to certification agreement
    audit_id: str = ""  # Reference to the final audit that granted certification
    # Organization details
    organization_name: str
    organization_name_ar: str = ""
    # Certification details
    standards: List[str] = []  # ISO 9001, ISO 14001, etc.
    scope: str = ""
    scope_ar: str = ""
    # Dates
    issue_date: str  # YYYY-MM-DD
    expiry_date: str  # YYYY-MM-DD (typically 3 years from issue)
    # Status
    status: str = "active"  # active, suspended, withdrawn, expired
    # QR Code
    verification_url: str = ""  # URL for QR code verification
    qr_code_data: str = ""  # Base64 encoded QR code image
    # Audit details
    lead_auditor: str = ""
    audit_team: List[str] = []
    # Timestamps
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: Optional[datetime] = None

class CertificateCreate(BaseModel):
    contract_id: str
    audit_id: str = ""
    standards: List[str] = []
    scope: str = ""
    scope_ar: str = ""
    lead_auditor: str = ""
    audit_team: List[str] = []

# ================= NOTIFICATION MODELS =================

class Notification(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    type: str  # form_submitted, proposal_sent, proposal_accepted, proposal_rejected, agreement_signed
    title: str
    message: str
    related_id: Optional[str] = None  # ID of related form/proposal/agreement
    related_type: Optional[str] = None  # form, proposal, agreement
    is_read: bool = False
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class NotificationCreate(BaseModel):
    type: str
    title: str
    message: str
    related_id: Optional[str] = None
    related_type: Optional[str] = None

# ================= CONTRACT REVIEW / AUDIT PROGRAM MODELS =================

class AuditTimeEntry(BaseModel):
    standard: str = ""  # ISO 9001, ISO 14001, etc.
    audit_type: str = ""  # SA1, SA2, SUR1, SUR2, RE
    num_days: float = 0
    percent_increase: float = 0
    percent_reduction: float = 0
    final_days: float = 0

class ContractReviewCreate(BaseModel):
    agreement_id: str  # Reference to certification agreement

class ContractReviewClientSubmit(BaseModel):
    """Data submitted by client"""
    consultant_name: str = ""
    consultant_affects_impartiality: bool = False
    consultant_impact_explanation: str = ""
    exclusions_justification: str = ""

class ContractReviewAdminUpdate(BaseModel):
    """Data filled by admin/auditor"""
    contract_review_date: str = ""
    risk_category: str = ""  # For QMS
    complexity_category: str = ""  # For other standards
    integration_level_percent: float = 100
    combined_audit_ability_percent: float = 100
    auditor_code_matched: bool = False
    # Audit time calculations
    audit_times: List[AuditTimeEntry] = []
    final_man_days: float = 0
    # Team assignment
    lead_auditor_id: str = ""
    lead_auditor_name: str = ""
    auditor_ids: List[str] = []
    auditor_names: List[str] = []
    other_team_members: str = ""
    technical_expert_needed: bool = False
    technical_expert_name: str = ""
    certification_decision_maker: str = ""
    # Approval
    prepared_by_name: str = ""
    prepared_by_date: str = ""
    reviewed_by_name: str = ""
    reviewed_by_date: str = ""

class ContractReview(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    access_token: str = Field(default_factory=lambda: str(uuid.uuid4()))
    agreement_id: str  # Reference to certification agreement
    # Auto-populated from agreement
    organization_name: str = ""
    standards: List[str] = []
    scope_of_services: str = ""
    total_employees: int = 0
    application_date: str = ""
    client_id: str = ""
    # Client-filled data
    consultant_name: str = ""
    consultant_affects_impartiality: bool = False
    consultant_impact_explanation: str = ""
    exclusions_justification: str = ""
    client_submitted: bool = False
    client_submitted_at: Optional[datetime] = None
    # Admin-filled data
    contract_review_date: str = ""
    risk_category: str = ""
    complexity_category: str = ""
    integration_level_percent: float = 100
    combined_audit_ability_percent: float = 100
    auditor_code_matched: bool = False
    audit_times: List[Dict[str, Any]] = []
    final_man_days: float = 0
    lead_auditor_id: str = ""
    lead_auditor_name: str = ""
    auditor_ids: List[str] = []
    auditor_names: List[str] = []
    other_team_members: str = ""
    technical_expert_needed: bool = False
    technical_expert_name: str = ""
    certification_decision_maker: str = ""
    prepared_by_name: str = ""
    prepared_by_date: str = ""
    reviewed_by_name: str = ""
    reviewed_by_date: str = ""
    # Status
    status: str = "pending_client"  # pending_client, pending_review, completed
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: Optional[datetime] = None

# ================= AUDIT PROGRAM MODELS (BACF6-05) =================

class AuditActivityEntry(BaseModel):
    """Single audit activity row in the program"""
    activity: str = ""  # Audit activity description
    audit_type: str = ""  # Desktop, On-site, etc.
    stage1: str = ""  # Stage 1 allocation (e.g., "1 day")
    stage2: str = ""  # Stage 2 allocation
    sur1: str = ""  # Surveillance 1 allocation
    sur2: str = ""  # Surveillance 2 allocation
    rc: str = ""  # Recertification allocation
    planned_date: str = ""  # YYYY-MM-DD

class AuditProgramCreate(BaseModel):
    """Create a new Audit Program from a Contract Review"""
    contract_review_id: str  # Reference to contract review

class AuditProgramUpdate(BaseModel):
    """Update audit program data"""
    num_shifts: int = 1
    activities: List[AuditActivityEntry] = []
    certification_manager: str = ""
    approval_date: str = ""
    notes: str = ""

class AuditProgram(BaseModel):
    """Audit Program (BACF6-05) - schedules audit stages"""
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    access_token: str = Field(default_factory=lambda: str(uuid.uuid4()))
    contract_review_id: str  # Reference to contract review
    agreement_id: str = ""  # Reference to certification agreement
    # Auto-populated from contract review/agreement
    organization_name: str = ""
    standards: List[str] = []
    scope_of_services: str = ""
    total_employees: int = 0
    # Program data
    num_shifts: int = 1
    activities: List[Dict[str, Any]] = []  # List of AuditActivityEntry dicts
    # Approval
    certification_manager: str = ""
    approval_date: str = ""
    notes: str = ""
    # Status
    status: str = "draft"  # draft, approved, in_progress, completed
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: Optional[datetime] = None

# ================= JOB ORDER MODELS (BACF6-06) =================

class JobOrderCreate(BaseModel):
    """Create a new Job Order from an Audit Program"""
    audit_program_id: str  # Reference to audit program
    auditor_id: str  # Selected auditor from system
    position: str = "Auditor"  # Auditor, Lead Auditor, Technical Expert
    audit_type: str = ""  # Stage 1, Stage 2, Surveillance 1, Surveillance 2, Recertification
    audit_date: str = ""  # YYYY-MM-DD

class JobOrderUpdate(BaseModel):
    """Update job order (admin)"""
    position: str = "Auditor"
    audit_type: str = ""
    audit_date: str = ""
    certification_manager: str = ""
    manager_approval_date: str = ""
    notes: str = ""

class JobOrderAuditorConfirmation(BaseModel):
    """Auditor confirmation of job order"""
    confirmed: bool = True
    unable_reason: str = ""  # If not confirmed, reason why

class JobOrder(BaseModel):
    """Job Order (BACF6-06) - Auditor appointment for audit"""
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    access_token: str = Field(default_factory=lambda: str(uuid.uuid4()))  # For auditor confirmation link
    audit_program_id: str  # Reference to audit program
    contract_review_id: str = ""  # Reference to contract review
    agreement_id: str = ""  # Reference to certification agreement
    # Auditor info
    auditor_id: str
    auditor_name: str = ""
    auditor_name_ar: str = ""
    auditor_email: str = ""
    position: str = "Auditor"  # Auditor, Lead Auditor, Technical Expert
    # Client/Audit details (from audit program)
    organization_name: str = ""
    organization_address: str = ""
    total_employees: int = 0
    phone: str = ""
    scope_of_services: str = ""
    standards: List[str] = []
    audit_type: str = ""  # Stage 1, Stage 2, Surveillance 1, Surveillance 2, Recertification
    audit_date: str = ""
    client_ref: str = ""
    # Manager approval
    certification_manager: str = ""
    manager_approval_date: str = ""
    manager_approved: bool = False
    # Auditor confirmation
    auditor_confirmed: bool = False
    auditor_confirmation_date: str = ""
    unable_reason: str = ""  # If auditor can't perform
    # Status
    status: str = "pending_approval"  # pending_approval, approved, pending_auditor, confirmed, rejected, completed
    notes: str = ""
    # Timestamps
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: Optional[datetime] = None

# ================= STAGE 1 AUDIT PLAN MODELS (BACF6-07) =================

class ScheduleEntry(BaseModel):
    """Single schedule entry in the audit plan"""
    date_time: str = ""  # Date and time range
    process: str = ""  # Process being audited
    process_owner: str = ""  # Process owner/auditee
    clauses: str = ""  # Applicable ISO clauses
    auditor: str = ""  # Assigned auditor name

class TeamMember(BaseModel):
    """Team member in the audit"""
    auditor_id: str = ""
    name: str = ""
    name_ar: str = ""
    role: str = "Auditor"  # Team Leader, Auditor, Technical Expert, Evaluator

class Stage1AuditPlanCreate(BaseModel):
    """Create a new Stage 1 Audit Plan from a confirmed Job Order"""
    job_order_id: str  # Reference to confirmed job order

class Stage1AuditPlanUpdate(BaseModel):
    """Update Stage 1 Audit Plan data"""
    # Client contact
    contact_person: str = ""
    contact_phone: str = ""
    contact_designation: str = ""
    contact_email: str = ""
    # Audit details
    audit_language: str = "English"
    audit_date_from: str = ""
    audit_date_to: str = ""
    # Team members (additional to team leader)
    team_members: List[TeamMember] = []
    # Schedule
    schedule_entries: List[ScheduleEntry] = []
    notes: str = ""

class Stage1AuditPlanClientResponse(BaseModel):
    """Client response to the audit plan"""
    accepted: bool = True
    change_requests: str = ""  # If not accepted, requested changes

class Stage1AuditPlan(BaseModel):
    """Stage 1 Audit Plan (BACF6-07) - Phase 1 audit planning"""
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    access_token: str = Field(default_factory=lambda: str(uuid.uuid4()))  # For client review link
    job_order_id: str  # Reference to job order
    audit_program_id: str = ""
    contract_review_id: str = ""
    # Client information (from job order)
    organization_name: str = ""
    file_no: str = ""
    address: str = ""
    plan_date: str = ""
    contact_person: str = ""
    contact_phone: str = ""
    contact_designation: str = ""
    contact_email: str = ""
    # Audit details (from job order)
    standards: List[str] = []
    audit_language: str = "English"
    audit_type: str = "Stage 1"
    audit_date_from: str = ""
    audit_date_to: str = ""
    scope: str = ""
    # Team (Team Leader from Job Order)
    team_leader: Dict[str, Any] = {}  # {auditor_id, name, name_ar, role}
    team_members: List[Dict[str, Any]] = []  # Additional team members
    # Schedule
    schedule_entries: List[Dict[str, Any]] = []  # List of ScheduleEntry dicts
    # Manager approval (internal)
    manager_approved: bool = False
    manager_name: str = ""
    manager_approval_date: str = ""
    # Client acceptance
    sent_to_client: bool = False
    client_accepted: bool = False
    client_acceptance_date: str = ""
    client_change_requests: str = ""
    # Status
    status: str = "draft"  # draft, pending_manager, manager_approved, pending_client, client_accepted, changes_requested, completed
    notes: str = ""
    # Timestamps
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: Optional[datetime] = None

# ================= STAGE 2 AUDIT PLAN MODELS (BACF6-08) =================

class Stage2AuditPlanCreate(BaseModel):
    """Create a new Stage 2 Audit Plan - from Stage 1 or independently"""
    stage1_plan_id: Optional[str] = None  # Optional - if from Stage 1 plan
    job_order_id: Optional[str] = None  # Optional - if creating independently
    audit_type: str = "Stage 2"  # Stage 2, Renewal, Surveillance

class Stage2AuditPlanUpdate(BaseModel):
    """Update Stage 2 Audit Plan data"""
    contact_person: str = ""
    contact_phone: str = ""
    contact_designation: str = ""
    contact_email: str = ""
    audit_language: str = "English"
    audit_type: str = "Stage 2"
    audit_date_from: str = ""
    audit_date_to: str = ""
    team_members: List[TeamMember] = []
    schedule_entries: List[ScheduleEntry] = []
    notes: str = ""

class Stage2AuditPlanClientResponse(BaseModel):
    """Client response to the Stage 2 audit plan"""
    accepted: bool = True
    change_requests: str = ""

class Stage2AuditPlan(BaseModel):
    """Stage 2 Audit Plan (BACF6-08) - Phase 2 audit planning"""
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    access_token: str = Field(default_factory=lambda: str(uuid.uuid4()))
    stage1_plan_id: Optional[str] = None  # Reference to Stage 1 plan (if applicable)
    job_order_id: str = ""
    audit_program_id: str = ""
    contract_review_id: str = ""
    # Client information
    organization_name: str = ""
    file_no: str = ""
    address: str = ""
    plan_date: str = ""
    contact_person: str = ""
    contact_phone: str = ""
    contact_designation: str = ""
    contact_email: str = ""
    # Audit details
    standards: List[str] = []
    audit_language: str = "English"
    audit_type: str = "Stage 2"  # Stage 2, Renewal, Surveillance
    audit_date_from: str = ""
    audit_date_to: str = ""
    scope: str = ""
    # Team
    team_leader: Dict[str, Any] = {}
    team_members: List[Dict[str, Any]] = []
    # Schedule
    schedule_entries: List[Dict[str, Any]] = []
    # Manager approval
    manager_approved: bool = False
    manager_name: str = ""
    manager_approval_date: str = ""
    # Client acceptance
    sent_to_client: bool = False
    client_accepted: bool = False
    client_acceptance_date: str = ""
    client_change_requests: str = ""
    # Status
    status: str = "draft"
    notes: str = ""
    # Timestamps
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: Optional[datetime] = None

# ================= OPENING & CLOSING MEETING MODELS (BACF6-09) =================

class MeetingAttendee(BaseModel):
    """Attendee for opening/closing meeting"""
    name: str = ""
    designation: str = ""
    opening_meeting_date: str = ""  # Date attended opening meeting
    closing_meeting_date: str = ""  # Date attended closing meeting

class OpeningClosingMeetingCreate(BaseModel):
    """Create Opening & Closing Meeting form - sent after Stage 1 audit"""
    stage1_plan_id: Optional[str] = None  # Reference to Stage 1 plan
    job_order_id: Optional[str] = None  # Alternative reference

class OpeningClosingMeetingSubmit(BaseModel):
    """Client submits the meeting attendance form"""
    attendees: List[MeetingAttendee] = []
    opening_meeting_notes: str = ""
    closing_meeting_notes: str = ""

class OpeningClosingMeeting(BaseModel):
    """Opening & Closing Meeting (BACF6-09) - Meeting attendance record"""
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    access_token: str = Field(default_factory=lambda: str(uuid.uuid4()))
    stage1_plan_id: Optional[str] = None
    job_order_id: str = ""
    audit_program_id: str = ""
    # Company info
    organization_name: str = ""
    file_no: str = ""
    address: str = ""
    # Audit info
    audit_type: str = ""
    audit_date: str = ""
    standards: List[str] = []
    # Meeting attendees
    attendees: List[Dict[str, Any]] = []  # List of MeetingAttendee dicts
    # Notes
    opening_meeting_notes: str = ""
    closing_meeting_notes: str = ""
    # Status
    status: str = "pending"  # pending, submitted
    sent_to_client: bool = False
    submitted_date: str = ""
    # Timestamps
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: Optional[datetime] = None

# ================= STAGE 1 AUDIT REPORT MODELS (BACF6-10) =================

class AuditFinding(BaseModel):
    """Finding in audit report"""
    department: str = ""
    finding: str = ""

class AuditConcern(BaseModel):
    """Area of concern in audit report"""
    department: str = ""
    concern: str = ""
    rating: int = 1  # 1=OFI, 2=Probable NC, 3=Not ready

class ChecklistItem(BaseModel):
    """Checklist item for audit"""
    requirement: str = ""
    status: str = "C"  # C=Conforming, NC=Non-Conforming, O=Observation
    comments: str = ""

class Stage1AuditReportCreate(BaseModel):
    """Create Stage 1 Audit Report"""
    stage1_plan_id: Optional[str] = None
    meeting_id: Optional[str] = None  # Opening/Closing Meeting reference

class Stage1AuditReportUpdate(BaseModel):
    """Update Stage 1 Audit Report"""
    # Change details
    employee_change: Optional[str] = None
    scope_change: Optional[str] = None
    integrated_system: Optional[str] = None
    additional_info: Optional[str] = None
    man_days_adequate: Optional[bool] = None
    # Findings
    positive_findings: Optional[List[Dict[str, Any]]] = None
    areas_of_concern: Optional[List[Dict[str, Any]]] = None
    # Declarations
    declarations: Optional[Dict[str, bool]] = None
    # Recommendation
    recommendation: Optional[str] = None  # proceed, not_proceed, further_stage1
    # Checklist
    checklist_items: Optional[List[Dict[str, Any]]] = None
    # Notes
    notes: Optional[str] = None

class Stage1AuditReport(BaseModel):
    """Stage 1 Audit Report (BACF6-10)"""
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    stage1_plan_id: str = ""
    meeting_id: str = ""
    job_order_id: str = ""
    audit_program_id: str = ""
    # Organization details
    organization_name: str = ""
    address: str = ""
    site_address: str = ""
    standards: List[str] = []
    num_employees: str = ""
    num_shifts: str = ""
    email: str = ""
    contact_person: str = ""
    phone: str = ""
    scope: str = ""
    ea_code: str = ""
    exclusions: str = ""
    # Audit team
    audit_team: Dict[str, Any] = {}  # lead_auditor, auditors, technical_experts
    audit_duration: str = ""
    start_date: str = ""
    end_date: str = ""
    # Brief & objectives
    organization_brief: str = ""
    audit_objective: str = "To evaluate the client's documented system, location planning and readiness for stage 2 audit."
    audit_criteria: str = "The requirements of a defined normative document or the certified management system requirements"
    # Change details
    employee_change: str = ""
    scope_change: str = ""
    integrated_system: str = ""
    additional_info: str = ""
    man_days_adequate: bool = True
    # Attendees (from meeting)
    attendees: List[Dict[str, Any]] = []
    # Findings
    positive_findings: List[Dict[str, Any]] = []  # List of AuditFinding dicts
    areas_of_concern: List[Dict[str, Any]] = []  # List of AuditConcern dicts
    # Team leader declarations
    declarations: Dict[str, bool] = {}
    # Recommendation
    recommendation: str = ""  # proceed, not_proceed, further_stage1
    # Checklist
    checklist_items: List[Dict[str, Any]] = []  # List of ChecklistItem dicts
    # Status
    status: str = "draft"  # draft, completed, approved
    completed_date: str = ""
    approved_by: str = ""
    approved_date: str = ""
    notes: str = ""
    # Timestamps
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: Optional[datetime] = None

# ================= STAGE 2 AUDIT REPORT MODELS (BACF6-11) =================

class Nonconformity(BaseModel):
    """Nonconformity finding in Stage 2 audit"""
    clause: str = ""  # Standard clause reference
    description: str = ""
    rating: int = 1  # 1=Minor NC, 2=Major NC

class OpportunityForImprovement(BaseModel):
    """Opportunity for improvement"""
    department: str = ""
    recommendation: str = ""

class Stage2AuditReportCreate(BaseModel):
    """Create Stage 2 Audit Report"""
    stage2_plan_id: Optional[str] = None
    stage1_report_id: Optional[str] = None

class Stage2AuditReportUpdate(BaseModel):
    """Update Stage 2 Audit Report"""
    # Change details
    employee_change: Optional[str] = None
    scope_change: Optional[str] = None
    integrated_system: Optional[str] = None
    additional_info: Optional[str] = None
    # Findings
    positive_findings: Optional[List[Dict[str, Any]]] = None
    opportunities_for_improvement: Optional[List[Dict[str, Any]]] = None
    nonconformities: Optional[List[Dict[str, Any]]] = None
    # Certification recommendations
    certification_recommendation: Optional[Dict[str, bool]] = None
    overall_recommendation: Optional[str] = None
    # Checklist
    checklist_items: Optional[List[Dict[str, Any]]] = None
    notes: Optional[str] = None

class Stage2AuditReport(BaseModel):
    """Stage 2 Audit Report (BACF6-11)"""
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    stage2_plan_id: str = ""
    stage1_report_id: str = ""
    job_order_id: str = ""
    audit_program_id: str = ""
    # Organization details
    organization_name: str = ""
    address: str = ""
    site_address: str = ""
    standards: List[str] = []
    num_employees: str = ""
    num_shifts: str = ""
    email: str = ""
    contact_person: str = ""
    phone: str = ""
    scope: str = ""
    ea_code: str = ""
    exclusions: str = ""
    # Audit team
    audit_team: Dict[str, Any] = {}
    audit_duration: str = ""
    start_date: str = ""
    end_date: str = ""
    # Change details
    employee_change: str = ""
    scope_change: str = ""
    integrated_system: str = ""
    additional_info: str = ""
    # Attendees
    attendees: List[Dict[str, Any]] = []
    # Findings
    positive_findings: List[Dict[str, Any]] = []
    opportunities_for_improvement: List[Dict[str, Any]] = []  # OFI list
    nonconformities: List[Dict[str, Any]] = []  # NC list with rating
    # Certification recommendations (checkboxes)
    certification_recommendation: Dict[str, bool] = {}  # issue_certificate, use_logo, etc.
    # Overall recommendation
    overall_recommendation: str = ""  # recommend_certification, recommend_minor_nc, major_nc_evidence, not_recommended
    # Checklist
    checklist_items: List[Dict[str, Any]] = []
    # Status
    status: str = "draft"  # draft, completed, approved
    completed_date: str = ""
    approved_by: str = ""
    approved_date: str = ""
    notes: str = ""
    # Timestamps
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: Optional[datetime] = None

# ================= AUDITOR NOTES MODELS (BACF6-12) =================

class AuditorNotesCreate(BaseModel):
    """Create Auditor Notes - can be from Stage 2 report or independently"""
    stage2_report_id: Optional[str] = None  # Optional link to Stage 2 report
    # If not from report, these fields can be provided
    client_name: str = ""
    location: str = ""
    standards: List[str] = []
    auditor_name: str = ""
    audit_type: str = ""  # Stage 1, Stage 2, Surveillance, Recertification
    audit_date: str = ""
    department: str = ""

class AuditorNotesUpdate(BaseModel):
    """Update Auditor Notes"""
    client_name: Optional[str] = None
    location: Optional[str] = None
    standards: Optional[List[str]] = None
    auditor_name: Optional[str] = None
    audit_type: Optional[str] = None
    audit_date: Optional[str] = None
    department: Optional[str] = None
    notes: Optional[str] = None
    notes_ar: Optional[str] = None

class AuditorNotes(BaseModel):
    """Auditor Notes (BACF6-12) - Created by auditor after audit"""
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    stage2_report_id: str = ""
    stage1_report_id: str = ""
    job_order_id: str = ""
    audit_program_id: str = ""
    contract_review_id: str = ""
    # Form fields
    client_name: str = ""
    client_name_ar: str = ""
    location: str = ""
    location_ar: str = ""
    standards: List[str] = []
    auditor_id: str = ""
    auditor_name: str = ""
    auditor_name_ar: str = ""
    audit_type: str = ""
    audit_date: str = ""
    department: str = ""
    department_ar: str = ""
    # Notes content
    notes: str = ""
    notes_ar: str = ""
    # Status
    status: str = "draft"  # draft, completed
    completed_at: Optional[datetime] = None
    # Timestamps
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: Optional[datetime] = None

# ================= NONCONFORMITY REPORT MODELS (BACF6-13) =================

class NonconformityItem(BaseModel):
    """Single nonconformity entry in the report"""
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    standard_clause: str = ""
    description: str = ""
    description_ar: str = ""
    nc_type: str = "minor"  # "major" or "minor"
    root_cause: str = ""
    root_cause_ar: str = ""
    corrections: str = ""
    corrections_ar: str = ""
    corrective_actions: str = ""
    corrective_actions_ar: str = ""
    verification_evidence: str = ""
    verification_decision: str = ""
    verification_date: str = ""
    verified_by: str = ""
    status: str = "open"  # open, closed, pending_verification

class NonconformityReportCreate(BaseModel):
    """Create Nonconformity Report"""
    stage2_report_id: Optional[str] = None
    client_name: str = ""
    certificate_no: str = ""
    standards: List[str] = []
    audit_type: str = ""
    audit_date: str = ""
    lead_auditor: str = ""
    management_representative: str = ""

class NonconformityReportUpdate(BaseModel):
    """Update Nonconformity Report"""
    client_name: Optional[str] = None
    certificate_no: Optional[str] = None
    standards: Optional[List[str]] = None
    audit_type: Optional[str] = None
    audit_date: Optional[str] = None
    lead_auditor: Optional[str] = None
    management_representative: Optional[str] = None
    nonconformities: Optional[List[Dict[str, Any]]] = None
    submission_deadline: Optional[str] = None
    verification_options: Optional[Dict[str, bool]] = None
    management_rep_date: Optional[str] = None
    audit_team_leader_date: Optional[str] = None
    evidence_submission_date: Optional[str] = None
    final_date: Optional[str] = None

class NonconformityReport(BaseModel):
    """Nonconformity Report (BACF6-13)"""
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    stage2_report_id: str = ""
    job_order_id: str = ""
    audit_program_id: str = ""
    contract_review_id: str = ""
    # Header fields
    client_name: str = ""
    client_name_ar: str = ""
    certificate_no: str = ""
    standards: List[str] = []
    audit_type: str = ""
    audit_date: str = ""
    lead_auditor: str = ""
    lead_auditor_ar: str = ""
    management_representative: str = ""
    management_representative_ar: str = ""
    # Nonconformities
    nonconformities: List[Dict[str, Any]] = []
    submission_deadline: str = ""
    # Verification options
    verification_options: Dict[str, bool] = {}
    # Signatures
    management_rep_date: str = ""
    audit_team_leader_date: str = ""
    evidence_submission_date: str = ""
    final_date: str = ""
    # Status and stats
    status: str = "draft"  # draft, sent_to_client, pending_verification, closed
    total_major: int = 0
    total_minor: int = 0
    closed_count: int = 0
    # Timestamps
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: Optional[datetime] = None

# ================= CERTIFICATE DATA MODELS (BACF6-14) =================

class CertificateDataCreate(BaseModel):
    """Create Certificate Data"""
    nc_report_id: Optional[str] = None
    stage2_report_id: Optional[str] = None
    client_name: str = ""
    standards: List[str] = []
    lead_auditor: str = ""
    audit_type: str = ""
    audit_date: str = ""

class CertificateDataUpdate(BaseModel):
    """Update Certificate Data"""
    client_name: Optional[str] = None
    standards: Optional[List[str]] = None
    lead_auditor: Optional[str] = None
    audit_type: Optional[str] = None
    audit_date: Optional[str] = None
    agreed_certification_scope: Optional[str] = None
    agreed_certification_scope_ar: Optional[str] = None
    ea_code: Optional[str] = None
    technical_category: Optional[str] = None
    company_data_local: Optional[str] = None
    certification_scope_local: Optional[str] = None
    company_data_english: Optional[str] = None
    certification_scope_english: Optional[str] = None

class CertificateDataClientConfirm(BaseModel):
    """Client confirmation"""
    client_signature: str = ""
    client_stamp: str = ""
    signature_date: str = ""

class CertificateData(BaseModel):
    """Certificate Data (BACF6-14)"""
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    access_token: str = Field(default_factory=lambda: str(uuid.uuid4()))
    nc_report_id: str = ""
    stage2_report_id: str = ""
    job_order_id: str = ""
    audit_program_id: str = ""
    contract_review_id: str = ""
    agreement_id: str = ""
    # Header fields
    client_name: str = ""
    client_name_ar: str = ""
    standards: List[str] = []
    lead_auditor: str = ""
    lead_auditor_ar: str = ""
    audit_type: str = ""
    audit_date: str = ""
    # Certification scope
    agreed_certification_scope: str = ""
    agreed_certification_scope_ar: str = ""
    ea_code: str = ""
    technical_category: str = ""
    # Company data - Local Language
    company_data_local: str = ""
    certification_scope_local: str = ""
    # Company data - English
    company_data_english: str = ""
    certification_scope_english: str = ""
    # Client confirmation
    client_signature: str = ""
    client_stamp: str = ""
    client_signature_date: str = ""
    client_confirmed: bool = False
    client_confirmed_at: Optional[datetime] = None
    # Certificate details
    certificate_number: str = ""
    issue_date: str = ""
    expiry_date: str = ""
    certificate_generated: bool = False
    # Status
    status: str = "draft"  # draft, sent_to_client, client_confirmed, certificate_issued
    # Timestamps
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: Optional[datetime] = None

# ================= TECHNICAL REVIEW MODELS (BACF6-15) =================

class TechnicalReviewChecklistItem(BaseModel):
    """Individual checklist item for technical review"""
    category: str = ""  # General, Application, Quotation, Man Days, Audit Team, Stage 1, Stage 2, Other
    item: str = ""
    item_ar: str = ""
    checked: Optional[bool] = None  # True=Yes, False=No, None=Not evaluated
    remarks: str = ""

class TechnicalReviewCreate(BaseModel):
    """Create Technical Review request"""
    stage2_report_id: str = ""  # Optional - link to Stage 2 report
    # Header info - can be auto-filled or manual
    client_name: str
    client_name_ar: str = ""
    location: str = ""
    scope: str = ""
    ea_code: str = ""
    standards: List[str] = []
    audit_type: str = ""  # Initial, Surveillance, Recertification
    audit_dates: str = ""
    # Audit team
    audit_team_members: List[str] = []
    technical_expert: str = ""

class TechnicalReviewUpdate(BaseModel):
    """Update Technical Review"""
    client_name: Optional[str] = None
    client_name_ar: Optional[str] = None
    location: Optional[str] = None
    scope: Optional[str] = None
    ea_code: Optional[str] = None
    standards: Optional[List[str]] = None
    audit_type: Optional[str] = None
    audit_dates: Optional[str] = None
    audit_team_members: Optional[List[str]] = None
    technical_expert: Optional[str] = None
    # Checklist items
    checklist_items: Optional[List[dict]] = None
    # Technical reviewer
    technical_reviewer: Optional[str] = None
    review_date: Optional[str] = None
    review_comments: Optional[str] = None
    # Certification decision
    certification_decision: Optional[str] = None  # issue_certificate, reject_certificate, needs_review
    decision_comments: Optional[str] = None
    # Approval
    approved_by: Optional[str] = None
    approval_date: Optional[str] = None

class TechnicalReview(BaseModel):
    """Technical Review and Certification Decision (BACF6-15)"""
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    stage2_report_id: str = ""  # Optional link to Stage 2 Audit Report
    job_order_id: str = ""
    audit_program_id: str = ""
    contract_review_id: str = ""
    agreement_id: str = ""
    # Header info
    client_name: str = ""
    client_name_ar: str = ""
    location: str = ""
    scope: str = ""
    ea_code: str = ""
    standards: List[str] = []
    audit_type: str = ""
    audit_dates: str = ""
    # Audit team
    audit_team_members: List[str] = []
    technical_expert: str = ""
    # Checklist items with Y/N and remarks
    checklist_items: List[dict] = []
    # Technical reviewer
    technical_reviewer: str = ""
    review_date: str = ""
    review_comments: str = ""
    # Certification Decision
    certification_decision: str = ""  # issue_certificate, reject_certificate, needs_review
    decision_comments: str = ""
    # Approval
    approved_by: str = ""
    approval_date: str = ""
    # Certificate generated
    certificate_id: str = ""  # If decision is issue_certificate
    certificate_number: str = ""
    # Status
    status: str = "draft"  # draft, under_review, decision_made, approved, certificate_issued
    # Timestamps
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: Optional[datetime] = None

# Default checklist items for Technical Review
DEFAULT_TECHNICAL_REVIEW_CHECKLIST = [
    {"category": "General", "item": "Correctness of Name, address, scope, Ref no.", "item_ar": "صحة الاسم والعنوان والنطاق ورقم المرجع", "checked": None, "remarks": ""},
    {"category": "General", "item": "Scope Terminology", "item_ar": "مصطلحات النطاق", "checked": None, "remarks": ""},
    {"category": "General", "item": "EA Code verification", "item_ar": "التحقق من رمز EA", "checked": None, "remarks": ""},
    {"category": "General", "item": "Conformance to audited processes", "item_ar": "المطابقة للعمليات المدققة", "checked": None, "remarks": ""},
    {"category": "Application", "item": "Application form complete and signed by both parties", "item_ar": "نموذج الطلب مكتمل وموقع من الطرفين", "checked": None, "remarks": ""},
    {"category": "Quotation", "item": "Quotation signed by both BAC & client representative", "item_ar": "عرض السعر موقع من بيان والعميل", "checked": None, "remarks": ""},
    {"category": "Man Days", "item": "Adequacy per IAF guidelines", "item_ar": "الكفاية وفقًا لإرشادات IAF", "checked": None, "remarks": ""},
    {"category": "Man Days", "item": "Adequacy on complexity", "item_ar": "الكفاية من حيث التعقيد", "checked": None, "remarks": ""},
    {"category": "Man Days", "item": "Reduction/change in man days justified", "item_ar": "تبرير التخفيض/التغيير في أيام العمل", "checked": None, "remarks": ""},
    {"category": "Audit Team", "item": "Team competent for scope", "item_ar": "الفريق مؤهل للنطاق", "checked": None, "remarks": ""},
    {"category": "Stage 1", "item": "Audit plan available and contains all requirements", "item_ar": "خطة التدقيق متوفرة وتحتوي على جميع المتطلبات", "checked": None, "remarks": ""},
    {"category": "Stage 1", "item": "Opening & Closing Meeting attendance available", "item_ar": "سجل حضور الاجتماع الافتتاحي والختامي متوفر", "checked": None, "remarks": ""},
    {"category": "Stage 1", "item": "Duration corresponds to contract review", "item_ar": "المدة تتوافق مع مراجعة العقد", "checked": None, "remarks": ""},
    {"category": "Stage 1", "item": "Audit report covers all areas including scope validation", "item_ar": "تقرير التدقيق يغطي جميع المجالات بما في ذلك التحقق من النطاق", "checked": None, "remarks": ""},
    {"category": "Stage 1", "item": "Findings clearly highlighted as concerns", "item_ar": "النتائج موضحة بوضوح كمخاوف", "checked": None, "remarks": ""},
    {"category": "Stage 2", "item": "Stage 2 audit plan available and complete", "item_ar": "خطة تدقيق المرحلة 2 متوفرة ومكتملة", "checked": None, "remarks": ""},
    {"category": "Stage 2", "item": "Opening & Closing Meeting attendance available", "item_ar": "سجل حضور الاجتماع الافتتاحي والختامي متوفر", "checked": None, "remarks": ""},
    {"category": "Stage 2", "item": "Stage 2 Audit Report correct and complete", "item_ar": "تقرير تدقيق المرحلة 2 صحيح ومكتمل", "checked": None, "remarks": ""},
    {"category": "Stage 2", "item": "Evidence for NC Closure recorded", "item_ar": "دليل إغلاق عدم المطابقة مسجل", "checked": None, "remarks": ""},
    {"category": "Other", "item": "Conflict of Interest and Confidentiality Undertaking", "item_ar": "إقرار تضارب المصالح والسرية", "checked": None, "remarks": ""},
    {"category": "Other", "item": "Customer Feedback Form reviewed", "item_ar": "مراجعة نموذج ملاحظات العميل", "checked": None, "remarks": ""},
    {"category": "Other", "item": "Certificate data request reviewed", "item_ar": "مراجعة طلب بيانات الشهادة", "checked": None, "remarks": ""},
]

# ================= CUSTOMER FEEDBACK MODELS (BACF6-16) =================

class FeedbackQuestion(BaseModel):
    """Individual feedback question with rating"""
    category: str = ""
    category_ar: str = ""
    question: str = ""
    question_ar: str = ""
    rating: Optional[int] = None  # 1-5 or None for N/A
    
class CustomerFeedbackCreate(BaseModel):
    """Create Customer Feedback request"""
    # Link to audit/certificate
    certificate_id: str = ""
    audit_id: str = ""
    # Organization info
    organization_name: str
    organization_name_ar: str = ""
    # Audit details
    audit_type: str = ""  # Initial, 1st Surveillance, 2nd Surveillance, Re-Certification
    standards: List[str] = []
    audit_date: str = ""
    lead_auditor: str = ""
    auditor: str = ""

class CustomerFeedbackSubmit(BaseModel):
    """Client submission data"""
    questions: List[dict] = []  # List of {question_id, rating}
    want_same_team: Optional[bool] = None
    suggestions: str = ""
    respondent_name: str = ""
    respondent_designation: str = ""

class CustomerFeedback(BaseModel):
    """Customer Feedback Survey (BACF6-16)"""
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    access_token: str = Field(default_factory=lambda: str(uuid.uuid4()))
    # Links
    certificate_id: str = ""
    audit_id: str = ""
    # Organization info
    organization_name: str = ""
    organization_name_ar: str = ""
    # Audit details
    audit_type: str = ""
    standards: List[str] = []
    audit_date: str = ""
    lead_auditor: str = ""
    auditor: str = ""
    # Feedback questions with ratings
    questions: List[dict] = []
    # Additional questions
    want_same_team: Optional[bool] = None
    suggestions: str = ""
    # Respondent info
    respondent_name: str = ""
    respondent_designation: str = ""
    submission_date: str = ""
    # Calculated scores
    overall_score: float = 0.0  # Percentage
    evaluation_result: str = ""  # excellent, good, average, unsatisfactory
    # Internal review
    reviewed_by: str = ""
    review_date: str = ""
    review_comments: str = ""
    # Status
    status: str = "pending"  # pending, submitted, reviewed
    # Timestamps
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    submitted_at: Optional[datetime] = None

# Default feedback questions
DEFAULT_FEEDBACK_QUESTIONS = [
    {"category": "BAC Office", "category_ar": "مكتب بيان", "question": "Responsiveness to your enquiries - Promptness", "question_ar": "الاستجابة لاستفساراتكم - السرعة", "rating": None},
    {"category": "BAC Office", "category_ar": "مكتب بيان", "question": "Accuracy of the quotes communicated to you", "question_ar": "دقة عروض الأسعار المقدمة لكم", "rating": None},
    {"category": "BAC Office", "category_ar": "مكتب بيان", "question": "Handling of your Complaint(s)", "question_ar": "التعامل مع شكواكم", "rating": None},
    {"category": "Audit Preparation", "category_ar": "التحضير للتدقيق", "question": "The audit Plan was sent sufficiently in advance", "question_ar": "تم إرسال خطة التدقيق مسبقاً بوقت كافٍ", "rating": None},
    {"category": "Audit Preparation", "category_ar": "التحضير للتدقيق", "question": "The audit team was well prepared for audit", "question_ar": "كان فريق التدقيق مستعداً جيداً للتدقيق", "rating": None},
    {"category": "Punctuality", "category_ar": "الالتزام بالمواعيد", "question": "The audit carried out as per the programme", "question_ar": "تم التدقيق وفقاً للبرنامج", "rating": None},
    {"category": "Audit", "category_ar": "التدقيق", "question": "Opening and closing meetings were professional", "question_ar": "كانت الاجتماعات الافتتاحية والختامية احترافية", "rating": None},
    {"category": "Audit", "category_ar": "التدقيق", "question": "Questions asked were relevant and easy to understand", "question_ar": "كانت الأسئلة ذات صلة وسهلة الفهم", "rating": None},
    {"category": "Audit", "category_ar": "التدقيق", "question": "The audit team gave enough explanation for your questions", "question_ar": "قدم فريق التدقيق شرحاً كافياً لأسئلتكم", "rating": None},
    {"category": "Audit", "category_ar": "التدقيق", "question": "The audit team was fair and impartial", "question_ar": "كان فريق التدقيق عادلاً ومحايداً", "rating": None},
    {"category": "Ethics", "category_ar": "الأخلاقيات", "question": "The audit team concentrated on the audit", "question_ar": "ركز فريق التدقيق على التدقيق", "rating": None},
    {"category": "Ethics", "category_ar": "الأخلاقيات", "question": "The audit team didn't make unreasonable demands", "question_ar": "لم يقدم فريق التدقيق طلبات غير معقولة", "rating": None},
    {"category": "Effectiveness", "category_ar": "الفعالية", "question": "Issues found were helpful for improving your system", "question_ar": "كانت النتائج مفيدة لتحسين نظامكم", "rating": None},
]

def calculate_feedback_score(questions: List[dict]) -> tuple:
    """Calculate overall score and evaluation result from ratings"""
    total_score = 0
    rated_count = 0
    
    for q in questions:
        rating = q.get('rating')
        if rating is not None and rating != 'na':
            try:
                rating_int = int(rating)
                if 1 <= rating_int <= 5:
                    total_score += rating_int
                    rated_count += 1
            except:
                pass
    
    if rated_count == 0:
        return 0.0, "unsatisfactory"
    
    # Calculate percentage (max is 5 per question)
    percentage = (total_score / (rated_count * 5)) * 100
    
    # Determine evaluation result
    if percentage >= 90:
        evaluation = "excellent"
    elif percentage >= 75:
        evaluation = "good"
    elif percentage >= 60:
        evaluation = "average"
    else:
        evaluation = "unsatisfactory"
    
    return round(percentage, 1), evaluation

# ================= PRE-TRANSFER REVIEW MODELS (BACF6-17) =================

class PreTransferChecklist(BaseModel):
    """Compliance checklist for pre-transfer review"""
    suspension_status: Optional[bool] = None  # True=Yes suspended, False=No, None=N/A
    threat_of_suspension: Optional[bool] = None
    minor_nc_outstanding: Optional[bool] = None
    major_nc_outstanding: Optional[bool] = None
    legal_representation: Optional[bool] = None
    complaints_handled: Optional[bool] = None
    within_bac_scope: Optional[bool] = None
    previous_reports_available: Optional[bool] = None

class PreTransferReviewCreate(BaseModel):
    """Create Pre-Transfer Review request"""
    # Client info
    client_name: str
    client_name_ar: str = ""
    client_address: str = ""
    client_phone: str = ""
    enquiry_reference: str = ""
    # Transfer details
    transfer_reason: str = ""
    existing_cb: str = ""  # Existing Certification Body
    certificate_number: str = ""
    validity: str = ""
    scope: str = ""
    sites: str = ""
    eac_code: str = ""
    standards: List[str] = []

class PreTransferReviewUpdate(BaseModel):
    """Update Pre-Transfer Review"""
    client_name: Optional[str] = None
    client_name_ar: Optional[str] = None
    client_address: Optional[str] = None
    client_phone: Optional[str] = None
    enquiry_reference: Optional[str] = None
    transfer_reason: Optional[str] = None
    existing_cb: Optional[str] = None
    certificate_number: Optional[str] = None
    validity: Optional[str] = None
    scope: Optional[str] = None
    sites: Optional[str] = None
    eac_code: Optional[str] = None
    standards: Optional[List[str]] = None
    # Checklist
    checklist: Optional[dict] = None
    certification_cycle_stage: Optional[str] = None
    # Attachments
    has_previous_audit_report: Optional[bool] = None
    has_previous_certificates: Optional[bool] = None
    # Decision
    transfer_decision: Optional[str] = None  # approved, rejected, pending
    decision_reason: Optional[str] = None
    # Review/Approval
    reviewed_by: Optional[str] = None
    review_date: Optional[str] = None
    approved_by: Optional[str] = None
    approval_date: Optional[str] = None

class PreTransferReview(BaseModel):
    """Pre-Transfer Review (BACF6-17)"""
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    # Client info
    client_name: str = ""
    client_name_ar: str = ""
    client_address: str = ""
    client_phone: str = ""
    enquiry_reference: str = ""
    # Transfer details
    transfer_reason: str = ""
    existing_cb: str = ""
    certificate_number: str = ""
    validity: str = ""
    scope: str = ""
    sites: str = ""
    eac_code: str = ""
    standards: List[str] = []
    # Compliance checklist
    checklist: dict = {}
    certification_cycle_stage: str = ""
    # Attachments
    has_previous_audit_report: bool = False
    has_previous_certificates: bool = False
    attachments: List[str] = []  # Document IDs
    # Decision
    transfer_decision: str = "pending"  # pending, approved, rejected
    decision_reason: str = ""
    # Review and Approval
    reviewed_by: str = ""
    review_date: str = ""
    approved_by: str = ""
    approval_date: str = ""
    # Status
    status: str = "draft"  # draft, under_review, decision_made
    # Timestamps
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: Optional[datetime] = None

# Default checklist items for Pre-Transfer Review
DEFAULT_PRETRANSFER_CHECKLIST = {
    "suspension_status": None,
    "threat_of_suspension": None,
    "minor_nc_outstanding": None,
    "major_nc_outstanding": None,
    "legal_representation": None,
    "complaints_handled": None,
    "within_bac_scope": None,
    "previous_reports_available": None
}

# ================= CERTIFIED CLIENTS REGISTRY MODELS (BAC-F6-19) =================

class CertifiedClientCreate(BaseModel):
    """Create Certified Client Record"""
    client_name: str
    client_name_ar: str = ""
    address: str = ""
    address_ar: str = ""
    contact_person: str = ""
    contact_number: str = ""
    scope: str = ""
    scope_ar: str = ""
    accreditation: List[str] = []  # ISO 9001:2015, ISO 14001:2015, etc.
    ea_code: str = ""  # EA Code or Food category
    certificate_number: str = ""
    issue_date: str = ""  # YYYY-MM-DD
    expiry_date: str = ""  # YYYY-MM-DD
    surveillance_1_date: str = ""  # YYYY-MM-DD
    surveillance_2_date: str = ""  # YYYY-MM-DD
    recertification_date: str = ""  # YYYY-MM-DD
    # Optional link to existing certificate
    linked_certificate_id: str = ""

class CertifiedClientUpdate(BaseModel):
    """Update Certified Client Record"""
    client_name: Optional[str] = None
    client_name_ar: Optional[str] = None
    address: Optional[str] = None
    address_ar: Optional[str] = None
    contact_person: Optional[str] = None
    contact_number: Optional[str] = None
    scope: Optional[str] = None
    scope_ar: Optional[str] = None
    accreditation: Optional[List[str]] = None
    ea_code: Optional[str] = None
    certificate_number: Optional[str] = None
    issue_date: Optional[str] = None
    expiry_date: Optional[str] = None
    surveillance_1_date: Optional[str] = None
    surveillance_2_date: Optional[str] = None
    recertification_date: Optional[str] = None
    status: Optional[str] = None
    notes: Optional[str] = None

class CertifiedClient(BaseModel):
    """Certified Client Registry Record (BAC-F6-19)"""
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    serial_number: int = 0  # Auto-incremented S.No.
    # Client Information
    client_name: str = ""
    client_name_ar: str = ""
    address: str = ""
    address_ar: str = ""
    contact_person: str = ""
    contact_number: str = ""
    # Certification Details
    scope: str = ""
    scope_ar: str = ""
    accreditation: List[str] = []  # Standards: ISO 9001:2015, etc.
    ea_code: str = ""  # EA Code/Food category
    certificate_number: str = ""
    # Important Dates
    issue_date: str = ""  # YYYY-MM-DD
    expiry_date: str = ""  # YYYY-MM-DD
    surveillance_1_date: str = ""  # YYYY-MM-DD
    surveillance_2_date: str = ""  # YYYY-MM-DD
    recertification_date: str = ""  # YYYY-MM-DD
    # Status
    status: str = "active"  # active, suspended, withdrawn, expired
    notes: str = ""
    # Link to system certificate (optional)
    linked_certificate_id: str = ""
    # Timestamps
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: Optional[datetime] = None

# ================= SUSPENDED CLIENTS REGISTRY MODELS (BAC-F6-20) =================

class SuspendedClientCreate(BaseModel):
    """Create Suspended Client Record"""
    client_id: str = ""  # Reference ID from certified clients
    client_name: str
    client_name_ar: str = ""
    address: str = ""
    address_ar: str = ""
    registration_date: str = ""  # Original certification date YYYY-MM-DD
    suspended_on: str = ""  # Suspension date YYYY-MM-DD
    reason_for_suspension: str = ""
    reason_for_suspension_ar: str = ""
    future_action: str = ""  # Planned action (reinstate, withdraw, etc.)
    remarks: str = ""
    # Link to certified client record
    linked_certified_client_id: str = ""

class SuspendedClientUpdate(BaseModel):
    """Update Suspended Client Record"""
    client_id: Optional[str] = None
    client_name: Optional[str] = None
    client_name_ar: Optional[str] = None
    address: Optional[str] = None
    address_ar: Optional[str] = None
    registration_date: Optional[str] = None
    suspended_on: Optional[str] = None
    reason_for_suspension: Optional[str] = None
    reason_for_suspension_ar: Optional[str] = None
    future_action: Optional[str] = None
    remarks: Optional[str] = None
    status: Optional[str] = None
    lifted_on: Optional[str] = None
    lifted_reason: Optional[str] = None

class SuspendedClient(BaseModel):
    """Suspended Client Registry Record (BAC-F6-20)"""
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    serial_number: int = 0  # Auto-incremented Sr. No.
    # Client Information
    client_id: str = ""  # Client reference ID
    client_name: str = ""
    client_name_ar: str = ""
    address: str = ""
    address_ar: str = ""
    # Suspension Details
    registration_date: str = ""  # Original certification date YYYY-MM-DD
    suspended_on: str = ""  # Suspension date YYYY-MM-DD
    reason_for_suspension: str = ""
    reason_for_suspension_ar: str = ""
    future_action: str = ""  # reinstate, withdraw, extend_suspension, under_review
    remarks: str = ""
    # Status tracking
    status: str = "suspended"  # suspended, reinstated, withdrawn
    lifted_on: str = ""  # Date suspension was lifted YYYY-MM-DD
    lifted_reason: str = ""  # Reason for lifting suspension
    # Link to certified client record
    linked_certified_client_id: str = ""
    # Timestamps
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: Optional[datetime] = None

# ================= AUDITOR MODELS =================

class AuditorAvailability(BaseModel):
    date: str  # YYYY-MM-DD
    is_available: bool = True
    reason: str = ""  # Reason if unavailable (vacation, sick, training, etc.)

class Auditor(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    employee_id: str = ""  # Employee ID/Badge number
    name: str
    name_ar: str = ""
    email: str
    phone: str = ""
    mobile: str = ""
    # Qualifications
    specializations: List[str] = []  # ISO 9001, ISO 14001, ISO 45001, etc.
    certification_level: str = "auditor"  # trainee, auditor, lead_auditor, technical_expert
    years_experience: int = 0
    certifications: List[str] = []  # List of certifications held
    # Status
    status: str = "active"  # active, inactive, on_leave
    # Availability
    availability: List[AuditorAvailability] = []
    default_available: bool = True  # Default availability if no specific date set
    max_audits_per_month: int = 10
    # Assignments tracking
    current_assignments: int = 0
    total_audits_completed: int = 0
    # Notes
    notes: str = ""
    # Timestamps
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: Optional[datetime] = None

class AuditorCreate(BaseModel):
    employee_id: str = ""
    name: str
    name_ar: str = ""
    email: str
    phone: str = ""
    mobile: str = ""
    specializations: List[str] = []
    certification_level: str = "auditor"
    years_experience: int = 0
    certifications: List[str] = []
    max_audits_per_month: int = 10
    notes: str = ""

class AuditAssignment(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    audit_id: str
    auditor_id: str
    role: str = "auditor"  # lead_auditor, auditor, technical_expert, observer
    status: str = "assigned"  # assigned, confirmed, completed, cancelled
    assigned_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    confirmed_at: Optional[datetime] = None
    notes: str = ""

# ================= TEMPLATE MODELS =================

class CertificationPackage(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    name_ar: str
    description: str
    description_ar: str
    standards: List[str]
    is_active: bool = True
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ProposalTemplate(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    name_ar: str
    description: str = ""
    default_fees: Dict[str, float] = {}  # initial_certification, surveillance_1, surveillance_2, recertification
    default_notes: str = ""
    default_validity_days: int = 30
    is_active: bool = True
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# Legacy models (kept for backward compatibility)
class QuotationCreate(BaseModel):
    form_id: str
    client_id: str
    client_email: EmailStr
    price: float
    details: str

class Quotation(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    form_id: str
    client_id: str
    price: float
    details: str
    status: str = "pending"  # pending, approved, rejected
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class QuotationResponse(BaseModel):
    status: str  # approved or rejected

class Contract(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    quotation_id: str
    proposal_id: str = ""  # Link to new proposal system
    client_id: str
    pdf_path: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# ================= APPLICATION FORM MODELS =================

class CurrentCertification(BaseModel):
    system: str = ""
    body: str = ""
    validUntil: str = ""

class ApplicationFormData(BaseModel):
    # Step 1: Company Information
    dateOfApplication: str = ""
    companyName: str = ""
    address: str = ""
    phoneNumber: str = ""
    website: str = ""
    email: str = ""
    contactPerson: str = ""
    designation: str = ""
    mobileNumber: str = ""
    contactEmail: str = ""
    legalStatus: str = ""
    
    # Step 2: Certification Selection
    certificationSchemes: List[str] = []
    certificationProgram: str = ""
    combinedAudit: str = ""
    combinedAuditSpecification: str = ""
    isInternalAuditCombined: str = ""
    isMRMCombined: str = ""
    isManualProceduresCombined: str = ""
    isSystemIntegrated: str = ""
    
    # Step 3: Sites & Employees
    numberOfSites: int = 1
    site1Address: str = ""
    site2Address: str = ""
    totalEmployees: str = ""
    locationShifts: str = ""
    fullTimeEmployees: str = ""
    partTimeEmployees: str = ""
    temporaryEmployees: str = ""
    unskilledWorkers: str = ""
    remoteEmployees: str = ""
    
    # Step 4: Existing Certifications
    isAlreadyCertified: str = ""
    currentCertifications: List[CurrentCertification] = []
    isConsultantInvolved: str = ""
    consultantName: str = ""
    transferReason: str = ""
    currentCertificateExpiry: str = ""
    keyBusinessProcesses: str = ""
    
    # Step 5: Management System Requirements
    # EMS
    hasEnvironmentAspectRegister: str = ""
    hasEnvironmentalManual: str = ""
    hasEnvironmentalAuditProgram: str = ""
    
    # FSMS
    numberOfHACCPStudies: str = ""
    numberOfProcessLines: str = ""
    processingType: str = ""
    
    # OHSMS
    hazardsIdentified: str = ""
    criticalRisks: str = ""
    
    # EnMS
    annualEnergyConsumption: str = ""
    numberOfEnergySources: str = ""
    numberOfSEUs: str = ""
    
    # Medical Devices
    productsInRange: str = ""
    medicalDeviceTypes: List[str] = []
    sterilizationType: str = ""
    numberOfDeviceFiles: str = ""
    applicableLegislations: str = ""
    exportCountries: str = ""
    productStandards: str = ""
    intendedUse: str = ""
    outsourceProcesses: str = ""
    
    # ISMS
    businessComplexity: str = ""
    processStandard: str = ""
    managementSystemLevel: str = ""
    itEnvironmentComplexity: str = ""
    outsourcingDependency: str = ""
    systemDevelopment: str = ""
    
    # Step 6: Declaration
    declarationName: str = ""
    declarationDesignation: str = ""
    declarationAgreed: bool = False

class ClientInfo(BaseModel):
    name: str
    company_name: str
    email: EmailStr
    phone: str
    mobile: str = ""  # Optional mobile number for easier customer communication

class AuditCalculationResult(BaseModel):
    certifications: Dict[str, Any] = {}
    total_md: float = 0
    reduction: float = 0
    final_total_md: float = 0
    phases: Dict[str, float] = {}

class ApplicationFormCreate(BaseModel):
    client_info: ClientInfo

class ApplicationForm(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    access_token: str = Field(default_factory=lambda: str(uuid.uuid4()))  # Public access token
    client_info: ClientInfo
    company_data: Optional[ApplicationFormData] = None
    audit_calculation: Optional[AuditCalculationResult] = None  # Auto-calculated on submit
    status: str = "pending"  # pending, submitted, under_review, approved, rejected
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    submitted_at: Optional[datetime] = None

class ApplicationFormUpdate(BaseModel):
    company_data: ApplicationFormData

class PublicApplicationFormResponse(BaseModel):
    id: str
    client_info: ClientInfo
    company_data: Optional[ApplicationFormData] = None
    status: str

# ================= HELPER FUNCTIONS =================

def hash_password(password: str) -> str:
    return pwd_context.hash(password)

def verify_password(plain_password: str, hashed_password: str) -> bool:
    return pwd_context.verify(plain_password, hashed_password)

def create_jwt_token(user_id: str, role: str) -> str:
    payload = {
        "user_id": user_id,
        "role": role,
        "exp": datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRATION_HOURS)
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

def decode_jwt_token(token: str) -> dict:
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        return payload
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token has expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

async def send_email(to: str, subject: str, body: str):
    """Send email using SMTP (mock implementation - logs to console)"""
    # For production, configure SMTP settings
    smtp_host = os.environ.get('SMTP_HOST', '')
    smtp_port = int(os.environ.get('SMTP_PORT', 587))
    smtp_user = os.environ.get('SMTP_USER', '')
    smtp_pass = os.environ.get('SMTP_PASS', '')
    
    if smtp_host and smtp_user and smtp_pass:
        # Real email sending
        message = MIMEMultipart()
        message["From"] = smtp_user
        message["To"] = to
        message["Subject"] = subject
        message.attach(MIMEText(body, "plain", "utf-8"))
        
        try:
            await aiosmtplib.send(
                message,
                hostname=smtp_host,
                port=smtp_port,
                username=smtp_user,
                password=smtp_pass,
                start_tls=True
            )
            logging.info(f"Email sent successfully to {to}")
        except Exception as e:
            logging.error(f"Failed to send email: {e}")
            raise HTTPException(status_code=500, detail=f"Failed to send email: {str(e)}")
    else:
        # Mock email - just log it
        logging.info(f"[MOCK EMAIL] To: {to}, Subject: {subject}")
        logging.info(f"[MOCK EMAIL] Body: {body[:200]}...")

def calculate_audit_from_form_data(company_data: ApplicationFormData) -> Dict:
    """
    Calculate audit time based on form data.
    Extracts employees, certifications, and risk info from the form.
    """
    # Get number of employees
    try:
        employees = int(company_data.totalEmployees) if company_data.totalEmployees else 10
    except (ValueError, TypeError):
        employees = 10  # Default
    
    # Get certifications selected
    certifications = company_data.certificationSchemes or []
    
    # Map form certification values to calculator format
    cert_mapping = {
        "ISO9001": "ISO9001",
        "ISO14001": "ISO14001",
        "ISO45001": "ISO45001",
        "ISO22000": "ISO22000",
        "ISO13485": "ISO13485",
        "ISO22301": "ISO22301",
        "ISO27001": "ISO27001",
        "ISO50001": "ISO50001",
    }
    
    mapped_certs = [cert_mapping.get(c, c) for c in certifications if c in cert_mapping]
    
    if not mapped_certs:
        return {
            "certifications": {},
            "total_md": 0,
            "reduction": 0,
            "final_total_md": 0,
            "phases": {}
        }
    
    # Get risk category (default to medium)
    risk_category = "medium"
    
    # Get food safety category if ISO 22000 is selected
    food_safety_category = company_data.processingType if company_data.processingType else "AI"
    
    # Check if integrated with ISO 9001
    integrated_with_9001 = "ISO9001" in mapped_certs
    
    # Get HACCP studies count
    try:
        haccp_studies = int(company_data.numberOfHACCPStudies) if company_data.numberOfHACCPStudies else 1
    except (ValueError, TypeError):
        haccp_studies = 1
    
    # Calculate
    result = calculate_total_audit_time(
        certifications=mapped_certs,
        employees=employees,
        risk_category=risk_category,
        food_safety_category=food_safety_category,
        haccp_studies=haccp_studies,
        integrated_with_9001=integrated_with_9001
    )
    
    return result

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)) -> dict:
    token = credentials.credentials
    payload = decode_jwt_token(token)
    return payload

async def require_admin(current_user: dict = Depends(get_current_user)) -> dict:
    if current_user.get("role") != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Admin access required")
    return current_user

async def send_email(to: str, subject: str, body: str):
    """Send email using SMTP"""
    try:
        smtp_host = os.environ.get('SMTP_HOST')
        smtp_port = int(os.environ.get('SMTP_PORT', 587))
        smtp_user = os.environ.get('SMTP_USER')
        smtp_pass = os.environ.get('SMTP_PASS')
        
        if not all([smtp_host, smtp_user, smtp_pass]):
            logging.warning("SMTP not configured, email not sent")
            return
        
        message = MIMEMultipart()
        message['From'] = smtp_user
        message['To'] = to
        message['Subject'] = subject
        message.attach(MIMEText(body, 'plain'))
        
        await aiosmtplib.send(
            message,
            hostname=smtp_host,
            port=smtp_port,
            username=smtp_user,
            password=smtp_pass,
            start_tls=True
        )
        logging.info(f"Email sent to {to}")
    except Exception as e:
        logging.error(f"Failed to send email: {e}")

def generate_legacy_contract_pdf(quotation: Quotation, user: dict) -> str:
    """Generate PDF contract (legacy - for old quotation system)"""
    filename = f"contract_{quotation.id}_{int(datetime.now().timestamp())}.pdf"
    filepath = CONTRACTS_DIR / filename
    
    c = canvas.Canvas(str(filepath), pagesize=letter)
    width, height = letter
    
    # Header
    c.setFont("Helvetica-Bold", 24)
    c.drawString(1 * inch, height - 1 * inch, "SERVICE CONTRACT")
    
    # Contract details
    c.setFont("Helvetica", 12)
    y = height - 2 * inch
    
    c.drawString(1 * inch, y, f"Contract ID: {quotation.id}")
    y -= 0.3 * inch
    
    c.drawString(1 * inch, y, f"Date: {datetime.now().strftime('%Y-%m-%d')}")
    y -= 0.3 * inch
    
    c.drawString(1 * inch, y, f"Client ID: {quotation.client_id}")
    y -= 0.5 * inch
    
    # Price
    c.setFont("Helvetica-Bold", 14)
    c.drawString(1 * inch, y, f"Total Price: ${quotation.price:,.2f}")
    y -= 0.5 * inch
    
    # Details
    c.setFont("Helvetica", 12)
    c.drawString(1 * inch, y, "Service Details:")
    y -= 0.3 * inch
    
    # Wrap text for details
    detail_lines = quotation.details.split('\n')
    for line in detail_lines[:10]:  # Limit to 10 lines
        if y < 2 * inch:
            break
        c.drawString(1.2 * inch, y, line[:80])  # Limit line length
        y -= 0.25 * inch
    
    # Footer
    c.setFont("Helvetica", 10)
    c.drawString(1 * inch, 1 * inch, "This contract is legally binding upon approval by both parties.")
    
    c.save()
    return str(filepath)

# ================= ROUTES =================

@api_router.get("/")
async def root():
    return {"message": "Service Contract Management API"}

# Authentication routes are now in routes/auth.py

# Form routes
@api_router.get("/defaults/signature")
async def get_default_signature():
    """Get the default Bayan signature as base64"""
    signature_path = ROOT_DIR / "assets" / "bayan-signature.png"
    if signature_path.exists():
        import base64
        with open(signature_path, "rb") as f:
            data = base64.b64encode(f.read()).decode('utf-8')
            return {"signature": f"data:image/png;base64,{data}"}
    return {"signature": ""}

@api_router.get("/defaults/stamp")
async def get_default_stamp():
    """Get the default Bayan stamp as base64"""
    stamp_path = ROOT_DIR / "assets" / "bayan-stamp.png"
    if stamp_path.exists():
        import base64
        with open(stamp_path, "rb") as f:
            data = base64.b64encode(f.read()).decode('utf-8')
            return {"stamp": f"data:image/png;base64,{data}"}
    return {"stamp": ""}

@api_router.get("/defaults/signatory")
async def get_default_signatory():
    """Get all default signatory details (name, title, signature, stamp)"""
    import base64
    result = {
        "issuer_name": "Abdullah Al-Rashid",
        "issuer_designation": "General Manager",
        "issuer_signature": "",
        "issuer_stamp": ""
    }
    
    signature_path = ROOT_DIR / "assets" / "bayan-signature.png"
    if signature_path.exists():
        with open(signature_path, "rb") as f:
            data = base64.b64encode(f.read()).decode('utf-8')
            result["issuer_signature"] = f"data:image/png;base64,{data}"
    
    stamp_path = ROOT_DIR / "assets" / "bayan-stamp.png"
    if stamp_path.exists():
        with open(stamp_path, "rb") as f:
            data = base64.b64encode(f.read()).decode('utf-8')
            result["issuer_stamp"] = f"data:image/png;base64,{data}"
    
    return result

@api_router.post("/forms", response_model=Form)
async def create_form(form_data: FormCreate, current_user: dict = Depends(require_admin)):
    form = Form(
        client_id=form_data.client_id,
        fields=form_data.fields
    )
    
    form_doc = form.model_dump()
    form_doc['created_at'] = form_doc['created_at'].isoformat()
    
    await db.forms.insert_one(form_doc)
    return form

@api_router.get("/forms", response_model=List[Form])
async def get_forms(current_user: dict = Depends(get_current_user)):
    query = {}
    # If client, only show their forms
    if current_user['role'] == UserRole.CLIENT:
        query = {"client_id": current_user['user_id']}
    
    forms = await db.forms.find(query, {"_id": 0}).to_list(1000)
    
    for form in forms:
        if isinstance(form['created_at'], str):
            form['created_at'] = datetime.fromisoformat(form['created_at'])
    
    return forms

@api_router.get("/forms/{form_id}", response_model=Form)
async def get_form(form_id: str, current_user: dict = Depends(get_current_user)):
    form = await db.forms.find_one({"id": form_id}, {"_id": 0})
    if not form:
        raise HTTPException(status_code=404, detail="Form not found")
    
    # Check access
    if current_user['role'] == UserRole.CLIENT and form['client_id'] != current_user['user_id']:
        raise HTTPException(status_code=403, detail="Access denied")
    
    if isinstance(form['created_at'], str):
        form['created_at'] = datetime.fromisoformat(form['created_at'])
    
    return Form(**form)

@api_router.post("/forms/{form_id}/submit", response_model=Form)
async def submit_form(form_id: str, submission: FormSubmission, current_user: dict = Depends(get_current_user)):
    form = await db.forms.find_one({"id": form_id})
    if not form:
        raise HTTPException(status_code=404, detail="Form not found")
    
    # Check access
    if current_user['role'] == UserRole.CLIENT and form['client_id'] != current_user['user_id']:
        raise HTTPException(status_code=403, detail="Access denied")
    
    # Update form
    await db.forms.update_one(
        {"id": form_id},
        {"$set": {"responses": submission.responses, "status": "submitted"}}
    )
    
    updated_form = await db.forms.find_one({"id": form_id}, {"_id": 0})
    if isinstance(updated_form['created_at'], str):
        updated_form['created_at'] = datetime.fromisoformat(updated_form['created_at'])
    
    return Form(**updated_form)

# Quotation routes
@api_router.post("/quotations", response_model=Quotation)
async def create_quotation(quotation_data: QuotationCreate, current_user: dict = Depends(require_admin)):
    quotation = Quotation(
        form_id=quotation_data.form_id,
        client_id=quotation_data.client_id,
        price=quotation_data.price,
        details=quotation_data.details
    )
    
    quotation_doc = quotation.model_dump()
    quotation_doc['created_at'] = quotation_doc['created_at'].isoformat()
    
    await db.quotations.insert_one(quotation_doc)
    
    # Send email notification
    await send_email(
        quotation_data.client_email,
        "New Quotation Received",
        f"You have received a new quotation for ${quotation_data.price}.\n\nDetails:\n{quotation_data.details}\n\nPlease log in to your account to review and respond."
    )
    
    return quotation

@api_router.get("/quotations", response_model=List[Quotation])
async def get_quotations(current_user: dict = Depends(get_current_user)):
    query = {}
    # If client, only show their quotations
    if current_user['role'] == UserRole.CLIENT:
        query = {"client_id": current_user['user_id']}
    
    quotations = await db.quotations.find(query, {"_id": 0}).to_list(1000)
    
    for quotation in quotations:
        if isinstance(quotation['created_at'], str):
            quotation['created_at'] = datetime.fromisoformat(quotation['created_at'])
    
    return quotations

@api_router.get("/quotations/{quotation_id}", response_model=Quotation)
async def get_quotation(quotation_id: str, current_user: dict = Depends(get_current_user)):
    quotation = await db.quotations.find_one({"id": quotation_id}, {"_id": 0})
    if not quotation:
        raise HTTPException(status_code=404, detail="Quotation not found")
    
    # Check access
    if current_user['role'] == UserRole.CLIENT and quotation['client_id'] != current_user['user_id']:
        raise HTTPException(status_code=403, detail="Access denied")
    
    if isinstance(quotation['created_at'], str):
        quotation['created_at'] = datetime.fromisoformat(quotation['created_at'])
    
    return Quotation(**quotation)

@api_router.post("/quotations/{quotation_id}/respond")
async def respond_to_quotation(quotation_id: str, response: QuotationResponse, current_user: dict = Depends(get_current_user)):
    quotation_doc = await db.quotations.find_one({"id": quotation_id})
    if not quotation_doc:
        raise HTTPException(status_code=404, detail="Quotation not found")
    
    # Check access
    if current_user['role'] == UserRole.CLIENT and quotation_doc['client_id'] != current_user['user_id']:
        raise HTTPException(status_code=403, detail="Access denied")
    
    # Validate status
    if response.status not in ["approved", "rejected", "modifications_requested"]:
        raise HTTPException(status_code=400, detail="Invalid status")
    
    # Update quotation
    await db.quotations.update_one(
        {"id": quotation_id},
        {"$set": {"status": response.status}}
    )
    
    # If approved, generate contract
    if response.status == "approved":
        quotation_obj = Quotation(**{
            **quotation_doc,
            'created_at': datetime.fromisoformat(quotation_doc['created_at']) if isinstance(quotation_doc['created_at'], str) else quotation_doc['created_at']
        })
        
        pdf_path = generate_contract_pdf(quotation_obj, current_user)
        
        contract = Contract(
            quotation_id=quotation_id,
            client_id=quotation_doc['client_id'],
            pdf_path=pdf_path
        )
        
        contract_doc = contract.model_dump()
        contract_doc['created_at'] = contract_doc['created_at'].isoformat()
        
        await db.contracts.insert_one(contract_doc)
        
        return {
            "message": "Quotation approved and contract generated",
            "contract": contract
        }
    
    return {"message": f"Quotation {response.status}"}

# Certification Agreement routes
@api_router.get("/certification-agreements")
async def get_certification_agreements(status: str = None, current_user: dict = Depends(get_current_user)):
    """Get all certification agreements (signed contracts)"""
    query = {}
    if status and status != 'all':
        query['status'] = status
    
    agreements = await db.certification_agreements.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return agreements

# Contract routes
@api_router.get("/contracts", response_model=List[Contract])
async def get_contracts(current_user: dict = Depends(get_current_user)):
    query = {}
    # If client, only show their contracts
    if current_user['role'] == UserRole.CLIENT:
        query = {"client_id": current_user['user_id']}
    
    # Sort by most recent first
    contracts = await db.contracts.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    
    for contract in contracts:
        if isinstance(contract['created_at'], str):
            contract['created_at'] = datetime.fromisoformat(contract['created_at'])
    
    return contracts

@api_router.get("/contracts/{contract_id}/download")
async def download_contract(contract_id: str, current_user: dict = Depends(get_current_user)):
    contract = await db.contracts.find_one({"id": contract_id})
    if not contract:
        raise HTTPException(status_code=404, detail="Contract not found")
    
    # Check access
    if current_user['role'] == UserRole.CLIENT and contract['client_id'] != current_user['user_id']:
        raise HTTPException(status_code=403, detail="Access denied")
    
    pdf_path = Path(contract['pdf_path'])
    if not pdf_path.exists():
        raise HTTPException(status_code=404, detail="Contract file not found")
    
    return FileResponse(
        path=pdf_path,
        filename=pdf_path.name,
        media_type='application/pdf'
    )

# ================= USER ROUTES =================

@api_router.get("/users/clients")
async def get_clients(current_user: dict = Depends(require_admin)):
    """Get all client users (admin only)"""
    clients = await db.users.find({"role": UserRole.CLIENT}, {"_id": 0, "password": 0}).to_list(1000)
    return clients

# ================= APPLICATION FORM ROUTES =================

@api_router.post("/application-forms", response_model=ApplicationForm)
async def create_application_form(form_data: ApplicationFormCreate, current_user: dict = Depends(require_admin)):
    """Admin creates an application form for a client (no login required for client)"""
    form = ApplicationForm(
        client_info=form_data.client_info
    )
    
    form_doc = form.model_dump()
    form_doc['created_at'] = form_doc['created_at'].isoformat()
    if form_doc.get('submitted_at'):
        form_doc['submitted_at'] = form_doc['submitted_at'].isoformat()
    
    await db.application_forms.insert_one(form_doc)
    return form

@api_router.get("/application-forms", response_model=List[ApplicationForm])
async def get_application_forms(current_user: dict = Depends(require_admin)):
    """Get all application forms (admin only) - sorted by most recent first"""
    # Only return forms that have client_info (new format), sorted by created_at descending
    forms = await db.application_forms.find({"client_info": {"$exists": True}}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    
    for form in forms:
        if isinstance(form.get('created_at'), str):
            form['created_at'] = datetime.fromisoformat(form['created_at'])
        if form.get('submitted_at') and isinstance(form['submitted_at'], str):
            form['submitted_at'] = datetime.fromisoformat(form['submitted_at'])
    
    return forms

@api_router.get("/application-forms/{form_id}", response_model=ApplicationForm)
async def get_application_form(form_id: str, current_user: dict = Depends(require_admin)):
    """Get a specific application form (admin only)"""
    form = await db.application_forms.find_one({"id": form_id}, {"_id": 0})
    if not form:
        raise HTTPException(status_code=404, detail="Application form not found")
    
    if isinstance(form.get('created_at'), str):
        form['created_at'] = datetime.fromisoformat(form['created_at'])
    if form.get('submitted_at') and isinstance(form['submitted_at'], str):
        form['submitted_at'] = datetime.fromisoformat(form['submitted_at'])
    
    return ApplicationForm(**form)

@api_router.post("/application-forms/{form_id}/send-email")
async def send_form_email(form_id: str, current_user: dict = Depends(require_admin)):
    """Send the form link to the client via email"""
    form = await db.application_forms.find_one({"id": form_id})
    if not form:
        raise HTTPException(status_code=404, detail="Application form not found")
    
    # Get frontend URL from environment or use default
    frontend_url = os.environ.get('FRONTEND_URL', 'https://audit-workflow-pro-1.preview.emergentagent.com')
    form_link = f"{frontend_url}/form/{form['access_token']}"
    
    client_info = form['client_info']
    email_body = f"""
مرحباً {client_info['name']}،

تم إنشاء نموذج طلب اعتماد لشركتكم {client_info['company_name']}.

يرجى النقر على الرابط التالي لملء النموذج:
{form_link}

مع أطيب التحيات،
بيان للتحقق والمطابقة

---

Hello {client_info['name']},

A certification application form has been created for your company {client_info['company_name']}.

Please click the following link to fill out the form:
{form_link}

Best regards,
Bayan Auditing & Conformity
"""
    
    await send_email(
        to=client_info['email'],
        subject="طلب اعتماد - Certification Application Form",
        body=email_body
    )
    
    return {"message": "Email sent successfully", "form_link": form_link}

# ================= PUBLIC FORM ACCESS (NO LOGIN REQUIRED) =================

@api_router.get("/public/form/{access_token}", response_model=PublicApplicationFormResponse)
async def get_public_form(access_token: str):
    """Public access to form using access token (no login required)"""
    form = await db.application_forms.find_one({"access_token": access_token}, {"_id": 0})
    if not form:
        raise HTTPException(status_code=404, detail="Form not found or invalid link")
    
    return PublicApplicationFormResponse(
        id=form['id'],
        client_info=form['client_info'],
        company_data=form.get('company_data'),
        status=form['status']
    )

@api_router.put("/public/form/{access_token}", response_model=PublicApplicationFormResponse)
async def update_public_form(access_token: str, update_data: ApplicationFormUpdate):
    """Public save draft (no login required)"""
    form = await db.application_forms.find_one({"access_token": access_token})
    if not form:
        raise HTTPException(status_code=404, detail="Form not found or invalid link")
    
    # Only allow updates if form is pending
    if form['status'] != 'pending':
        raise HTTPException(status_code=400, detail="Cannot update a submitted form")
    
    # Update form data
    await db.application_forms.update_one(
        {"access_token": access_token},
        {"$set": {"company_data": update_data.company_data.model_dump()}}
    )
    
    updated_form = await db.application_forms.find_one({"access_token": access_token}, {"_id": 0})
    
    return PublicApplicationFormResponse(
        id=updated_form['id'],
        client_info=updated_form['client_info'],
        company_data=updated_form.get('company_data'),
        status=updated_form['status']
    )

@api_router.post("/public/form/{access_token}/submit", response_model=PublicApplicationFormResponse)
async def submit_public_form(access_token: str, update_data: ApplicationFormUpdate):
    """Public form submission (no login required)"""
    form = await db.application_forms.find_one({"access_token": access_token})
    if not form:
        raise HTTPException(status_code=404, detail="Form not found or invalid link")
    
    # Only allow submission if form is pending
    if form['status'] != 'pending':
        raise HTTPException(status_code=400, detail="Form has already been submitted")
    
    # Validate required fields
    company_data = update_data.company_data
    if not company_data.companyName or not company_data.declarationAgreed:
        raise HTTPException(status_code=400, detail="Company name and declaration agreement are required")
    
    # Calculate audit time automatically
    audit_calculation = calculate_audit_from_form_data(company_data)
    
    # Update and submit
    submitted_at = datetime.now(timezone.utc)
    await db.application_forms.update_one(
        {"access_token": access_token},
        {
            "$set": {
                "company_data": company_data.model_dump(),
                "audit_calculation": audit_calculation,
                "status": "submitted",
                "submitted_at": submitted_at.isoformat()
            }
        }
    )
    
    updated_form = await db.application_forms.find_one({"access_token": access_token}, {"_id": 0})
    
    # Create notification for admin
    client_name = form.get('client_info', {}).get('company_name', 'Unknown')
    await create_notification(
        notification_type="form_submitted",
        title="طلب اعتماد جديد",
        message=f"قامت {client_name} بتقديم طلب اعتماد جديد",
        related_id=updated_form['id'],
        related_type="form"
    )
    
    return PublicApplicationFormResponse(
        id=updated_form['id'],
        client_info=updated_form['client_info'],
        company_data=updated_form.get('company_data'),
        status=updated_form['status']
    )

# ================= PROPOSAL ROUTES =================

@api_router.post("/proposals", response_model=Proposal)
async def create_proposal(proposal_data: ProposalCreate, current_user: dict = Depends(require_admin)):
    """Admin creates a proposal/quotation for a client"""
    # Calculate total amount
    fees = proposal_data.service_fees
    total = fees.initial_certification + fees.surveillance_1 + fees.surveillance_2 + fees.recertification
    
    proposal = Proposal(
        application_form_id=proposal_data.application_form_id,
        organization_name=proposal_data.organization_name,
        organization_address=proposal_data.organization_address,
        organization_phone=proposal_data.organization_phone,
        contact_person=proposal_data.contact_person,
        contact_position=proposal_data.contact_position,
        contact_email=proposal_data.contact_email,
        standards=proposal_data.standards,
        scope=proposal_data.scope,
        total_employees=proposal_data.total_employees,
        number_of_sites=proposal_data.number_of_sites,
        audit_duration=proposal_data.audit_duration,
        service_fees=proposal_data.service_fees,
        total_amount=total,
        notes=proposal_data.notes,
        validity_days=proposal_data.validity_days,
        issuer_name=proposal_data.issuer_name,  # Use value from form
        issuer_designation=proposal_data.issuer_designation,  # Use value from form
        issuer_signature=proposal_data.issuer_signature,  # First party signature
        issuer_stamp=proposal_data.issuer_stamp,  # First party stamp
        issued_date=datetime.now(timezone.utc)
    )
    
    proposal_doc = proposal.model_dump()
    proposal_doc['created_at'] = proposal_doc['created_at'].isoformat()
    proposal_doc['issued_date'] = proposal_doc['issued_date'].isoformat() if proposal_doc['issued_date'] else None
    
    await db.proposals.insert_one(proposal_doc)
    
    # Update application form status to under_review
    await db.application_forms.update_one(
        {"id": proposal_data.application_form_id},
        {"$set": {"status": "under_review"}}
    )
    
    return proposal

@api_router.get("/proposals", response_model=List[Proposal])
async def get_proposals(current_user: dict = Depends(require_admin)):
    """Get all proposals (admin only) - sorted by most recent first"""
    proposals = await db.proposals.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    
    for p in proposals:
        if isinstance(p.get('created_at'), str):
            p['created_at'] = datetime.fromisoformat(p['created_at'])
        if p.get('issued_date') and isinstance(p['issued_date'], str):
            p['issued_date'] = datetime.fromisoformat(p['issued_date'])
        if p.get('client_response_date') and isinstance(p['client_response_date'], str):
            p['client_response_date'] = datetime.fromisoformat(p['client_response_date'])
    
    return proposals

@api_router.get("/proposals/{proposal_id}", response_model=Proposal)
async def get_proposal(proposal_id: str, current_user: dict = Depends(require_admin)):
    """Get a specific proposal"""
    proposal = await db.proposals.find_one({"id": proposal_id}, {"_id": 0})
    if not proposal:
        raise HTTPException(status_code=404, detail="Proposal not found")
    
    if isinstance(proposal.get('created_at'), str):
        proposal['created_at'] = datetime.fromisoformat(proposal['created_at'])
    if proposal.get('issued_date') and isinstance(proposal['issued_date'], str):
        proposal['issued_date'] = datetime.fromisoformat(proposal['issued_date'])
    
    return Proposal(**proposal)

@api_router.post("/proposals/{proposal_id}/send")
async def send_proposal(proposal_id: str, current_user: dict = Depends(require_admin)):
    """Send proposal link to client via email"""
    proposal = await db.proposals.find_one({"id": proposal_id})
    if not proposal:
        raise HTTPException(status_code=404, detail="Proposal not found")
    
    frontend_url = os.environ.get('FRONTEND_URL', 'https://audit-workflow-pro-1.preview.emergentagent.com')
    proposal_link = f"{frontend_url}/proposal/{proposal['access_token']}"
    
    email_body = f"""
مرحباً {proposal['contact_person']}،

نرفق لكم عرض السعر للحصول على شهادة الاعتماد لشركتكم {proposal['organization_name']}.

يرجى النقر على الرابط التالي لمراجعة العرض والموافقة عليه:
{proposal_link}

العرض صالح لمدة {proposal['validity_days']} يوماً.

مع أطيب التحيات،
بيان للتحقق والمطابقة

---

Hello {proposal['contact_person']},

Please find attached the price quotation for certification of your organization {proposal['organization_name']}.

Click the following link to review and accept the proposal:
{proposal_link}

This proposal is valid for {proposal['validity_days']} days.

Best regards,
Bayan Auditing & Conformity
"""
    
    await send_email(
        to=proposal['contact_email'],
        subject=f"عرض سعر - Price Quotation - {proposal['organization_name']}",
        body=email_body
    )
    
    return {"message": "Proposal sent successfully", "proposal_link": proposal_link}

# ================= PUBLIC PROPOSAL ACCESS =================

@api_router.get("/public/proposal/{access_token}", response_model=PublicProposalResponse)
async def get_public_proposal(access_token: str):
    """Public access to proposal (no login required)"""
    proposal = await db.proposals.find_one({"access_token": access_token}, {"_id": 0})
    if not proposal:
        raise HTTPException(status_code=404, detail="Proposal not found or invalid link")
    
    return PublicProposalResponse(
        id=proposal['id'],
        organization_name=proposal['organization_name'],
        organization_address=proposal.get('organization_address', ''),
        contact_person=proposal['contact_person'],
        contact_email=proposal['contact_email'],
        standards=proposal['standards'],
        scope=proposal['scope'],
        total_employees=proposal['total_employees'],
        number_of_sites=proposal['number_of_sites'],
        audit_duration=proposal['audit_duration'],
        service_fees=proposal['service_fees'],
        total_amount=proposal['total_amount'],
        notes=proposal.get('notes', ''),
        validity_days=proposal['validity_days'],
        status=proposal['status'],
        issuer_name=proposal.get('issuer_name', ''),
        issuer_designation=proposal.get('issuer_designation', ''),
        issued_date=datetime.fromisoformat(proposal['issued_date']) if proposal.get('issued_date') else None,
        modification_comment=proposal.get('modification_comment', ''),
        modification_requested_changes=proposal.get('modification_requested_changes', '')
    )

@api_router.post("/public/proposal/{access_token}/respond")
async def respond_to_proposal(access_token: str, response: ProposalResponse):
    """Client accepts or rejects proposal (no login required)"""
    proposal = await db.proposals.find_one({"access_token": access_token})
    if not proposal:
        raise HTTPException(status_code=404, detail="Proposal not found or invalid link")
    
    if proposal['status'] != 'pending':
        raise HTTPException(status_code=400, detail="Proposal has already been responded to")
    
    update_data = {
        "status": response.status,
        "client_response_date": datetime.now(timezone.utc).isoformat(),
        "client_signatory_name": response.signatory_name,
        "client_signatory_designation": response.signatory_designation,
    }
    
    if response.status == "rejected":
        update_data["rejection_reason"] = response.rejection_reason
    
    await db.proposals.update_one(
        {"access_token": access_token},
        {"$set": update_data}
    )
    
    # Create notification for admin
    org_name = proposal.get('organization_name', 'Unknown')
    if response.status == "accepted":
        await create_notification(
            notification_type="proposal_accepted",
            title="تم قبول العرض",
            message=f"قامت {org_name} بقبول عرض السعر",
            related_id=proposal['id'],
            related_type="proposal"
        )
    else:
        await create_notification(
            notification_type="proposal_rejected",
            title="تم رفض العرض",
            message=f"قامت {org_name} برفض عرض السعر",
            related_id=proposal['id'],
            related_type="proposal"
        )
    
    # If accepted, update application form status and create contract
    if response.status == "accepted":
        await db.application_forms.update_one(
            {"id": proposal['application_form_id']},
            {"$set": {"status": "approved"}}
        )
        
        # TODO: Generate contract PDF
        contract = Contract(
            quotation_id="",
            proposal_id=proposal['id'],
            client_id=proposal['organization_name'],
            pdf_path=""  # Will be generated
        )
        contract_doc = contract.model_dump()
        contract_doc['created_at'] = contract_doc['created_at'].isoformat()
        await db.contracts.insert_one(contract_doc)
    
    return {"message": f"Proposal {response.status} successfully"}

# ================= PUBLIC PROPOSAL PDF DOWNLOAD =================

@api_router.get("/public/proposals/{access_token}/bilingual_pdf")
async def get_public_proposal_bilingual_pdf(access_token: str):
    """Public access to download bilingual proposal PDF (no login required)"""
    proposal = await db.proposals.find_one({"access_token": access_token}, {"_id": 0})
    if not proposal:
        raise HTTPException(status_code=404, detail="Proposal not found or invalid link")
    
    # Generate the PDF
    pdf_path = await generate_bilingual_proposal_pdf_file(proposal)
    
    return FileResponse(
        pdf_path,
        media_type='application/pdf',
        filename=f'proposal_{proposal["id"][:8]}_ar_en.pdf'
    )

# ================= CERTIFICATION AGREEMENT ROUTES =================

@api_router.post("/public/agreement/{access_token}/submit")
async def submit_certification_agreement(access_token: str, agreement_data: CertificationAgreementSubmit):
    """Submit certification agreement form (no login required)"""
    # Find the proposal by access token
    proposal = await db.proposals.find_one({"access_token": access_token})
    if not proposal:
        raise HTTPException(status_code=404, detail="Proposal not found or invalid link")
    
    # Check if proposal is accepted
    if proposal['status'] != 'accepted':
        raise HTTPException(status_code=400, detail="Proposal must be accepted before signing agreement")
    
    # Check if agreement already exists
    existing_agreement = await db.certification_agreements.find_one({"proposal_access_token": access_token})
    if existing_agreement:
        raise HTTPException(status_code=400, detail="Agreement has already been submitted")
    
    # Validate signature is provided
    if not agreement_data.signature_image:
        raise HTTPException(status_code=422, detail="Digital signature is required")
    
    # Validate company seal is provided
    if not agreement_data.stamp_image:
        raise HTTPException(status_code=422, detail="Company seal/stamp is required")
    
    # Create certification agreement
    agreement = CertificationAgreement(
        proposal_id=proposal['id'],
        proposal_access_token=access_token,
        organization_name=agreement_data.organization_name,
        organization_address=agreement_data.organization_address,
        selected_standards=agreement_data.selected_standards,
        other_standard=agreement_data.other_standard,
        scope_of_services=agreement_data.scope_of_services,
        sites=agreement_data.sites,
        signatory_name=agreement_data.signatory_name,
        signatory_position=agreement_data.signatory_position,
        signatory_date=agreement_data.signatory_date,
        acknowledgements=agreement_data.acknowledgements.model_dump(),
        signature_image=agreement_data.signature_image,
        stamp_image=agreement_data.stamp_image
    )
    
    agreement_doc = agreement.model_dump()
    agreement_doc['created_at'] = agreement_doc['created_at'].isoformat()
    
    await db.certification_agreements.insert_one(agreement_doc)
    
    # Update proposal status to indicate agreement signed
    await db.proposals.update_one(
        {"access_token": access_token},
        {"$set": {"status": "agreement_signed"}}
    )
    
    # Update application form status
    await db.application_forms.update_one(
        {"id": proposal['application_form_id']},
        {"$set": {"status": "agreement_signed"}}
    )
    
    # Create notification for admin
    await create_notification(
        notification_type="agreement_signed",
        title="اتفاقية جديدة موقعة",
        message=f"قامت {agreement_data.organization_name} بتوقيع اتفاقية المنح",
        related_id=agreement.id,
        related_type="agreement"
    )
    
    # ===== AUTO-SCHEDULE AUDITS =====
    # Automatically create audit schedules based on the proposal's audit duration
    try:
        audit_duration = proposal.get('audit_duration', {})
        organization_name = proposal.get('organization_name', '')
        sites = agreement_data.sites or ['Main Site']
        
        # Calculate audit dates starting from today
        from dateutil.relativedelta import relativedelta
        base_date = datetime.now(timezone.utc)
        
        # Define the audit schedule (audit type, days from contract signing, duration field)
        audit_schedule_plan = [
            ("stage_1", 30, "stage_1", "Initial Audit - Stage 1 | التدقيق الأولي - المرحلة 1"),
            ("stage_2", 60, "stage_2", "Initial Audit - Stage 2 | التدقيق الأولي - المرحلة 2"),
            ("surveillance_1", 365, "surveillance_1", "Surveillance Audit 1 | تدقيق المراقبة 1"),
            ("surveillance_2", 730, "surveillance_2", "Surveillance Audit 2 | تدقيق المراقبة 2"),
            ("recertification", 1065, "recertification", "Recertification Audit | تدقيق إعادة الاعتماد"),
        ]
        
        created_audits = []
        for audit_type, days_offset, duration_field, audit_notes in audit_schedule_plan:
            duration_days = audit_duration.get(duration_field, 1)
            if duration_days and duration_days > 0:
                scheduled_date = base_date + timedelta(days=days_offset)
                
                # Create audit for each site
                for site_idx, site in enumerate(sites):
                    site_name = site if isinstance(site, str) else site.get('name', f'Site {site_idx + 1}')
                    
                    audit_entry = {
                        "id": str(uuid.uuid4()),
                        "contract_id": proposal['id'],
                        "site_id": str(site_idx),
                        "organization_name": organization_name,
                        "site_name": site_name,
                        "audit_type": audit_type,
                        "scheduled_date": scheduled_date.strftime("%Y-%m-%d"),
                        "scheduled_time": "09:00",
                        "duration_days": int(duration_days),
                        "auditors": "",
                        "notes": audit_notes,
                        "status": "scheduled",
                        "is_recurring": False,
                        "recurrence_type": "",
                        "recurrence_end_date": "",
                        "parent_audit_id": "",
                        "calendar_event_id": "",
                        "calendar_synced": False,
                        "sms_reminder_sent": False,
                        "created_at": datetime.now(timezone.utc).isoformat(),
                        "auto_generated": True  # Mark as auto-generated
                    }
                    
                    await db.audit_schedules.insert_one(audit_entry)
                    created_audits.append(audit_entry['id'])
        
        print(f"Auto-scheduled {len(created_audits)} audits for contract {proposal['id']}")
        
        # Create notification for auto-scheduled audits
        if created_audits:
            await create_notification(
                notification_type="audits_scheduled",
                title="تم جدولة التدقيقات تلقائياً",
                message=f"تم إنشاء {len(created_audits)} جدولة تدقيق تلقائياً لـ {organization_name}",
                related_id=agreement.id,
                related_type="audit_schedule"
            )
    except Exception as e:
        print(f"Warning: Failed to auto-schedule audits: {e}")
        # Don't fail the agreement signing if audit scheduling fails
    
    return {"message": "Certification agreement submitted successfully", "agreement_id": agreement.id}

@api_router.get("/public/agreement/{access_token}")
async def get_certification_agreement(access_token: str):
    """Get certification agreement status"""
    agreement = await db.certification_agreements.find_one({"proposal_access_token": access_token}, {"_id": 0})
    if agreement:
        return {"status": "submitted", "agreement": agreement}
    
    # Check if proposal exists and is accepted
    proposal = await db.proposals.find_one({"access_token": access_token}, {"_id": 0})
    if not proposal:
        raise HTTPException(status_code=404, detail="Proposal not found")
    
    if proposal['status'] != 'accepted':
        raise HTTPException(status_code=400, detail="Proposal is not accepted")
    
    return {"status": "pending", "proposal": proposal}

# ================= PDF CONTRACT GENERATION =================

@api_router.get("/contracts/{agreement_id}/pdf")
async def generate_contract_pdf_endpoint(agreement_id: str, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Generate PDF Grant Agreement for a signed agreement (Admin only)"""
    await get_current_user(credentials)
    
    # Get agreement
    agreement = await db.certification_agreements.find_one({"id": agreement_id}, {"_id": 0})
    if not agreement:
        raise HTTPException(status_code=404, detail="Agreement not found")
    
    # Get proposal for issuer details
    proposal = await db.proposals.find_one({"id": agreement['proposal_id']}, {"_id": 0})
    if not proposal:
        raise HTTPException(status_code=404, detail="Proposal not found")
    
    # Prepare agreement data with correct field mappings
    agreement_data = {
        'organization_name': agreement.get('organization_name', ''),
        'organization_address': agreement.get('organization_address', ''),
        'selected_standards': agreement.get('selected_standards', []),
        'standards': agreement.get('selected_standards', []),  # alias
        'scope': agreement.get('scope_of_services', ''),
        'scope_of_services': agreement.get('scope_of_services', ''),  # alias
        'sites': agreement.get('sites', []),
        'signatory_name': agreement.get('signatory_name', ''),
        'signatory_position': agreement.get('signatory_position', ''),
        'signatory_date': agreement.get('signatory_date', ''),
        'signature_image': agreement.get('signature_image', ''),
        'stamp_image': agreement.get('stamp_image', ''),
        # BAC signatory from proposal
        'issuer_name': proposal.get('issuer_name', 'Abdullah Al-Rashid'),
        'issuer_designation': proposal.get('issuer_designation', 'General Manager'),
    }
    
    # Generate PDF
    try:
        pdf_path = CONTRACTS_DIR / f"grant_agreement_{agreement_id[:8]}.pdf"
        
        generate_grant_agreement_pdf(agreement_data, str(pdf_path))
        
        # Read PDF bytes
        with open(pdf_path, 'rb') as f:
            pdf_bytes = f.read()
        
        # Update agreement status
        await db.certification_agreements.update_one(
            {"id": agreement_id},
            {"$set": {"status": "contract_generated"}}
        )
        
        # Return PDF
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={
                "Content-Disposition": f"attachment; filename=grant_agreement_{agreement_id[:8]}.pdf"
            }
        )
    except Exception as e:
        logging.error(f"Error generating Grant Agreement PDF: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error generating PDF: {str(e)}")

@api_router.get("/public/contracts/{access_token}/pdf")
async def get_public_contract_pdf(access_token: str):
    """Get PDF Grant Agreement for a signed agreement (Public - client access)
    Generates a professional bilingual PDF with all terms and conditions"""
    # Get agreement by access token
    agreement = await db.certification_agreements.find_one({"proposal_access_token": access_token}, {"_id": 0})
    if not agreement:
        raise HTTPException(status_code=404, detail="Agreement not found")
    
    # Get proposal for issuer details
    proposal = await db.proposals.find_one({"id": agreement['proposal_id']}, {"_id": 0})
    if not proposal:
        raise HTTPException(status_code=404, detail="Proposal not found")
    
    # Prepare agreement data
    agreement_data = {
        'organization_name': agreement.get('organization_name', ''),
        'organization_address': agreement.get('organization_address', ''),
        'selected_standards': agreement.get('selected_standards', []),
        'scope': agreement.get('scope_of_services', ''),
        'sites': agreement.get('sites', []),
        'signatory_name': agreement.get('signatory_name', ''),
        'signatory_position': agreement.get('signatory_position', ''),
        'signatory_date': agreement.get('signatory_date', ''),
        'issuer_name': proposal.get('issuer_name', 'Abdullah Al-Rashid'),
        'issuer_designation': proposal.get('issuer_designation', 'General Manager'),
    }
    
    # Generate PDF
    try:
        pdf_bytes = generate_grant_agreement_pdf(agreement_data)
        
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={
                "Content-Disposition": f"attachment; filename=grant_agreement_{agreement['id'][:8]}.pdf"
            }
        )
    except Exception as e:
        logging.error(f"Error generating Grant Agreement PDF: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error generating PDF: {str(e)}")

# ================= BILINGUAL PDF CONTRACT GENERATION =================

@api_router.get("/contracts/{agreement_id}/pdf/bilingual")
async def generate_bilingual_contract_pdf_endpoint(agreement_id: str, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Generate PDF Grant Agreement for a signed agreement (Admin only)
    Generates a professional bilingual PDF with all terms and conditions"""
    await get_current_user(credentials)
    
    # Get agreement
    agreement = await db.certification_agreements.find_one({"id": agreement_id}, {"_id": 0})
    if not agreement:
        raise HTTPException(status_code=404, detail="Agreement not found")
    
    # Get proposal for issuer details
    proposal = await db.proposals.find_one({"id": agreement['proposal_id']}, {"_id": 0})
    if not proposal:
        raise HTTPException(status_code=404, detail="Proposal not found")
    
    # Prepare agreement data
    agreement_data = {
        'organization_name': agreement.get('organization_name', ''),
        'organization_address': agreement.get('organization_address', ''),
        'selected_standards': agreement.get('selected_standards', []),
        'scope': agreement.get('scope_of_services', ''),
        'sites': agreement.get('sites', []),
        'signatory_name': agreement.get('signatory_name', ''),
        'signatory_position': agreement.get('signatory_position', ''),
        'signatory_date': agreement.get('signatory_date', ''),
        'issuer_name': proposal.get('issuer_name', 'Abdullah Al-Rashid'),
        'issuer_designation': proposal.get('issuer_designation', 'General Manager'),
    }
    
    # Generate PDF
    try:
        pdf_bytes = generate_grant_agreement_pdf(agreement_data)
        
        # Update agreement status
        await db.certification_agreements.update_one(
            {"id": agreement_id},
            {"$set": {"status": "contract_generated"}}
        )
        
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={
                "Content-Disposition": f"attachment; filename=grant_agreement_{agreement_id[:8]}.pdf"
            }
        )
    except Exception as e:
        logging.error(f"Error generating Grant Agreement PDF: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error generating PDF: {str(e)}")

@api_router.get("/public/contracts/{access_token}/pdf/bilingual")
async def get_public_bilingual_contract_pdf(access_token: str):
    """Get PDF Grant Agreement for a signed agreement (Public - client access)
    Generates a professional bilingual PDF with all terms and conditions"""
    # Get agreement by access token
    agreement = await db.certification_agreements.find_one({"proposal_access_token": access_token}, {"_id": 0})
    if not agreement:
        raise HTTPException(status_code=404, detail="Agreement not found")
    
    # Get proposal for issuer details
    proposal = await db.proposals.find_one({"id": agreement['proposal_id']}, {"_id": 0})
    if not proposal:
        raise HTTPException(status_code=404, detail="Proposal not found")
    
    # Prepare agreement data
    agreement_data = {
        'organization_name': agreement.get('organization_name', ''),
        'organization_address': agreement.get('organization_address', ''),
        'selected_standards': agreement.get('selected_standards', []),
        'scope': agreement.get('scope_of_services', ''),
        'sites': agreement.get('sites', []),
        'signatory_name': agreement.get('signatory_name', ''),
        'signatory_position': agreement.get('signatory_position', ''),
        'signatory_date': agreement.get('signatory_date', ''),
        'issuer_name': proposal.get('issuer_name', 'Abdullah Al-Rashid'),
        'issuer_designation': proposal.get('issuer_designation', 'General Manager'),
    }
    
    # Generate PDF
    try:
        pdf_bytes = generate_grant_agreement_pdf(agreement_data)
        
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={
                "Content-Disposition": f"attachment; filename=grant_agreement_{agreement['id'][:8]}.pdf"
            }
        )
    except Exception as e:
        logging.error(f"Error generating Grant Agreement PDF: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error generating PDF: {str(e)}")

# ================= NOTIFICATION HELPER =================
# Note: Notification routes are now in routes/notifications.py

async def create_notification(notification_type: str, title: str, message: str, related_id: str = None, related_type: str = None):
    """Helper function to create a notification"""
    notification = Notification(
        type=notification_type,
        title=title,
        message=message,
        related_id=related_id,
        related_type=related_type
    )
    notification_doc = notification.model_dump()
    notification_doc['created_at'] = notification_doc['created_at'].isoformat()
    await db.notifications.insert_one(notification_doc)
    return notification

# ================= TEMPLATE ROUTES =================

@api_router.get("/templates/packages")
async def get_certification_packages(credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Get all certification packages"""
    await get_current_user(credentials)
    packages = await db.certification_packages.find({"is_active": True}, {"_id": 0}).to_list(100)
    return packages

@api_router.post("/templates/packages")
async def create_certification_package(package: CertificationPackage, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Create a new certification package"""
    await get_current_user(credentials)
    package_doc = package.model_dump()
    package_doc['created_at'] = package_doc['created_at'].isoformat()
    await db.certification_packages.insert_one(package_doc)
    return {"message": "Package created", "id": package.id}

@api_router.delete("/templates/packages/{package_id}")
async def delete_certification_package(package_id: str, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Delete a certification package"""
    await get_current_user(credentials)
    await db.certification_packages.update_one(
        {"id": package_id},
        {"$set": {"is_active": False}}
    )
    return {"message": "Package deleted"}

@api_router.put("/templates/packages/{package_id}")
async def update_certification_package(package_id: str, package: CertificationPackage, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Update a certification package"""
    await get_current_user(credentials)
    update_data = {
        "name": package.name,
        "name_ar": package.name_ar,
        "description": package.description,
        "description_ar": package.description_ar,
        "standards": package.standards
    }
    await db.certification_packages.update_one(
        {"id": package_id},
        {"$set": update_data}
    )
    return {"message": "Package updated", "id": package_id}

@api_router.get("/templates/proposals")
async def get_proposal_templates(credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Get all proposal templates"""
    await get_current_user(credentials)
    templates = await db.proposal_templates.find({"is_active": True}, {"_id": 0}).to_list(100)
    return templates

@api_router.post("/templates/proposals")
async def create_proposal_template(template: ProposalTemplate, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Create a new proposal template"""
    await get_current_user(credentials)
    template_doc = template.model_dump()
    template_doc['created_at'] = template_doc['created_at'].isoformat()
    await db.proposal_templates.insert_one(template_doc)
    return {"message": "Template created", "id": template.id}

@api_router.delete("/templates/proposals/{template_id}")
async def delete_proposal_template(template_id: str, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Delete a proposal template"""
    await get_current_user(credentials)
    await db.proposal_templates.update_one(
        {"id": template_id},
        {"$set": {"is_active": False}}
    )
    return {"message": "Template deleted"}

@api_router.put("/templates/proposals/{template_id}")
async def update_proposal_template(template_id: str, template: ProposalTemplate, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Update a proposal template"""
    await get_current_user(credentials)
    update_data = {
        "name": template.name,
        "name_ar": template.name_ar,
        "description": template.description,
        "default_fees": template.default_fees,
        "default_notes": template.default_notes,
        "default_validity_days": template.default_validity_days
    }
    await db.proposal_templates.update_one(
        {"id": template_id},
        {"$set": update_data}
    )
    return {"message": "Template updated", "id": template_id}

# ================= REPORTS ROUTES =================

@api_router.get("/reports/submissions")
async def get_submission_statistics(credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Get submission statistics"""
    await get_current_user(credentials)
    
    # Get counts by status
    total_forms = await db.application_forms.count_documents({})
    submitted_forms = await db.application_forms.count_documents({"status": {"$in": ["submitted", "under_review", "approved", "agreement_signed"]}})
    pending_forms = await db.application_forms.count_documents({"status": "pending"})
    
    # Get monthly statistics (last 6 months)
    from datetime import datetime, timedelta
    monthly_stats = []
    for i in range(5, -1, -1):
        start_date = datetime.now().replace(day=1) - timedelta(days=i*30)
        end_date = start_date + timedelta(days=30)
        
        count = await db.application_forms.count_documents({
            "created_at": {
                "$gte": start_date.isoformat(),
                "$lt": end_date.isoformat()
            }
        })
        monthly_stats.append({
            "month": start_date.strftime("%b %Y"),
            "count": count
        })
    
    # Conversion rate
    total_proposals = await db.proposals.count_documents({})
    accepted_proposals = await db.proposals.count_documents({"status": {"$in": ["accepted", "agreement_signed"]}})
    conversion_rate = (accepted_proposals / total_proposals * 100) if total_proposals > 0 else 0
    
    return {
        "total_forms": total_forms,
        "submitted_forms": submitted_forms,
        "pending_forms": pending_forms,
        "monthly_stats": monthly_stats,
        "total_proposals": total_proposals,
        "accepted_proposals": accepted_proposals,
        "conversion_rate": round(conversion_rate, 1)
    }

@api_router.get("/reports/revenue")
async def get_revenue_statistics(credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Get revenue statistics"""
    await get_current_user(credentials)
    
    # Get all proposals
    proposals = await db.proposals.find({}, {"_id": 0}).to_list(1000)
    
    total_quoted = sum(p.get('total_amount', 0) for p in proposals)
    accepted_revenue = sum(p.get('total_amount', 0) for p in proposals if p.get('status') in ['accepted', 'agreement_signed'])
    pending_revenue = sum(p.get('total_amount', 0) for p in proposals if p.get('status') == 'sent')
    rejected_revenue = sum(p.get('total_amount', 0) for p in proposals if p.get('status') == 'rejected')
    
    # Get agreements (contracts)
    agreements = await db.certification_agreements.count_documents({})
    
    # Revenue by month (last 6 months)
    from datetime import datetime, timedelta
    monthly_revenue = []
    for i in range(5, -1, -1):
        start_date = datetime.now().replace(day=1) - timedelta(days=i*30)
        end_date = start_date + timedelta(days=30)
        
        month_proposals = [p for p in proposals if p.get('status') in ['accepted', 'agreement_signed']]
        # Filter by date if issued_date exists
        month_total = sum(
            p.get('total_amount', 0) for p in month_proposals
            if p.get('issued_date') and start_date.isoformat() <= p.get('issued_date', '') < end_date.isoformat()
        )
        
        monthly_revenue.append({
            "month": start_date.strftime("%b %Y"),
            "amount": month_total
        })
    
    return {
        "total_quoted": total_quoted,
        "accepted_revenue": accepted_revenue,
        "pending_revenue": pending_revenue,
        "rejected_revenue": rejected_revenue,
        "total_contracts": agreements,
        "monthly_revenue": monthly_revenue,
        "currency": "SAR"
    }

@api_router.get("/reports/filtered")
async def get_filtered_report(
    credentials: HTTPAuthorizationCredentials = Depends(security),
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    status: Optional[str] = None,
    standard: Optional[str] = None
):
    """Get filtered report data with comprehensive filters"""
    await get_current_user(credentials)
    
    # Build query for forms
    forms_query = {}
    if status and status != "all":
        forms_query["status"] = status
    
    # Date filtering
    if start_date or end_date:
        forms_query["created_at"] = {}
        if start_date:
            forms_query["created_at"]["$gte"] = start_date
        if end_date:
            forms_query["created_at"]["$lte"] = end_date
    
    # Get filtered forms
    forms = await db.application_forms.find(forms_query, {"_id": 0}).to_list(1000)
    
    # Filter by standard if specified
    if standard and standard != "all":
        forms = [f for f in forms if f.get('company_data') and f.get('company_data', {}).get('certificationSchemes') and standard in f.get('company_data', {}).get('certificationSchemes', [])]
    
    # Build query for proposals
    proposals_query = {}
    if status and status != "all":
        proposals_query["status"] = status
    if start_date or end_date:
        proposals_query["created_at"] = {}
        if start_date:
            proposals_query["created_at"]["$gte"] = start_date
        if end_date:
            proposals_query["created_at"]["$lte"] = end_date
    
    # Get filtered proposals
    proposals = await db.proposals.find(proposals_query, {"_id": 0}).to_list(1000)
    
    # Filter proposals by standard if specified
    if standard and standard != "all":
        proposals = [p for p in proposals if standard in p.get('standards', [])]
    
    # Calculate statistics
    total_forms = len(forms)
    submitted_forms = len([f for f in forms if f.get('status') in ['submitted', 'under_review', 'approved', 'agreement_signed']])
    pending_forms = len([f for f in forms if f.get('status') == 'pending'])
    
    total_proposals = len(proposals)
    accepted_proposals = len([p for p in proposals if p.get('status') in ['accepted', 'agreement_signed']])
    pending_proposals = len([p for p in proposals if p.get('status') == 'pending'])
    rejected_proposals = len([p for p in proposals if p.get('status') == 'rejected'])
    modification_requested = len([p for p in proposals if p.get('status') == 'modification_requested'])
    
    # Revenue calculations
    total_quoted = sum(p.get('total_amount', 0) for p in proposals)
    accepted_revenue = sum(p.get('total_amount', 0) for p in proposals if p.get('status') in ['accepted', 'agreement_signed'])
    pending_revenue = sum(p.get('total_amount', 0) for p in proposals if p.get('status') == 'pending')
    rejected_revenue = sum(p.get('total_amount', 0) for p in proposals if p.get('status') == 'rejected')
    
    conversion_rate = (accepted_proposals / total_proposals * 100) if total_proposals > 0 else 0
    
    # Standards breakdown
    standards_count = {}
    for p in proposals:
        for std in p.get('standards', []):
            standards_count[std] = standards_count.get(std, 0) + 1
    
    # Status breakdown for forms
    form_status_breakdown = {}
    for f in forms:
        status_val = f.get('status', 'unknown')
        form_status_breakdown[status_val] = form_status_breakdown.get(status_val, 0) + 1
    
    # Status breakdown for proposals
    proposal_status_breakdown = {}
    for p in proposals:
        status_val = p.get('status', 'unknown')
        proposal_status_breakdown[status_val] = proposal_status_breakdown.get(status_val, 0) + 1
    
    return {
        "forms": {
            "total": total_forms,
            "submitted": submitted_forms,
            "pending": pending_forms,
            "status_breakdown": form_status_breakdown,
            "data": forms[:50]  # Return first 50 for detail view
        },
        "proposals": {
            "total": total_proposals,
            "accepted": accepted_proposals,
            "pending": pending_proposals,
            "rejected": rejected_proposals,
            "modification_requested": modification_requested,
            "status_breakdown": proposal_status_breakdown,
            "data": proposals[:50]
        },
        "revenue": {
            "total_quoted": total_quoted,
            "accepted": accepted_revenue,
            "pending": pending_revenue,
            "rejected": rejected_revenue,
            "currency": "SAR"
        },
        "conversion_rate": round(conversion_rate, 1),
        "standards_breakdown": standards_count,
        "filters_applied": {
            "start_date": start_date,
            "end_date": end_date,
            "status": status,
            "standard": standard
        }
    }

# ================= MODIFICATION REQUEST ROUTES =================

@api_router.post("/public/proposal/{access_token}/request_modification")
async def request_proposal_modification(access_token: str, request: ModificationRequest):
    """Client requests modification to a proposal (no login required)"""
    proposal = await db.proposals.find_one({"access_token": access_token})
    if not proposal:
        raise HTTPException(status_code=404, detail="Proposal not found or invalid link")
    
    if proposal['status'] != 'pending':
        raise HTTPException(status_code=400, detail="Proposal has already been responded to")
    
    if not request.comment:
        raise HTTPException(status_code=400, detail="Please provide a comment explaining the requested modifications")
    
    # Update proposal with modification request
    update_data = {
        "status": "modification_requested",
        "modification_comment": request.comment,
        "modification_requested_changes": request.requested_changes,
        "modification_requested_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.proposals.update_one(
        {"access_token": access_token},
        {"$set": update_data}
    )
    
    # Create notification for admin
    org_name = proposal.get('organization_name', 'Unknown')
    await create_notification(
        notification_type="modification_requested",
        title="طلب تعديل على العرض",
        message=f"طلبت {org_name} تعديلات على عرض السعر: {request.comment[:100]}",
        related_id=proposal['id'],
        related_type="proposal"
    )
    
    return {"message": "Modification request submitted successfully"}

# ================= CUSTOMER PORTAL / TRACKING ROUTES =================

@api_router.get("/public/track/{tracking_id}")
async def track_order(tracking_id: str):
    """Public endpoint to track order status by tracking ID (form ID, access token, or proposal access token)"""
    
    # Try to find by form ID first
    form = await db.application_forms.find_one({"id": tracking_id}, {"_id": 0})
    
    # If not found, try by form access token
    if not form:
        form = await db.application_forms.find_one({"access_token": tracking_id}, {"_id": 0})
    
    # If still not found, try by proposal access token
    proposal_by_token = None
    if not form:
        proposal_by_token = await db.proposals.find_one({"access_token": tracking_id}, {"_id": 0})
        if proposal_by_token:
            # Get the form linked to this proposal
            form = await db.application_forms.find_one({"id": proposal_by_token['application_form_id']}, {"_id": 0})
    
    if not form:
        raise HTTPException(status_code=404, detail="Order not found")
    
    # Get proposal if exists
    proposal = await db.proposals.find_one({"application_form_id": form['id']}, {"_id": 0})
    
    # Get agreement if exists
    agreement = None
    if proposal:
        agreement = await db.certification_agreements.find_one({"proposal_id": proposal['id']}, {"_id": 0})
    
    # Determine current status
    current_status = form.get('status', 'pending')
    if agreement:
        current_status = 'agreement_signed'
    elif proposal and proposal.get('status') in ['accepted', 'agreement_signed']:
        current_status = 'accepted'
    elif proposal:
        current_status = 'under_review'
    
    # Build timeline dates
    timeline_dates = {
        'pending': form.get('created_at')
    }
    if form.get('submitted_at'):
        timeline_dates['submitted'] = form.get('submitted_at')
    if proposal:
        timeline_dates['under_review'] = proposal.get('created_at')
        if proposal.get('status') in ['accepted', 'agreement_signed']:
            timeline_dates['accepted'] = proposal.get('client_response_date')
    if agreement:
        timeline_dates['agreement_signed'] = agreement.get('created_at')
        if agreement.get('status') == 'contract_generated':
            timeline_dates['contract_generated'] = agreement.get('created_at')
    
    # Get standards from form data
    standards = []
    if form.get('company_data') and form['company_data'].get('certificationSchemes'):
        standards = form['company_data']['certificationSchemes']
    elif proposal:
        standards = proposal.get('standards', [])
    
    response = {
        "tracking_id": form['id'],
        "company_name": form.get('client_info', {}).get('company_name', 'Unknown'),
        "contact_email": form.get('client_info', {}).get('email', ''),
        "contact_phone": form.get('client_info', {}).get('phone', ''),
        "created_at": form.get('created_at'),
        "submitted_at": form.get('submitted_at'),
        "current_status": current_status,
        "standards": standards,
        "timeline_dates": timeline_dates,
        "contract_available": agreement is not None and agreement.get('signature_image')
    }
    
    # Add proposal info if available
    if proposal:
        response["proposal"] = {
            "id": proposal['id'],
            "access_token": proposal.get('access_token'),
            "total_amount": proposal.get('total_amount', 0),
            "status": proposal.get('status', 'pending')
        }
    
    return response

# ================= SITE MANAGEMENT =================
# Note: Site routes are now in routes/sites.py

class Site(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    contract_id: str  # Links to proposal/contract
    name: str
    address: str
    city: str = ""
    country: str = "Saudi Arabia"
    contact_name: str = ""
    contact_email: str = ""
    contact_phone: str = ""
    is_main_site: bool = False
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class SiteCreate(BaseModel):
    contract_id: str
    name: str
    address: str
    city: str = ""
    country: str = "Saudi Arabia"
    contact_name: str = ""
    contact_email: str = ""
    contact_phone: str = ""
    is_main_site: bool = False

# ================= AUDIT SCHEDULING ROUTES =================

class AuditSchedule(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    contract_id: str  # Links to proposal/contract
    site_id: str = ""  # Optional - for multi-site
    organization_name: str = ""
    site_name: str = ""
    audit_type: str = "initial"  # initial, surveillance, recertification
    scheduled_date: str
    scheduled_time: str = "09:00"
    duration_days: int = 1
    auditors: str = ""
    notes: str = ""
    status: str = "scheduled"  # scheduled, in_progress, completed, cancelled
    # Recurring event fields
    is_recurring: bool = False
    recurrence_type: str = ""  # weekly, monthly, quarterly, yearly
    recurrence_end_date: str = ""
    parent_audit_id: str = ""  # For generated recurring events
    # Google Calendar sync
    calendar_event_id: str = ""
    calendar_synced: bool = False
    # SMS notification
    sms_reminder_sent: bool = False
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class AuditScheduleCreate(BaseModel):
    contract_id: str
    site_id: str = ""
    audit_type: str = "initial"
    scheduled_date: str
    scheduled_time: str = "09:00"
    duration_days: int = 1
    auditors: str = ""
    notes: str = ""
    # Recurring event fields
    is_recurring: bool = False
    recurrence_type: str = ""  # weekly, monthly, quarterly, yearly
    recurrence_end_date: str = ""

# ================= AUDITOR MANAGEMENT ROUTES =================

@api_router.get("/auditors")
async def get_auditors(
    credentials: HTTPAuthorizationCredentials = Depends(security),
    status: Optional[str] = None,
    specialization: Optional[str] = None
):
    """Get all auditors with optional filtering"""
    await get_current_user(credentials)
    
    query = {}
    if status:
        query["status"] = status
    if specialization:
        query["specializations"] = specialization
    
    auditors = await db.auditors.find(query, {"_id": 0}).sort("name", 1).to_list(100)
    
    # Calculate current assignments for each auditor
    for auditor in auditors:
        assignments = await db.audit_assignments.count_documents({
            "auditor_id": auditor['id'],
            "status": {"$in": ["assigned", "confirmed"]}
        })
        auditor['current_assignments'] = assignments
    
    return auditors

@api_router.get("/auditors/available")
async def get_available_auditors(
    date: str,
    specialization: Optional[str] = None,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Get auditors available on a specific date"""
    await get_current_user(credentials)
    
    query = {"status": "active"}
    if specialization:
        query["specializations"] = specialization
    
    auditors = await db.auditors.find(query, {"_id": 0}).to_list(100)
    
    available_auditors = []
    for auditor in auditors:
        # Check if auditor has specific unavailability for this date
        is_available = auditor.get('default_available', True)
        for avail in auditor.get('availability', []):
            if avail.get('date') == date:
                is_available = avail.get('is_available', True)
                break
        
        if is_available:
            # Check if auditor is already assigned to an audit on this date
            existing = await db.audit_schedules.count_documents({
                "scheduled_date": date,
                "auditors": {"$regex": auditor['id']}
            })
            
            # Check max audits per month
            month_start = date[:7] + "-01"
            month_end = date[:7] + "-31"
            monthly_count = await db.audit_assignments.count_documents({
                "auditor_id": auditor['id'],
                "status": {"$in": ["assigned", "confirmed", "completed"]}
            })
            
            auditor['is_available'] = existing == 0 and monthly_count < auditor.get('max_audits_per_month', 10)
            auditor['existing_audits_today'] = existing
            auditor['monthly_audits'] = monthly_count
            available_auditors.append(auditor)
    
    return available_auditors

@api_router.post("/auditors")
async def create_auditor(auditor_data: AuditorCreate, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Create a new auditor"""
    await get_current_user(credentials)
    
    # Check if email already exists
    existing = await db.auditors.find_one({"email": auditor_data.email})
    if existing:
        raise HTTPException(status_code=400, detail="Auditor with this email already exists")
    
    auditor = Auditor(**auditor_data.model_dump())
    auditor_doc = auditor.model_dump()
    auditor_doc['created_at'] = auditor_doc['created_at'].isoformat()
    
    await db.auditors.insert_one(auditor_doc)
    
    return {"message": "Auditor created", "auditor_id": auditor.id}

@api_router.get("/auditors/{auditor_id}")
async def get_auditor(auditor_id: str, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Get auditor details"""
    await get_current_user(credentials)
    
    auditor = await db.auditors.find_one({"id": auditor_id}, {"_id": 0})
    if not auditor:
        raise HTTPException(status_code=404, detail="Auditor not found")
    
    # Get recent assignments
    assignments = await db.audit_assignments.find(
        {"auditor_id": auditor_id},
        {"_id": 0}
    ).sort("assigned_at", -1).to_list(20)
    
    auditor['recent_assignments'] = assignments
    
    return auditor

@api_router.put("/auditors/{auditor_id}")
async def update_auditor(auditor_id: str, updates: dict, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Update auditor details"""
    await get_current_user(credentials)
    
    auditor = await db.auditors.find_one({"id": auditor_id})
    if not auditor:
        raise HTTPException(status_code=404, detail="Auditor not found")
    
    updates['updated_at'] = datetime.now(timezone.utc).isoformat()
    await db.auditors.update_one({"id": auditor_id}, {"$set": updates})
    
    return {"message": "Auditor updated"}

@api_router.delete("/auditors/{auditor_id}")
async def delete_auditor(auditor_id: str, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Delete an auditor (soft delete - sets status to inactive)"""
    await get_current_user(credentials)
    
    auditor = await db.auditors.find_one({"id": auditor_id})
    if not auditor:
        raise HTTPException(status_code=404, detail="Auditor not found")
    
    # Check for active assignments
    active_assignments = await db.audit_assignments.count_documents({
        "auditor_id": auditor_id,
        "status": {"$in": ["assigned", "confirmed"]}
    })
    
    if active_assignments > 0:
        raise HTTPException(status_code=400, detail=f"Cannot delete auditor with {active_assignments} active assignments")
    
    # Soft delete
    await db.auditors.update_one(
        {"id": auditor_id},
        {"$set": {"status": "inactive", "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    return {"message": "Auditor deactivated"}

@api_router.post("/auditors/{auditor_id}/availability")
async def set_auditor_availability(
    auditor_id: str,
    availability_data: dict,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Set auditor availability for specific dates"""
    await get_current_user(credentials)
    
    auditor = await db.auditors.find_one({"id": auditor_id})
    if not auditor:
        raise HTTPException(status_code=404, detail="Auditor not found")
    
    # Update availability array
    current_availability = auditor.get('availability', [])
    date = availability_data.get('date')
    
    # Remove existing entry for this date
    current_availability = [a for a in current_availability if a.get('date') != date]
    
    # Add new entry
    current_availability.append({
        "date": date,
        "is_available": availability_data.get('is_available', True),
        "reason": availability_data.get('reason', '')
    })
    
    await db.auditors.update_one(
        {"id": auditor_id},
        {"$set": {
            "availability": current_availability,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"message": "Availability updated"}

# ================= AUDIT ASSIGNMENT ROUTES =================

@api_router.post("/audit-schedules/{audit_id}/assign")
async def assign_auditor_to_audit(
    audit_id: str,
    assignment_data: dict,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Assign an auditor to an audit"""
    await get_current_user(credentials)
    
    # Verify audit exists
    audit = await db.audit_schedules.find_one({"id": audit_id})
    if not audit:
        raise HTTPException(status_code=404, detail="Audit not found")
    
    # Verify auditor exists
    auditor_id = assignment_data.get('auditor_id')
    auditor = await db.auditors.find_one({"id": auditor_id})
    if not auditor:
        raise HTTPException(status_code=404, detail="Auditor not found")
    
    # Check if already assigned
    existing = await db.audit_assignments.find_one({
        "audit_id": audit_id,
        "auditor_id": auditor_id,
        "status": {"$ne": "cancelled"}
    })
    if existing:
        raise HTTPException(status_code=400, detail="Auditor already assigned to this audit")
    
    # Create assignment
    assignment = AuditAssignment(
        audit_id=audit_id,
        auditor_id=auditor_id,
        role=assignment_data.get('role', 'auditor'),
        notes=assignment_data.get('notes', '')
    )
    
    assignment_doc = assignment.model_dump()
    assignment_doc['assigned_at'] = assignment_doc['assigned_at'].isoformat()
    await db.audit_assignments.insert_one(assignment_doc)
    
    # Update audit with auditor name
    current_auditors = audit.get('auditors', '')
    auditor_name = auditor.get('name', '')
    if current_auditors:
        current_auditors += f", {auditor_name}"
    else:
        current_auditors = auditor_name
    
    await db.audit_schedules.update_one(
        {"id": audit_id},
        {"$set": {"auditors": current_auditors}}
    )
    
    return {"message": "Auditor assigned", "assignment_id": assignment.id}

@api_router.get("/audit-schedules/{audit_id}/assignments")
async def get_audit_assignments(audit_id: str, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Get all auditor assignments for an audit"""
    await get_current_user(credentials)
    
    assignments = await db.audit_assignments.find(
        {"audit_id": audit_id, "status": {"$ne": "cancelled"}},
        {"_id": 0}
    ).to_list(20)
    
    # Enrich with auditor details
    for assignment in assignments:
        auditor = await db.auditors.find_one({"id": assignment['auditor_id']}, {"_id": 0})
        if auditor:
            assignment['auditor'] = {
                "id": auditor['id'],
                "name": auditor['name'],
                "name_ar": auditor.get('name_ar', ''),
                "email": auditor['email'],
                "certification_level": auditor.get('certification_level', ''),
                "specializations": auditor.get('specializations', [])
            }
    
    return assignments

@api_router.delete("/audit-assignments/{assignment_id}")
async def remove_auditor_assignment(assignment_id: str, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Remove an auditor from an audit"""
    await get_current_user(credentials)
    
    assignment = await db.audit_assignments.find_one({"id": assignment_id})
    if not assignment:
        raise HTTPException(status_code=404, detail="Assignment not found")
    
    # Cancel assignment
    await db.audit_assignments.update_one(
        {"id": assignment_id},
        {"$set": {"status": "cancelled"}}
    )
    
    # Update audit auditors list
    audit = await db.audit_schedules.find_one({"id": assignment['audit_id']})
    if audit:
        auditor = await db.auditors.find_one({"id": assignment['auditor_id']})
        if auditor:
            current_auditors = audit.get('auditors', '')
            auditor_name = auditor.get('name', '')
            # Remove auditor name from list
            auditor_list = [a.strip() for a in current_auditors.split(',') if a.strip() != auditor_name]
            await db.audit_schedules.update_one(
                {"id": assignment['audit_id']},
                {"$set": {"auditors": ', '.join(auditor_list)}}
            )
    
    return {"message": "Assignment removed"}

# ================= AUDIT SCHEDULE ROUTES =================

@api_router.get("/audit-schedules")
async def get_audit_schedules(credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Get all audit schedules"""
    await get_current_user(credentials)
    audits = await db.audit_schedules.find({}, {"_id": 0}).to_list(1000)
    
    # Enrich with organization names
    for audit in audits:
        if audit.get('contract_id'):
            proposal = await db.proposals.find_one({"id": audit['contract_id']}, {"_id": 0})
            if proposal:
                audit['organization_name'] = proposal.get('organization_name', '')
        if audit.get('site_id'):
            site = await db.sites.find_one({"id": audit['site_id']}, {"_id": 0})
            if site:
                audit['site_name'] = site.get('name', '')
    
    return audits

@api_router.post("/audit-schedules")
async def create_audit_schedule(audit_data: AuditScheduleCreate, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Create a new audit schedule with optional recurring events"""
    await get_current_user(credentials)
    
    # Get organization name from contract
    organization_name = ""
    contact_phone = ""
    contact_email = ""
    if audit_data.contract_id:
        proposal = await db.proposals.find_one({"id": audit_data.contract_id}, {"_id": 0})
        if proposal:
            organization_name = proposal.get('organization_name', '')
            contact_phone = proposal.get('contact_phone', '')
            contact_email = proposal.get('contact_email', '')
    
    # Get site name if provided
    site_name = ""
    if audit_data.site_id:
        site = await db.sites.find_one({"id": audit_data.site_id}, {"_id": 0})
        if site:
            site_name = site.get('name', '')
    
    # Create the main audit
    audit = AuditSchedule(
        contract_id=audit_data.contract_id,
        site_id=audit_data.site_id,
        organization_name=organization_name,
        site_name=site_name,
        audit_type=audit_data.audit_type,
        scheduled_date=audit_data.scheduled_date,
        scheduled_time=audit_data.scheduled_time,
        duration_days=audit_data.duration_days,
        auditors=audit_data.auditors,
        notes=audit_data.notes,
        is_recurring=audit_data.is_recurring,
        recurrence_type=audit_data.recurrence_type,
        recurrence_end_date=audit_data.recurrence_end_date
    )
    
    audit_doc = audit.model_dump()
    audit_doc['created_at'] = audit_doc['created_at'].isoformat()
    
    await db.audit_schedules.insert_one(audit_doc)
    
    created_audits = [audit.id]
    
    # Generate recurring events if enabled
    if audit_data.is_recurring and audit_data.recurrence_type and audit_data.recurrence_end_date:
        recurring_audits = generate_recurring_audits(
            parent_audit=audit,
            organization_name=organization_name,
            site_name=site_name,
            recurrence_type=audit_data.recurrence_type,
            end_date=audit_data.recurrence_end_date
        )
        
        for recurring_audit in recurring_audits:
            recurring_doc = recurring_audit.model_dump()
            recurring_doc['created_at'] = recurring_doc['created_at'].isoformat()
            await db.audit_schedules.insert_one(recurring_doc)
            created_audits.append(recurring_audit.id)
    
    return {"message": "Audit scheduled", "id": audit.id, "total_created": len(created_audits)}


def generate_recurring_audits(parent_audit: AuditSchedule, organization_name: str, site_name: str, recurrence_type: str, end_date: str) -> List[AuditSchedule]:
    """Generate recurring audit schedules based on recurrence type"""
    from dateutil.relativedelta import relativedelta
    
    recurring_audits = []
    
    try:
        start_date = datetime.strptime(parent_audit.scheduled_date, "%Y-%m-%d")
        end_date_dt = datetime.strptime(end_date, "%Y-%m-%d")
        
        # Calculate next dates based on recurrence type
        delta = None
        if recurrence_type == "weekly":
            delta = relativedelta(weeks=1)
        elif recurrence_type == "monthly":
            delta = relativedelta(months=1)
        elif recurrence_type == "quarterly":
            delta = relativedelta(months=3)
        elif recurrence_type == "yearly":
            delta = relativedelta(years=1)
        else:
            return recurring_audits
        
        current_date = start_date + delta
        
        while current_date <= end_date_dt:
            recurring_audit = AuditSchedule(
                contract_id=parent_audit.contract_id,
                site_id=parent_audit.site_id,
                organization_name=organization_name,
                site_name=site_name,
                audit_type=parent_audit.audit_type,
                scheduled_date=current_date.strftime("%Y-%m-%d"),
                scheduled_time=parent_audit.scheduled_time,
                duration_days=parent_audit.duration_days,
                auditors=parent_audit.auditors,
                notes=parent_audit.notes,
                is_recurring=True,
                recurrence_type=recurrence_type,
                parent_audit_id=parent_audit.id
            )
            recurring_audits.append(recurring_audit)
            current_date += delta
        
    except Exception as e:
        logging.error(f"Error generating recurring audits: {e}")
    
    return recurring_audits

@api_router.put("/audit-schedules/{audit_id}")
async def update_audit_schedule(audit_id: str, audit_data: AuditScheduleCreate, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Update an audit schedule"""
    await get_current_user(credentials)
    
    update_data = {
        "contract_id": audit_data.contract_id,
        "site_id": audit_data.site_id,
        "audit_type": audit_data.audit_type,
        "scheduled_date": audit_data.scheduled_date,
        "scheduled_time": audit_data.scheduled_time,
        "duration_days": audit_data.duration_days,
        "auditors": audit_data.auditors,
        "notes": audit_data.notes
    }
    
    await db.audit_schedules.update_one(
        {"id": audit_id},
        {"$set": update_data}
    )
    return {"message": "Audit updated", "id": audit_id}

@api_router.delete("/audit-schedules/{audit_id}")
async def delete_audit_schedule(audit_id: str, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Delete an audit schedule"""
    await get_current_user(credentials)
    await db.audit_schedules.delete_one({"id": audit_id})
    return {"message": "Audit deleted"}

# ================= CUSTOMER CONTACT HISTORY =================
# Note: Contact routes are now in routes/contacts.py

class ContactRecord(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    customer_id: str  # Links to form or proposal ID
    customer_name: str = ""
    contact_type: str = "call"  # call, email, meeting, other
    subject: str
    notes: str
    contact_date: str
    follow_up_date: str = ""
    follow_up_completed: bool = False
    created_by: str = ""
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ContactRecordCreate(BaseModel):
    customer_id: str
    contact_type: str = "call"
    subject: str
    notes: str
    contact_date: str
    follow_up_date: str = ""

# ================= INVOICE & PAYMENT ROUTES =================

async def generate_invoice_number():
    """Generate unique invoice number: INV-YYYY-XXXX"""
    year = datetime.now().year
    # Find the last invoice number for this year
    last_invoice = await db.invoices.find_one(
        {"invoice_number": {"$regex": f"^INV-{year}-"}},
        sort=[("invoice_number", -1)]
    )
    if last_invoice:
        last_num = int(last_invoice['invoice_number'].split('-')[-1])
        new_num = last_num + 1
    else:
        new_num = 1
    return f"INV-{year}-{new_num:04d}"

@api_router.get("/invoices")
async def get_invoices(
    credentials: HTTPAuthorizationCredentials = Depends(security),
    status: Optional[str] = None,
    contract_id: Optional[str] = None
):
    """Get all invoices with optional filtering"""
    await get_current_user(credentials)
    
    query = {}
    if status:
        query["status"] = status
    if contract_id:
        query["contract_id"] = contract_id
    
    invoices = await db.invoices.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    
    # Check for overdue invoices and update status
    today = datetime.now().strftime("%Y-%m-%d")
    for invoice in invoices:
        if invoice.get('status') == 'sent' and invoice.get('due_date') and invoice.get('due_date') < today:
            await db.invoices.update_one(
                {"id": invoice['id']},
                {"$set": {"status": "overdue"}}
            )
            invoice['status'] = 'overdue'
    
    return invoices

@api_router.get("/invoices/stats")
async def get_invoice_stats(credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Get invoice statistics"""
    await get_current_user(credentials)
    
    invoices = await db.invoices.find({}, {"_id": 0}).to_list(1000)
    today = datetime.now().strftime("%Y-%m-%d")
    
    total_invoiced = sum(inv.get('total_amount', 0) for inv in invoices)
    total_paid = sum(inv.get('paid_amount', 0) for inv in invoices if inv.get('status') == 'paid')
    total_pending = sum(inv.get('total_amount', 0) for inv in invoices if inv.get('status') in ['sent', 'viewed'])
    total_overdue = sum(inv.get('total_amount', 0) for inv in invoices if inv.get('status') == 'overdue' or (inv.get('status') == 'sent' and inv.get('due_date', '') < today))
    
    return {
        "total_invoices": len(invoices),
        "total_invoiced": total_invoiced,
        "total_paid": total_paid,
        "total_pending": total_pending,
        "total_overdue": total_overdue,
        "paid_count": len([inv for inv in invoices if inv.get('status') == 'paid']),
        "pending_count": len([inv for inv in invoices if inv.get('status') in ['sent', 'viewed']]),
        "overdue_count": len([inv for inv in invoices if inv.get('status') == 'overdue' or (inv.get('status') == 'sent' and inv.get('due_date', '') < today)]),
        "draft_count": len([inv for inv in invoices if inv.get('status') == 'draft'])
    }

@api_router.post("/invoices")
async def create_invoice(invoice_data: InvoiceCreate, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Create a new invoice from a contract/proposal"""
    await get_current_user(credentials)
    
    # Get contract/proposal details
    proposal = await db.proposals.find_one({"id": invoice_data.contract_id}, {"_id": 0})
    if not proposal:
        raise HTTPException(status_code=404, detail="Contract not found")
    
    # Generate invoice number
    invoice_number = await generate_invoice_number()
    
    # Calculate due date based on payment terms
    issue_date = datetime.now()
    if invoice_data.due_date:
        due_date = invoice_data.due_date
    else:
        days_map = {"due_on_receipt": 0, "net_15": 15, "net_30": 30, "net_60": 60}
        days = days_map.get(invoice_data.payment_terms, 30)
        due_date = (issue_date + timedelta(days=days)).strftime("%Y-%m-%d")
    
    # Create invoice items from proposal if not provided
    items = invoice_data.items
    if not items:
        service_fees = proposal.get('service_fees', {})
        items = []
        
        if service_fees.get('initial_certification', 0) > 0:
            items.append(InvoiceItem(
                description="Initial Certification Fee",
                description_ar="رسوم الشهادة الأولية",
                quantity=1,
                unit_price=service_fees['initial_certification'],
                total=service_fees['initial_certification']
            ).model_dump())
        
        if service_fees.get('surveillance_1', 0) > 0:
            items.append(InvoiceItem(
                description="Surveillance Audit 1 Fee",
                description_ar="رسوم تدقيق المراقبة 1",
                quantity=1,
                unit_price=service_fees['surveillance_1'],
                total=service_fees['surveillance_1']
            ).model_dump())
        
        if service_fees.get('surveillance_2', 0) > 0:
            items.append(InvoiceItem(
                description="Surveillance Audit 2 Fee",
                description_ar="رسوم تدقيق المراقبة 2",
                quantity=1,
                unit_price=service_fees['surveillance_2'],
                total=service_fees['surveillance_2']
            ).model_dump())
        
        if service_fees.get('recertification', 0) > 0:
            items.append(InvoiceItem(
                description="Recertification Fee",
                description_ar="رسوم إعادة الاعتماد",
                quantity=1,
                unit_price=service_fees['recertification'],
                total=service_fees['recertification']
            ).model_dump())
    
    # Calculate totals
    subtotal = sum(item.get('total', item.get('unit_price', 0) * item.get('quantity', 1)) for item in items)
    tax_amount = subtotal * (invoice_data.tax_rate / 100)
    total_amount = subtotal + tax_amount
    
    invoice = Invoice(
        invoice_number=invoice_number,
        contract_id=invoice_data.contract_id,
        organization_name=proposal.get('organization_name', ''),
        organization_address=proposal.get('organization_address', ''),
        contact_email=proposal.get('contact_email', ''),
        contact_phone=proposal.get('organization_phone', ''),
        items=items,
        subtotal=subtotal,
        tax_rate=invoice_data.tax_rate,
        tax_amount=tax_amount,
        total_amount=total_amount,
        currency=proposal.get('service_fees', {}).get('currency', 'SAR'),
        issue_date=issue_date.strftime("%Y-%m-%d"),
        due_date=due_date,
        payment_terms=invoice_data.payment_terms,
        notes=invoice_data.notes,
        status="draft"
    )
    
    invoice_doc = invoice.model_dump()
    invoice_doc['created_at'] = invoice_doc['created_at'].isoformat()
    
    await db.invoices.insert_one(invoice_doc)
    
    return {"message": "Invoice created", "invoice_id": invoice.id, "invoice_number": invoice_number}

@api_router.get("/invoices/{invoice_id}")
async def get_invoice(invoice_id: str, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Get invoice details"""
    await get_current_user(credentials)
    
    invoice = await db.invoices.find_one({"id": invoice_id}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    return invoice

@api_router.put("/invoices/{invoice_id}")
async def update_invoice(invoice_id: str, updates: dict, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Update invoice details"""
    await get_current_user(credentials)
    
    invoice = await db.invoices.find_one({"id": invoice_id}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    # Don't allow editing paid invoices
    if invoice.get('status') == 'paid':
        raise HTTPException(status_code=400, detail="Cannot edit paid invoice")
    
    updates['updated_at'] = datetime.now(timezone.utc).isoformat()
    
    # Recalculate totals if items changed
    if 'items' in updates:
        subtotal = sum(item.get('total', item.get('unit_price', 0) * item.get('quantity', 1)) for item in updates['items'])
        tax_rate = updates.get('tax_rate', invoice.get('tax_rate', 15))
        tax_amount = subtotal * (tax_rate / 100)
        updates['subtotal'] = subtotal
        updates['tax_amount'] = tax_amount
        updates['total_amount'] = subtotal + tax_amount
    
    await db.invoices.update_one({"id": invoice_id}, {"$set": updates})
    
    return {"message": "Invoice updated"}

@api_router.post("/invoices/{invoice_id}/send")
async def send_invoice(invoice_id: str, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Mark invoice as sent (would send email in production)"""
    await get_current_user(credentials)
    
    invoice = await db.invoices.find_one({"id": invoice_id}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    await db.invoices.update_one(
        {"id": invoice_id},
        {"$set": {
            "status": "sent",
            "sent_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    # Create notification
    await create_notification(
        notification_type="invoice_sent",
        title="تم إرسال فاتورة",
        message=f"تم إرسال الفاتورة {invoice.get('invoice_number')} إلى {invoice.get('organization_name')}",
        related_id=invoice_id,
        related_type="invoice"
    )
    
    return {"message": "Invoice sent"}

@api_router.post("/invoices/{invoice_id}/record-payment")
async def record_payment(invoice_id: str, payment_data: dict, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Record a payment against an invoice"""
    current_user = await get_current_user(credentials)
    
    invoice = await db.invoices.find_one({"id": invoice_id}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    amount = payment_data.get('amount', 0)
    if amount <= 0:
        raise HTTPException(status_code=400, detail="Payment amount must be positive")
    
    # Create payment record
    payment = PaymentRecord(
        invoice_id=invoice_id,
        amount=amount,
        payment_date=payment_data.get('payment_date', datetime.now().strftime("%Y-%m-%d")),
        payment_method=payment_data.get('payment_method', 'bank_transfer'),
        reference=payment_data.get('reference', ''),
        notes=payment_data.get('notes', ''),
        recorded_by=current_user.get('email', '')
    )
    
    payment_doc = payment.model_dump()
    payment_doc['created_at'] = payment_doc['created_at'].isoformat()
    await db.payments.insert_one(payment_doc)
    
    # Update invoice
    new_paid_amount = invoice.get('paid_amount', 0) + amount
    new_status = 'paid' if new_paid_amount >= invoice.get('total_amount', 0) else invoice.get('status')
    
    await db.invoices.update_one(
        {"id": invoice_id},
        {"$set": {
            "paid_amount": new_paid_amount,
            "status": new_status,
            "paid_date": payment_data.get('payment_date') if new_status == 'paid' else None,
            "payment_method": payment_data.get('payment_method', ''),
            "payment_reference": payment_data.get('reference', ''),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    # Create notification
    await create_notification(
        notification_type="payment_received",
        title="تم استلام دفعة",
        message=f"تم استلام دفعة بقيمة {amount} {invoice.get('currency', 'SAR')} للفاتورة {invoice.get('invoice_number')}",
        related_id=invoice_id,
        related_type="invoice"
    )
    
    return {"message": "Payment recorded", "new_status": new_status, "paid_amount": new_paid_amount}

@api_router.get("/invoices/{invoice_id}/payments")
async def get_invoice_payments(invoice_id: str, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Get all payments for an invoice"""
    await get_current_user(credentials)
    
    payments = await db.payments.find({"invoice_id": invoice_id}, {"_id": 0}).sort("payment_date", -1).to_list(100)
    return payments

@api_router.delete("/invoices/{invoice_id}")
async def delete_invoice(invoice_id: str, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Delete a draft invoice"""
    await get_current_user(credentials)
    
    invoice = await db.invoices.find_one({"id": invoice_id}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    if invoice.get('status') != 'draft':
        raise HTTPException(status_code=400, detail="Only draft invoices can be deleted")
    
    await db.invoices.delete_one({"id": invoice_id})
    return {"message": "Invoice deleted"}

# ================= REPORT EXPORT ROUTES =================

@api_router.get("/reports/export")
async def export_report(
    credentials: HTTPAuthorizationCredentials = Depends(security),
    format: str = "excel",  # excel or pdf
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    status: Optional[str] = None,
    standard: Optional[str] = None
):
    """Export reports to Excel or PDF"""
    await get_current_user(credentials)
    
    # Build query for forms
    forms_query = {}
    if status and status != "all":
        forms_query["status"] = status
    if start_date or end_date:
        forms_query["created_at"] = {}
        if start_date:
            forms_query["created_at"]["$gte"] = start_date
        if end_date:
            forms_query["created_at"]["$lte"] = end_date
    
    # Get data
    forms = await db.application_forms.find(forms_query, {"_id": 0}).to_list(1000)
    proposals = await db.proposals.find({}, {"_id": 0}).to_list(1000)
    
    # Filter by standard if specified
    if standard and standard != "all":
        forms = [f for f in forms if f.get('company_data') and standard in f.get('company_data', {}).get('certificationSchemes', [])]
        proposals = [p for p in proposals if standard in p.get('standards', [])]
    
    if format == "excel":
        # Generate Excel file
        import io
        try:
            import openpyxl
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            wb = Workbook()
            
            # Forms Sheet
            ws_forms = wb.active
            ws_forms.title = "Application Forms"
            
            # Headers
            headers = ["Company Name", "Contact", "Email", "Status", "Standards", "Created Date", "Submitted Date"]
            header_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            
            for col, header in enumerate(headers, 1):
                cell = ws_forms.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            
            # Data rows
            for row, form in enumerate(forms, 2):
                client_info = form.get('client_info') or {}
                company_data = form.get('company_data') or {}
                
                ws_forms.cell(row=row, column=1, value=client_info.get('company_name', ''))
                ws_forms.cell(row=row, column=2, value=client_info.get('name', ''))
                ws_forms.cell(row=row, column=3, value=client_info.get('email', ''))
                ws_forms.cell(row=row, column=4, value=form.get('status', ''))
                ws_forms.cell(row=row, column=5, value=', '.join(company_data.get('certificationSchemes', [])))
                ws_forms.cell(row=row, column=6, value=form.get('created_at', '')[:10] if form.get('created_at') else '')
                ws_forms.cell(row=row, column=7, value=form.get('submitted_at', '')[:10] if form.get('submitted_at') else '')
            
            # Adjust column widths
            for col in ws_forms.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws_forms.column_dimensions[column].width = min(max_length + 2, 50)
            
            # Proposals Sheet
            ws_proposals = wb.create_sheet("Proposals")
            
            headers = ["Organization", "Contact", "Standards", "Status", "Total Amount", "Issued Date"]
            for col, header in enumerate(headers, 1):
                cell = ws_proposals.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            
            for row, proposal in enumerate(proposals, 2):
                ws_proposals.cell(row=row, column=1, value=proposal.get('organization_name', ''))
                ws_proposals.cell(row=row, column=2, value=proposal.get('contact_person', ''))
                ws_proposals.cell(row=row, column=3, value=', '.join(proposal.get('standards', [])))
                ws_proposals.cell(row=row, column=4, value=proposal.get('status', ''))
                ws_proposals.cell(row=row, column=5, value=proposal.get('total_amount', 0))
                ws_proposals.cell(row=row, column=6, value=proposal.get('issued_date', '')[:10] if proposal.get('issued_date') else '')
            
            # Adjust column widths for proposals
            for col in ws_proposals.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws_proposals.column_dimensions[column].width = min(max_length + 2, 50)
            
            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            return Response(
                content=output.getvalue(),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": f"attachment; filename=report_{datetime.now().strftime('%Y%m%d')}.xlsx"
                }
            )
        except ImportError:
            raise HTTPException(status_code=500, detail="Excel export library not installed")
    
    elif format == "pdf":
        # Generate PDF report
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.units import inch
        import io
        
        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=18, textColor=colors.HexColor('#1B365D'))
        
        elements = []
        
        # Title
        elements.append(Paragraph("Reports Summary", title_style))
        elements.append(Spacer(1, 20))
        
        # Summary stats
        total_forms = len(forms)
        submitted_forms = len([f for f in forms if f.get('status') in ['submitted', 'under_review', 'approved', 'agreement_signed']])
        total_proposals = len(proposals)
        accepted_proposals = len([p for p in proposals if p.get('status') in ['accepted', 'agreement_signed']])
        total_revenue = sum(p.get('total_amount', 0) for p in proposals if p.get('status') in ['accepted', 'agreement_signed'])
        
        summary_data = [
            ["Metric", "Value"],
            ["Total Forms", str(total_forms)],
            ["Submitted Forms", str(submitted_forms)],
            ["Total Proposals", str(total_proposals)],
            ["Accepted Proposals", str(accepted_proposals)],
            ["Total Revenue (SAR)", f"{total_revenue:,.0f}"]
        ]
        
        summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1B365D')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(summary_table)
        elements.append(Spacer(1, 30))
        
        # Forms table
        elements.append(Paragraph("Application Forms", styles['Heading2']))
        elements.append(Spacer(1, 10))
        
        forms_data = [["Company", "Status", "Standards"]]
        for form in forms[:20]:  # Limit to 20 rows
            client_info = form.get('client_info') or {}
            company_data = form.get('company_data') or {}
            forms_data.append([
                (client_info.get('company_name', '') or '')[:30],
                form.get('status', ''),
                ', '.join(company_data.get('certificationSchemes', []))[:30]
            ])
        
        forms_table = Table(forms_data, colWidths=[2.5*inch, 1.5*inch, 2.5*inch])
        forms_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1B365D')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
        ]))
        
        elements.append(forms_table)
        
        doc.build(elements)
        output.seek(0)
        
        return Response(
            content=output.getvalue(),
            media_type="application/pdf",
            headers={
                "Content-Disposition": f"attachment; filename=report_{datetime.now().strftime('%Y%m%d')}.pdf"
            }
        )
    
    raise HTTPException(status_code=400, detail="Invalid format. Use 'excel' or 'pdf'")

# ================= DOCUMENT MANAGEMENT =================
# Note: Document routes are now in routes/documents.py

class Document(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    related_id: str  # Links to form, proposal, or contract ID
    related_type: str  # form, proposal, contract
    name: str
    file_type: str
    file_size: int = 0
    file_data: str  # Base64 encoded file data
    uploaded_by: str = ""
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class DocumentCreate(BaseModel):
    related_id: str
    related_type: str
    name: str
    file_type: str
    file_data: str  # Base64 encoded

# ================= SEED DATA =================

async def seed_default_templates():
    """Seed default certification packages and proposal templates"""
    
    # Check if already seeded
    existing_packages = await db.certification_packages.count_documents({})
    if existing_packages > 0:
        return
    
    # Default certification packages
    packages = [
        CertificationPackage(
            name="QMS Basic",
            name_ar="نظام إدارة الجودة الأساسي",
            description="ISO 9001 Quality Management System certification",
            description_ar="شهادة نظام إدارة الجودة ISO 9001",
            standards=["ISO9001"]
        ),
        CertificationPackage(
            name="EMS Basic",
            name_ar="نظام الإدارة البيئية الأساسي",
            description="ISO 14001 Environmental Management System certification",
            description_ar="شهادة نظام الإدارة البيئية ISO 14001",
            standards=["ISO14001"]
        ),
        CertificationPackage(
            name="OHS Basic",
            name_ar="نظام السلامة والصحة المهنية الأساسي",
            description="ISO 45001 Occupational Health & Safety certification",
            description_ar="شهادة نظام السلامة والصحة المهنية ISO 45001",
            standards=["ISO45001"]
        ),
        CertificationPackage(
            name="Integrated Management System",
            name_ar="نظام الإدارة المتكامل",
            description="Combined QMS, EMS, and OHS certification (ISO 9001, 14001, 45001)",
            description_ar="شهادة نظام الإدارة المتكامل (ISO 9001, 14001, 45001)",
            standards=["ISO9001", "ISO14001", "ISO45001"]
        ),
        CertificationPackage(
            name="Food Safety",
            name_ar="سلامة الغذاء",
            description="ISO 22000 Food Safety Management System certification",
            description_ar="شهادة نظام إدارة سلامة الغذاء ISO 22000",
            standards=["ISO22000"]
        ),
        CertificationPackage(
            name="Information Security",
            name_ar="أمن المعلومات",
            description="ISO 27001 Information Security Management certification",
            description_ar="شهادة نظام إدارة أمن المعلومات ISO 27001",
            standards=["ISO27001"]
        )
    ]
    
    for pkg in packages:
        pkg_doc = pkg.model_dump()
        pkg_doc['created_at'] = pkg_doc['created_at'].isoformat()
        await db.certification_packages.insert_one(pkg_doc)
    
    # Default proposal templates
    templates = [
        ProposalTemplate(
            name="Standard Pricing",
            name_ar="الأسعار القياسية",
            description="Standard pricing for most certifications",
            default_fees={
                "initial_certification": 15000,
                "surveillance_1": 8000,
                "surveillance_2": 8000,
                "recertification": 12000
            },
            default_notes="العرض شامل كافة التكاليف عدا السفر والإقامة",
            default_validity_days=30
        ),
        ProposalTemplate(
            name="Small Business",
            name_ar="المنشآت الصغيرة",
            description="Discounted pricing for small businesses (< 50 employees)",
            default_fees={
                "initial_certification": 10000,
                "surveillance_1": 5000,
                "surveillance_2": 5000,
                "recertification": 8000
            },
            default_notes="عرض خاص للمنشآت الصغيرة - شامل كافة التكاليف عدا السفر والإقامة",
            default_validity_days=30
        ),
        ProposalTemplate(
            name="Enterprise",
            name_ar="المؤسسات الكبيرة",
            description="Pricing for large enterprises (500+ employees)",
            default_fees={
                "initial_certification": 25000,
                "surveillance_1": 12000,
                "surveillance_2": 12000,
                "recertification": 20000
            },
            default_notes="عرض للمؤسسات الكبيرة - يشمل تدقيق متعدد المواقع",
            default_validity_days=45
        )
    ]
    
    for tmpl in templates:
        tmpl_doc = tmpl.model_dump()
        tmpl_doc['created_at'] = tmpl_doc['created_at'].isoformat()
        await db.proposal_templates.insert_one(tmpl_doc)
    
    logger.info("Default templates seeded successfully")

# ================= GOOGLE CALENDAR INTEGRATION =================

# Google Calendar OAuth Configuration
GOOGLE_CLIENT_ID = os.environ.get('GOOGLE_CLIENT_ID', '')
GOOGLE_CLIENT_SECRET = os.environ.get('GOOGLE_CLIENT_SECRET', '')
GOOGLE_REDIRECT_URI = os.environ.get('GOOGLE_REDIRECT_URI', '')

# Check if Google Calendar is configured
GOOGLE_CALENDAR_ENABLED = bool(GOOGLE_CLIENT_ID and GOOGLE_CLIENT_SECRET)

@api_router.get("/calendar/status")
async def get_calendar_status(credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Check if Google Calendar integration is enabled and connected"""
    current_user = await get_current_user(credentials)
    
    # Check if configured
    if not GOOGLE_CALENDAR_ENABLED:
        return {
            "enabled": False,
            "configured": False,
            "connected": False,
            "message": "Google Calendar integration not configured. Please add GOOGLE_CLIENT_ID and GOOGLE_CLIENT_SECRET to environment."
        }
    
    # Check if user has connected their calendar
    user = await db.users.find_one({"email": current_user['email']}, {"_id": 0})
    has_tokens = bool(user and user.get('google_tokens'))
    
    return {
        "enabled": True,
        "configured": True,
        "connected": has_tokens,
        "message": "Connected to Google Calendar" if has_tokens else "Click 'Connect Calendar' to enable sync"
    }

@api_router.get("/calendar/auth/url")
async def get_calendar_auth_url(credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Get Google Calendar OAuth authorization URL"""
    await get_current_user(credentials)
    
    if not GOOGLE_CALENDAR_ENABLED:
        raise HTTPException(status_code=400, detail="Google Calendar integration not configured")
    
    # Build authorization URL
    scopes = "https://www.googleapis.com/auth/calendar"
    auth_url = (
        f"https://accounts.google.com/o/oauth2/auth?"
        f"client_id={GOOGLE_CLIENT_ID}&"
        f"redirect_uri={GOOGLE_REDIRECT_URI}&"
        f"response_type=code&"
        f"scope={scopes}&"
        f"access_type=offline&"
        f"prompt=consent"
    )
    
    return {"authorization_url": auth_url}

@api_router.get("/oauth/calendar/callback")
async def calendar_oauth_callback(code: str):
    """Handle Google Calendar OAuth callback"""
    import requests
    
    if not GOOGLE_CALENDAR_ENABLED:
        raise HTTPException(status_code=400, detail="Google Calendar integration not configured")
    
    # Exchange code for tokens
    token_response = requests.post('https://oauth2.googleapis.com/token', data={
        'code': code,
        'client_id': GOOGLE_CLIENT_ID,
        'client_secret': GOOGLE_CLIENT_SECRET,
        'redirect_uri': GOOGLE_REDIRECT_URI,
        'grant_type': 'authorization_code'
    })
    
    if token_response.status_code != 200:
        raise HTTPException(status_code=400, detail="Failed to exchange code for tokens")
    
    tokens = token_response.json()
    
    # Get user email
    user_response = requests.get(
        'https://www.googleapis.com/oauth2/v2/userinfo',
        headers={'Authorization': f'Bearer {tokens["access_token"]}'}
    )
    
    if user_response.status_code != 200:
        raise HTTPException(status_code=400, detail="Failed to get user info")
    
    user_info = user_response.json()
    
    # Save tokens to user record
    await db.users.update_one(
        {"email": user_info['email']},
        {"$set": {"google_tokens": tokens}},
        upsert=False
    )
    
    # Redirect back to dashboard
    frontend_url = os.environ.get('FRONTEND_URL', '')
    return RedirectResponse(url=f"{frontend_url}/audit-scheduling?calendar_connected=true")

@api_router.post("/calendar/sync-audit")
async def sync_audit_to_calendar(
    audit_id: str,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Sync an audit schedule to Google Calendar"""
    import requests
    
    current_user = await get_current_user(credentials)
    
    if not GOOGLE_CALENDAR_ENABLED:
        raise HTTPException(status_code=400, detail="Google Calendar integration not configured")
    
    # Get user tokens
    user = await db.users.find_one({"email": current_user['email']}, {"_id": 0})
    if not user or not user.get('google_tokens'):
        raise HTTPException(status_code=400, detail="Google Calendar not connected")
    
    tokens = user['google_tokens']
    
    # Get audit schedule
    audit = await db.audit_schedules.find_one({"id": audit_id}, {"_id": 0})
    if not audit:
        raise HTTPException(status_code=404, detail="Audit schedule not found")
    
    # Get contract details for more info
    contract = await db.proposals.find_one({"id": audit.get('contract_id')}, {"_id": 0})
    org_name = contract.get('organization_name', 'Client') if contract else 'Client'
    
    # Create calendar event
    event = {
        'summary': f"Audit: {org_name}",
        'description': f"Audit Type: {audit.get('audit_type', 'N/A')}\nAuditor: {audit.get('auditor', 'N/A')}\nSite: {audit.get('site', 'N/A')}\nNotes: {audit.get('notes', '')}",
        'start': {
            'date': audit.get('start_date'),
            'timeZone': 'Asia/Riyadh'
        },
        'end': {
            'date': audit.get('end_date'),
            'timeZone': 'Asia/Riyadh'
        },
        'location': audit.get('site', ''),
        'reminders': {
            'useDefault': False,
            'overrides': [
                {'method': 'email', 'minutes': 24 * 60},
                {'method': 'popup', 'minutes': 60}
            ]
        }
    }
    
    # Create event in Google Calendar
    response = requests.post(
        'https://www.googleapis.com/calendar/v3/calendars/primary/events',
        headers={
            'Authorization': f'Bearer {tokens["access_token"]}',
            'Content-Type': 'application/json'
        },
        json=event
    )
    
    if response.status_code == 401:
        # Token expired, need to re-authenticate
        raise HTTPException(status_code=401, detail="Google Calendar token expired. Please reconnect.")
    
    if response.status_code not in [200, 201]:
        raise HTTPException(status_code=500, detail=f"Failed to create calendar event: {response.text}")
    
    event_data = response.json()
    
    # Save calendar event ID to audit
    await db.audit_schedules.update_one(
        {"id": audit_id},
        {"$set": {"calendar_event_id": event_data.get('id')}}
    )
    
    return {"message": "Audit synced to Google Calendar", "event_id": event_data.get('id')}

@api_router.delete("/calendar/disconnect")
async def disconnect_calendar(credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Disconnect Google Calendar"""
    current_user = await get_current_user(credentials)
    
    await db.users.update_one(
        {"email": current_user['email']},
        {"$unset": {"google_tokens": ""}}
    )
    
    return {"message": "Google Calendar disconnected"}


# ================= SMS NOTIFICATIONS (TWILIO) =================

# Twilio configuration
TWILIO_ACCOUNT_SID = os.environ.get('TWILIO_ACCOUNT_SID', '')
TWILIO_AUTH_TOKEN = os.environ.get('TWILIO_AUTH_TOKEN', '')
TWILIO_PHONE_NUMBER = os.environ.get('TWILIO_PHONE_NUMBER', '')
TWILIO_ENABLED = bool(TWILIO_ACCOUNT_SID and TWILIO_AUTH_TOKEN and TWILIO_PHONE_NUMBER)

class SMSRequest(BaseModel):
    to_phone: str
    message: str
    customer_name: str = ""

class SMSNotificationSettings(BaseModel):
    enabled: bool = True
    reminder_days_before: int = 1  # Days before audit to send reminder
    send_on_proposal_accepted: bool = True
    send_on_agreement_signed: bool = True
    send_on_audit_scheduled: bool = True

@api_router.get("/sms/status")
async def get_sms_status(credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Get SMS integration status"""
    await get_current_user(credentials)
    
    return {
        "enabled": TWILIO_ENABLED,
        "configured": TWILIO_ENABLED,
        "simulation_mode": not TWILIO_ENABLED,
        "message": "SMS notifications active" if TWILIO_ENABLED else "SMS running in simulation mode (no Twilio credentials)"
    }

@api_router.post("/sms/send")
async def send_sms(sms_data: SMSRequest, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Send SMS notification"""
    await get_current_user(credentials)
    
    if TWILIO_ENABLED:
        # Real Twilio integration
        try:
            from twilio.rest import Client
            client = Client(TWILIO_ACCOUNT_SID, TWILIO_AUTH_TOKEN)
            
            message = client.messages.create(
                body=sms_data.message,
                from_=TWILIO_PHONE_NUMBER,
                to=sms_data.to_phone
            )
            
            # Log the SMS
            sms_log = {
                "id": str(uuid.uuid4()),
                "to_phone": sms_data.to_phone,
                "customer_name": sms_data.customer_name,
                "message": sms_data.message,
                "status": "sent",
                "twilio_sid": message.sid,
                "sent_at": datetime.now(timezone.utc).isoformat()
            }
            await db.sms_logs.insert_one(sms_log)
            
            return {"success": True, "message": "SMS sent successfully", "sid": message.sid}
            
        except Exception as e:
            logging.error(f"Twilio SMS error: {e}")
            raise HTTPException(status_code=500, detail=f"Failed to send SMS: {str(e)}")
    else:
        # Simulation mode
        sms_log = {
            "id": str(uuid.uuid4()),
            "to_phone": sms_data.to_phone,
            "customer_name": sms_data.customer_name,
            "message": sms_data.message,
            "status": "simulated",
            "sent_at": datetime.now(timezone.utc).isoformat()
        }
        await db.sms_logs.insert_one(sms_log)
        
        logging.info(f"[SIMULATED SMS] To: {sms_data.to_phone}, Message: {sms_data.message}")
        
        return {
            "success": True, 
            "message": "SMS simulated (Twilio not configured)",
            "simulated": True,
            "log_id": sms_log["id"]
        }

@api_router.get("/sms/logs")
async def get_sms_logs(credentials: HTTPAuthorizationCredentials = Depends(security), limit: int = 50):
    """Get SMS notification logs"""
    await get_current_user(credentials)
    
    logs = await db.sms_logs.find({}, {"_id": 0}).sort("sent_at", -1).limit(limit).to_list(limit)
    return logs

@api_router.post("/sms/send-audit-reminder")
async def send_audit_reminder(audit_id: str, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Send SMS reminder for an upcoming audit"""
    await get_current_user(credentials)
    
    # Get audit details
    audit = await db.audit_schedules.find_one({"id": audit_id}, {"_id": 0})
    if not audit:
        raise HTTPException(status_code=404, detail="Audit not found")
    
    # Get contract/proposal for contact info
    proposal = await db.proposals.find_one({"id": audit.get('contract_id')}, {"_id": 0})
    if not proposal:
        raise HTTPException(status_code=404, detail="Contract not found")
    
    contact_phone = proposal.get('contact_phone', '')
    if not contact_phone:
        raise HTTPException(status_code=400, detail="No phone number available for this client")
    
    # Create reminder message (bilingual)
    message_ar = f"تذكير: موعد التدقيق المجدول لشركة {audit.get('organization_name', '')} في {audit.get('scheduled_date')} الساعة {audit.get('scheduled_time', '09:00')}. بيان للتحقق والمطابقة"
    message_en = f"Reminder: Audit scheduled for {audit.get('organization_name', '')} on {audit.get('scheduled_date')} at {audit.get('scheduled_time', '09:00')}. BAYAN Auditing"
    
    full_message = f"{message_ar}\n\n{message_en}"
    
    # Send the SMS
    sms_request = SMSRequest(
        to_phone=contact_phone,
        message=full_message,
        customer_name=proposal.get('contact_name', '')
    )
    
    result = await send_sms(sms_request, credentials)
    
    # Mark audit as reminder sent
    await db.audit_schedules.update_one(
        {"id": audit_id},
        {"$set": {"sms_reminder_sent": True, "sms_reminder_sent_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    return result


# ================= MULTILINGUAL PDF GENERATION =================

@api_router.get("/proposals/{proposal_id}/bilingual_pdf")
async def generate_bilingual_proposal_pdf(proposal_id: str, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Generate bilingual (Arabic + English) PDF for a proposal/quotation"""
    await get_current_user(credentials)
    
    proposal = await db.proposals.find_one({"id": proposal_id}, {"_id": 0})
    if not proposal:
        raise HTTPException(status_code=404, detail="Proposal not found")
    
    # Generate the PDF
    pdf_path = await generate_bilingual_proposal_pdf_file(proposal)
    
    return FileResponse(
        pdf_path,
        media_type='application/pdf',
        filename=f'proposal_{proposal_id}_bilingual.pdf'
    )

async def generate_bilingual_proposal_pdf_file(proposal: dict) -> str:
    """Generate a bilingual proposal PDF with company stamp (no signature - not yet approved)"""
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    import arabic_reshaper
    from bidi.algorithm import get_display
    import base64
    import io
    
    # Register Arabic fonts
    font_path = ROOT_DIR / "fonts" / "Amiri-Regular.ttf"
    font_bold_path = ROOT_DIR / "fonts" / "Amiri-Bold.ttf"
    arabic_font_available = False
    
    if font_path.exists():
        try:
            pdfmetrics.registerFont(TTFont('Amiri', str(font_path)))
            arabic_font_available = True
            if font_bold_path.exists():
                pdfmetrics.registerFont(TTFont('Amiri-Bold', str(font_bold_path)))
        except Exception as e:
            print(f"Error registering Arabic font: {e}")
            arabic_font_available = False
    
    # Logo path
    logo_path = ROOT_DIR / "assets" / "bayan-logo.png"
    
    pdf_path = CONTRACTS_DIR / f"proposal_{proposal['id']}_bilingual.pdf"
    c = canvas.Canvas(str(pdf_path), pagesize=A4)
    width, height = A4
    
    def draw_arabic_text(text, x, y, font_size=11, bold=False):
        """Draw Arabic text with proper reshaping"""
        if arabic_font_available:
            try:
                reshaped = arabic_reshaper.reshape(str(text))
                bidi_text = get_display(reshaped)
                font_name = 'Amiri-Bold' if bold and font_bold_path.exists() else 'Amiri'
                c.setFont(font_name, font_size)
                c.drawRightString(x, y, bidi_text)
                return
            except Exception as e:
                print(f"Error drawing Arabic text: {e}")
        c.setFont('Helvetica-Bold' if bold else 'Helvetica', font_size)
        c.drawRightString(x, y, str(text))
    
    def draw_section_header(en_text, ar_text, y_pos):
        """Draw bilingual section header"""
        c.setFillColor(colors.HexColor('#1e3a5f'))
        c.setFont('Helvetica-Bold', 12)
        c.drawString(50, y_pos, en_text)
        draw_arabic_text(ar_text, width - 50, y_pos, 12, bold=True)
        c.setStrokeColor(colors.HexColor('#1e3a5f'))
        c.setLineWidth(1)
        c.line(50, y_pos - 5, width - 50, y_pos - 5)
        return y_pos - 25
    
    def has_arabic_chars(text):
        """Check if text contains Arabic characters"""
        if not text:
            return False
        return any('\u0600' <= char <= '\u06FF' or '\u0750' <= char <= '\u077F' for char in str(text))
    
    def draw_text_value(text, x, y, font_size=9):
        """Draw text value using appropriate font based on content"""
        text_str = str(text) if text else 'N/A'
        if has_arabic_chars(text_str) and arabic_font_available:
            # Use Arabic font for Arabic text
            try:
                reshaped = arabic_reshaper.reshape(text_str)
                bidi_text = get_display(reshaped)
                c.setFont('Amiri', font_size)
                c.drawString(x, y, bidi_text)
            except Exception as e:
                print(f"Error rendering Arabic value: {e}")
                c.setFont('Helvetica', font_size)
                c.drawString(x, y, text_str)
        else:
            c.setFont('Helvetica', font_size)
            c.drawString(x, y, text_str)
    
    def draw_field(label_en, label_ar, value, y_pos):
        """Draw bilingual field with value on both sides - separated columns to prevent overlap"""
        value_str = str(value) if value else 'N/A'
        # Truncate long values to prevent overflow (max ~25 chars for each side)
        value_display = value_str[:30] + '...' if len(value_str) > 30 else value_str
        c.setFillColor(colors.black)
        # English side (left half of page) - label at 50, value at 145
        c.setFont('Helvetica-Bold', 9)
        c.drawString(50, y_pos, f"{label_en}:")
        # Draw value using appropriate font (Amiri for Arabic, Helvetica for English)
        draw_text_value(value_display, 145, y_pos, 9)
        # Arabic side (right half of page) - value at center-right, label at far right
        draw_arabic_text(f"{label_ar}:", width - 50, y_pos, 9, bold=True)
        draw_arabic_text(value_display, width - 160, y_pos, 9)
        return y_pos - 18
    
    # Header with logo
    c.setFillColor(colors.HexColor('#1e3a5f'))
    c.rect(0, height - 100, width, 100, fill=True, stroke=False)
    
    # Draw white background for logo then logo
    if logo_path.exists():
        try:
            # White rounded rectangle background for logo
            c.setFillColor(colors.white)
            c.roundRect(25, height - 95, 80, 80, 5, fill=True, stroke=False)
            # Draw logo
            c.drawImage(str(logo_path), 30, height - 90, width=70, height=70, preserveAspectRatio=True, mask='auto')
        except:
            pass
    
    # Title
    c.setFillColor(colors.white)
    c.setFont('Helvetica-Bold', 22)
    c.drawCentredString(width/2, height - 45, "PRICE QUOTATION")
    
    # Arabic title
    if arabic_font_available:
        try:
            reshaped = arabic_reshaper.reshape("عرض السعر")
            bidi_text = get_display(reshaped)
            c.setFont('Amiri-Bold' if font_bold_path.exists() else 'Amiri', 18)
            c.drawCentredString(width/2, height - 70, bidi_text)
        except:
            pass
    
    # Reference info
    c.setFont('Helvetica', 9)
    c.drawRightString(width - 30, height - 30, f"Ref: {proposal.get('id', 'N/A')[:8].upper()}")
    c.drawRightString(width - 30, height - 45, f"Date: {proposal.get('issued_date', 'N/A')[:10] if proposal.get('issued_date') else 'N/A'}")
    
    y = height - 130
    
    # Section 1: Client Information
    y = draw_section_header("1. CLIENT INFORMATION", "١. معلومات العميل", y)
    y = draw_field("Organization", "المنظمة", proposal.get('organization_name', 'N/A'), y)
    y = draw_field("Contact Person", "جهة الاتصال", proposal.get('contact_person', proposal.get('contact_name', 'N/A')), y)
    y = draw_field("Email", "البريد الإلكتروني", proposal.get('contact_email', 'N/A'), y)
    y = draw_field("Phone", "الهاتف", proposal.get('contact_phone', 'N/A'), y)
    y -= 10
    
    # Section 2: Certification Scope
    y = draw_section_header("2. CERTIFICATION SCOPE", "٢. نطاق الاعتماد", y)
    standards = proposal.get('standards', [])
    standards_text = ', '.join(standards) if standards else 'N/A'
    y = draw_field("Standards", "المعايير", standards_text, y)
    y = draw_field("Scope", "النطاق", proposal.get('scope', 'N/A'), y)
    y -= 10
    
    # Section 3: Audit Duration
    y = draw_section_header("3. AUDIT DURATION (Days)", "٣. مدة التدقيق (أيام)", y)
    audit_duration = proposal.get('audit_duration', {})
    y = draw_field("Stage 1", "المرحلة الأولى", audit_duration.get('stage_1', 'N/A'), y)
    y = draw_field("Stage 2", "المرحلة الثانية", audit_duration.get('stage_2', 'N/A'), y)
    y = draw_field("Surveillance 1", "تدقيق المراقبة 1", audit_duration.get('surveillance_1', 'N/A'), y)
    y = draw_field("Surveillance 2", "تدقيق المراقبة 2", audit_duration.get('surveillance_2', 'N/A'), y)
    y = draw_field("Recertification", "إعادة الاعتماد", audit_duration.get('recertification', 'N/A'), y)
    y -= 10
    
    # Section 4: Service Fees
    y = draw_section_header("4. SERVICE FEES", "٤. رسوم الخدمة", y)
    service_fees = proposal.get('service_fees', {})
    currency = service_fees.get('currency', 'SAR')
    
    def format_fee(amount):
        return f"{currency} {amount:,.2f}" if amount else f"{currency} 0.00"
    
    y = draw_field("Initial Certification", "الاعتماد الأولي", format_fee(service_fees.get('initial_certification', 0)), y)
    y = draw_field("Surveillance 1 Fee", "رسوم المراقبة 1", format_fee(service_fees.get('surveillance_1', 0)), y)
    y = draw_field("Surveillance 2 Fee", "رسوم المراقبة 2", format_fee(service_fees.get('surveillance_2', 0)), y)
    y = draw_field("Recertification Fee", "رسوم إعادة الاعتماد", format_fee(service_fees.get('recertification', 0)), y)
    y -= 5
    
    # Total Amount (highlighted)
    c.setFillColor(colors.HexColor('#e8f4e8'))
    c.rect(50, y - 5, width - 100, 25, fill=True, stroke=False)
    c.setStrokeColor(colors.HexColor('#1e3a5f'))
    c.rect(50, y - 5, width - 100, 25, fill=False, stroke=True)
    c.setFillColor(colors.HexColor('#1e3a5f'))
    c.setFont('Helvetica-Bold', 12)
    c.drawString(60, y + 3, "TOTAL AMOUNT:")
    c.drawString(200, y + 3, format_fee(proposal.get('total_amount', 0)))
    draw_arabic_text(":المبلغ الإجمالي", width - 60, y + 3, 12, bold=True)
    y -= 35
    
    # Section 5: Validity
    y = draw_section_header("5. VALIDITY", "٥. الصلاحية", y)
    y = draw_field("Valid Until", "صالح حتى", proposal.get('valid_until', 'N/A'), y)
    y = draw_field("Validity Days", "أيام الصلاحية", f"{proposal.get('validity_days', 30)} days", y)
    y -= 10
    
    # Section 6: Status
    status_ar = {
        'draft': 'مسودة',
        'sent': 'تم الإرسال للعميل',
        'pending': 'قيد الانتظار',
        'accepted': 'مقبول',
        'rejected': 'مرفوض',
        'agreement_signed': 'تم الإرسال للعميل',
        'modification_requested': 'طلب تعديل'
    }
    status_en = {
        'draft': 'Draft',
        'sent': 'Sent to Client',
        'pending': 'Pending',
        'accepted': 'Accepted',
        'rejected': 'Rejected',
        'agreement_signed': 'Sent to Client',
        'modification_requested': 'Modification Requested'
    }
    status = proposal.get('status', 'pending')
    y = draw_section_header("6. STATUS", "٦. الحالة", y)
    c.setFillColor(colors.black)
    c.setFont('Helvetica', 11)
    c.drawString(50, y, f"Current Status: {status_en.get(status, status.replace('_', ' ').title())}")
    draw_arabic_text(f"الحالة الحالية: {status_ar.get(status, status)}", width - 50, y, 11)
    y -= 30
    
    # Company Seal Section - Use official seal image
    seal_path = ROOT_DIR / "assets" / "company-seal.png"
    c.setFillColor(colors.HexColor('#1e3a5f'))
    c.setFont('Helvetica-Bold', 10)
    # Draw bilingual header for seal section
    c.drawString(width/2 - 80, y, "Company Seal /")
    draw_arabic_text("ختم الشركة", width/2 + 80, y, 10, bold=True)
    y -= 10
    
    # Draw the official company seal image - positioned with sufficient margin above footer (footer is 0-45)
    # Seal needs to be placed at minimum y=50 to avoid footer overlap
    seal_y = max(y - 80, 55)  # Ensure seal stays above footer with 10px buffer
    if seal_path.exists():
        try:
            c.drawImage(str(seal_path), width/2 - 40, seal_y, width=80, height=80, preserveAspectRatio=True, mask='auto')
        except Exception as e:
            print(f"Error drawing seal: {e}")
    
    # Footer with proper Arabic text rendering
    c.setFillColor(colors.HexColor('#1e3a5f'))
    c.rect(0, 0, width, 45, fill=True, stroke=False)
    c.setFillColor(colors.white)
    c.setFont('Helvetica', 9)
    # English part
    c.drawString(width/2 - 180, 28, "BAYAN Auditing & Conformity |")
    # Arabic part with proper reshaping
    if arabic_font_available:
        try:
            footer_ar = arabic_reshaper.reshape("بيان للتحقق والمطابقة")
            c.setFont('Amiri', 10)
            c.drawString(width/2 + 10, 26, get_display(footer_ar))
        except:
            pass
    c.setFont('Helvetica', 9)
    c.drawCentredString(width/2, 10, "3879 Al Khadar Street, Riyadh, 12282, Saudi Arabia")
    
    c.save()
    return str(pdf_path)


@api_router.get("/forms/{form_id}/bilingual_pdf")
async def generate_bilingual_form_pdf(form_id: str, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Generate bilingual (Arabic + English) PDF for a submitted form"""
    await get_current_user(credentials)
    
    form = await db.application_forms.find_one({"id": form_id}, {"_id": 0})
    if not form:
        raise HTTPException(status_code=404, detail="Form not found")
    
    # Generate the PDF
    pdf_path = await generate_bilingual_form_pdf_file(form)
    
    return FileResponse(
        pdf_path,
        media_type='application/pdf',
        filename=f'form_submission_{form_id}_bilingual.pdf'
    )

async def generate_bilingual_form_pdf_file(form: dict) -> str:
    """Generate a professional bilingual form submission PDF"""
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.lib import colors
    import arabic_reshaper
    from bidi.algorithm import get_display
    
    # Register Arabic font
    font_path = ROOT_DIR / "fonts" / "Amiri-Regular.ttf"
    font_bold_path = ROOT_DIR / "fonts" / "Amiri-Bold.ttf"
    arabic_font_available = False
    if font_path.exists():
        try:
            pdfmetrics.registerFont(TTFont('Amiri', str(font_path)))
            arabic_font_available = True
            if font_bold_path.exists():
                pdfmetrics.registerFont(TTFont('Amiri-Bold', str(font_bold_path)))
        except:
            pass
    
    logo_path = ROOT_DIR / "assets" / "bayan-logo.png"
    pdf_path = CONTRACTS_DIR / f"form_{form['id']}_bilingual.pdf"
    c = canvas.Canvas(str(pdf_path), pagesize=A4)
    width, height = A4
    
    # Colors
    primary_color = colors.HexColor('#1e3a5f')  # Dark blue for header/footer
    section_color = colors.HexColor('#4a7c9b')  # Muted blue for section titles
    light_bg = colors.HexColor('#f0f4f8')
    accent_color = colors.HexColor('#2563eb')
    
    def draw_arabic(text, x, y, size=10, bold=False, right_align=False):
        """Draw Arabic text with proper reshaping"""
        if arabic_font_available:
            try:
                reshaped = arabic_reshaper.reshape(str(text))
                bidi_text = get_display(reshaped)
                font = 'Amiri-Bold' if bold and font_bold_path.exists() else 'Amiri'
                c.setFont(font, size)
                if right_align:
                    c.drawRightString(x, y, bidi_text)
                else:
                    c.drawString(x, y, bidi_text)
            except:
                c.setFont('Helvetica', size)
                c.drawString(x, y, str(text))
        else:
            c.setFont('Helvetica', size)
            c.drawString(x, y, str(text))
    
    def has_arabic(text):
        if not text: return False
        return any('\u0600' <= c <= '\u06FF' for c in str(text))
    
    def draw_value(text, x, y, size=10):
        """Draw value with appropriate font"""
        text_str = str(text) if text else 'N/A'
        if has_arabic(text_str) and arabic_font_available:
            try:
                reshaped = arabic_reshaper.reshape(text_str)
                bidi_text = get_display(reshaped)
                c.setFont('Amiri', size)
                c.drawString(x, y, bidi_text)
            except:
                c.setFont('Helvetica', size)
                c.drawString(x, y, text_str)
        else:
            c.setFont('Helvetica', size)
            c.drawString(x, y, text_str)
    
    # ============ HEADER ============
    c.setFillColor(primary_color)
    c.rect(0, height - 80, width, 80, fill=True, stroke=False)
    
    # Logo with white background
    if logo_path.exists():
        try:
            c.setFillColor(colors.white)
            c.roundRect(20, height - 70, 60, 60, 5, fill=True, stroke=False)
            c.drawImage(str(logo_path), 23, height - 67, width=54, height=54, preserveAspectRatio=True, mask='auto')
        except: pass
    
    # Title
    c.setFillColor(colors.white)
    c.setFont('Helvetica-Bold', 20)
    c.drawCentredString(width/2, height - 40, "APPLICATION FORM")
    if arabic_font_available:
        try:
            ar_title = get_display(arabic_reshaper.reshape("طلب المنح"))
            c.setFont('Amiri-Bold', 16)
            c.drawCentredString(width/2, height - 60, ar_title)
        except: pass
    
    # Reference and Date
    c.setFont('Helvetica', 9)
    c.drawRightString(width - 25, height - 30, f"Ref: {form.get('id', 'N/A')[:8].upper()}")
    c.drawRightString(width - 25, height - 45, f"Date: {str(form.get('submitted_at', form.get('created_at', '')))[:10]}")
    
    company_data = form.get('company_data', {})
    client_info = form.get('client_info', {})
    
    # ============ SECTION DRAWING FUNCTIONS ============
    def draw_section_box(title_en, title_ar, y_start, fields, box_height=None):
        """Draw a section with background box and fields
        Layout: English Label | VALUE (centered) | Arabic Label
        Dropdown values are translated to Arabic
        """
        if box_height is None:
            box_height = len(fields) * 18 + 30
        
        # Translation map for dropdown values (show full Arabic label as in dropdown)
        translations = {
            # Legal Status - match dropdown exactly
            'private': 'شركة خاصة',
            'public': 'شركة عامة',
            'government': 'مؤسسة حكومية',
            'public_sector': 'قطاع عام',
            'sole_proprietorship': 'ملكية فردية',
            'non-profit': 'منظمة غير ربحية',
            'partnership': 'شراكة',
            'other': 'أخرى',
            # Certification Program
            'initial': 'اعتماد أولي',
            'renewal': 'تجديد الاعتماد',
            'transfer': 'نقل الاعتماد',
            'surveillance': 'تدقيق مراقبة',
            # Yes/No
            'yes': 'نعم',
            'no': 'لا',
            'Yes': 'نعم',
            'No': 'لا',
            'true': 'نعم',
            'false': 'لا',
            'True': 'نعم',
            'False': 'لا',
        }
        
        # Section header bar (different color from main header/footer)
        c.setFillColor(section_color)
        c.rect(30, y_start - 22, width - 60, 22, fill=True, stroke=False)
        c.setFillColor(colors.white)
        c.setFont('Helvetica-Bold', 10)
        c.drawString(40, y_start - 16, title_en)
        if arabic_font_available:
            try:
                ar = get_display(arabic_reshaper.reshape(title_ar))
                c.setFont('Amiri-Bold', 10)
                c.drawRightString(width - 40, y_start - 16, ar)
            except: pass
        
        # Content box
        c.setFillColor(light_bg)
        c.rect(30, y_start - box_height, width - 60, box_height - 22, fill=True, stroke=False)
        c.setStrokeColor(colors.HexColor('#d1d5db'))
        c.setLineWidth(0.5)
        c.rect(30, y_start - box_height, width - 60, box_height, fill=False, stroke=True)
        
        # Draw fields: English Label | Value (center) | Arabic Label
        y = y_start - 40
        center_x = width / 2
        
        for label_en, label_ar, value in fields:
            c.setFillColor(colors.black)
            val_str = str(value) if value else 'N/A'
            val_display = val_str[:35] + '...' if len(val_str) > 35 else val_str
            
            # Translate dropdown values to Arabic
            if val_display in translations:
                val_display = translations[val_display]
            elif val_display.lower() in translations:
                val_display = translations[val_display.lower()]
            
            # LEFT: English label
            c.setFont('Helvetica-Bold', 9)
            c.drawString(40, y, f"{label_en}:")
            
            # CENTER: Value (with appropriate font)
            if has_arabic(val_display):
                if arabic_font_available:
                    try:
                        reshaped = arabic_reshaper.reshape(val_display)
                        bidi_val = get_display(reshaped)
                        c.setFont('Amiri', 9)
                        c.drawCentredString(center_x, y, bidi_val)
                    except:
                        c.setFont('Helvetica', 9)
                        c.drawCentredString(center_x, y, val_display)
                else:
                    c.setFont('Helvetica', 9)
                    c.drawCentredString(center_x, y, val_display)
            else:
                c.setFont('Helvetica', 9)
                c.drawCentredString(center_x, y, val_display)
            
            # RIGHT: Arabic label
            if arabic_font_available:
                try:
                    ar_label = get_display(arabic_reshaper.reshape(f"{label_ar}:"))
                    c.setFont('Amiri-Bold', 9)
                    c.drawRightString(width - 40, y, ar_label)
                except: pass
            
            y -= 18
        
        return y_start - box_height - 8
    
    # ============ PAGE 1 CONTENT ============
    y = height - 100
    
    # Section 1: Company Information
    company_fields = [
        ("Company Name", "اسم الشركة", company_data.get('companyName', client_info.get('company_name', 'N/A'))),
        ("Legal Status", "الحالة القانونية", company_data.get('legalStatus', 'N/A')),
        ("Address", "العنوان", company_data.get('address', 'N/A')),
        ("Phone", "الهاتف", company_data.get('phoneNumber', client_info.get('phone', 'N/A'))),
        ("Email", "البريد الإلكتروني", company_data.get('email', client_info.get('email', 'N/A'))),
        ("Website", "الموقع الإلكتروني", company_data.get('website', 'N/A')),
    ]
    y = draw_section_box("1. COMPANY INFORMATION", "١. معلومات الشركة", y, company_fields)
    
    # Consistent spacing between sections on page 1
    section_gap = 18
    y -= section_gap
    
    # Section 2: Contact Information
    contact_fields = [
        ("Contact Person", "جهة الاتصال", company_data.get('contactPerson', client_info.get('name', 'N/A'))),
        ("Designation", "المنصب", company_data.get('designation', 'N/A')),
        ("Mobile", "الجوال", company_data.get('mobileNumber', 'N/A')),
        ("Contact Email", "البريد الإلكتروني", company_data.get('contactEmail', 'N/A')),
    ]
    y = draw_section_box("2. CONTACT INFORMATION", "٢. معلومات الاتصال", y, contact_fields)
    y -= section_gap
    
    # Section 3: Organization Details
    org_fields = [
        ("Total Employees", "إجمالي الموظفين", company_data.get('totalEmployees', 'N/A')),
        ("Full-Time Employees", "دوام كامل", company_data.get('fullTimeEmployees', 'N/A')),
        ("Part-Time Employees", "دوام جزئي", company_data.get('partTimeEmployees', 'N/A')),
        ("Number of Sites", "عدد المواقع", company_data.get('numberOfSites', 'N/A')),
        ("Location Shifts", "الورديات", company_data.get('locationShifts', 'N/A')),
        ("Key Business Processes", "العمليات الرئيسية", company_data.get('keyBusinessProcesses', 'N/A')),
    ]
    y = draw_section_box("3. ORGANIZATION DETAILS", "٣. تفاصيل المنظمة", y, org_fields)
    y -= section_gap
    
    # Section 4: Certification Standards (moved to page 1)
    standards = company_data.get('certificationSchemes', [])
    standards_text = ', '.join(standards) if standards else 'N/A'
    cert_fields = [
        ("Selected Standards", "المعايير المختارة", standards_text),
        ("Certification Program", "برنامج الاعتماد", company_data.get('certificationProgram', 'N/A')),
        ("Already Certified?", "معتمد حالياً؟", company_data.get('isAlreadyCertified', 'N/A')),
    ]
    if company_data.get('otherStandard'):
        cert_fields.append(("Other Standard", "معيار آخر", company_data.get('otherStandard')))
    y = draw_section_box("4. CERTIFICATION STANDARDS", "٤. معايير الاعتماد", y, cert_fields)
    y -= section_gap
    
    # Section 5: Sites Information (on page 1)
    site_fields = []
    site1 = company_data.get('site1Address', '')
    site2 = company_data.get('site2Address', '')
    if site1:
        site_fields.append(("Site 1 Address", "عنوان الموقع 1", site1))
    if site2:
        site_fields.append(("Site 2 Address", "عنوان الموقع 2", site2))
    if not site_fields:
        site_fields.append(("Primary Location", "الموقع الرئيسي", "Main office only"))
    y = draw_section_box("5. SITES INFORMATION", "٥. معلومات المواقع", y, site_fields)
    y -= section_gap
    
    # Section 6: Consultant Information (on page 1)
    consultant_fields = [
        ("Consultant Used?", "هل تم استخدام مستشار؟", company_data.get('isConsultantInvolved', 'N/A')),
    ]
    if company_data.get('consultantName'):
        consultant_fields.append(("Consultant Name", "اسم المستشار", company_data.get('consultantName')))
    y = draw_section_box("6. CONSULTANT", "٦. المستشار", y, consultant_fields)
    
    # Footer for page 1
    c.setFillColor(primary_color)
    c.rect(0, 0, width, 35, fill=True, stroke=False)
    c.setFillColor(colors.white)
    c.setFont('Helvetica', 9)
    c.drawCentredString(width/2 - 60, 18, "BAYAN Auditing & Conformity")
    if arabic_font_available:
        try:
            ar_footer = get_display(arabic_reshaper.reshape("بيان للتحقق والمطابقة"))
            c.setFont('Amiri', 9)
            c.drawString(width/2 + 20, 16, ar_footer)
        except: pass
    c.setFont('Helvetica', 8)
    c.drawCentredString(width/2, 6, "Page 1 of 2")
    
    # ============ PAGE 2 ============
    c.showPage()
    
    # Header for page 2
    c.setFillColor(primary_color)
    c.rect(0, height - 50, width, 50, fill=True, stroke=False)
    if logo_path.exists():
        try:
            c.setFillColor(colors.white)
            c.roundRect(20, height - 45, 40, 40, 4, fill=True, stroke=False)
            c.drawImage(str(logo_path), 22, height - 43, width=36, height=36, preserveAspectRatio=True, mask='auto')
        except: pass
    c.setFillColor(colors.white)
    c.setFont('Helvetica-Bold', 14)
    c.drawCentredString(width/2, height - 30, "APPLICATION FORM (Continued)")
    
    y = height - 70
    
    # Section 7: Declaration (on page 2)
    decl_fields = [
        ("Declared By", "المُقِر", company_data.get('declarationName', 'N/A')),
        ("Designation", "المنصب", company_data.get('declarationDesignation', 'N/A')),
        ("Agreement", "الموافقة", "Yes / نعم" if company_data.get('declarationAgreed') else "No / لا"),
    ]
    y = draw_section_box("7. DECLARATION", "٧. الإقرار", y, decl_fields)
    
    # ============ IMPORTANT NOTES SECTION (on page 2) ============
    c.setFillColor(section_color)
    c.rect(30, y - 22, width - 60, 22, fill=True, stroke=False)
    c.setFillColor(colors.white)
    # Draw English part with Helvetica
    c.setFont('Helvetica-Bold', 10)
    c.drawString(40, y - 16, "IMPORTANT NOTES /")
    # Draw Arabic part with Amiri font
    if arabic_font_available:
        try:
            ar_header = get_display(arabic_reshaper.reshape("ملاحظات هامة"))
            c.setFont('Amiri-Bold', 11)
            c.drawRightString(width - 40, y - 16, ar_header)
        except:
            pass
    
    # Notes content box
    notes_box_height = 85
    c.setFillColor(light_bg)
    c.rect(30, y - 22 - notes_box_height, width - 60, notes_box_height, fill=True, stroke=False)
    c.setStrokeColor(colors.HexColor('#d1d5db'))
    c.setLineWidth(0.5)
    c.rect(30, y - 22 - notes_box_height, width - 60, notes_box_height + 22, fill=False, stroke=True)
    
    c.setFillColor(colors.black)
    c.setFont('Helvetica', 9)
    # English notes (left side)
    notes_english = [
        "• Please ensure all information provided is accurate and up-to-date.",
        "• Incomplete applications may delay the certification process.",
        "• Additional documentation may be requested during the review process.",
        "• For inquiries, contact us at: info@bayan.sa or +966 11 XXX XXXX",
    ]
    # Arabic notes (right side) - matching English content
    notes_arabic = [
        "• يرجى التأكد من صحة جميع المعلومات المقدمة وحداثتها",
        "• قد تؤدي الطلبات غير المكتملة إلى تأخير عملية الاعتماد",
        "• قد يتم طلب وثائق إضافية خلال عملية المراجعة",
        "• للاستفسارات، تواصل معنا على: info@bayan.sa أو +966 11 XXX XXXX",
    ]
    
    ny = y - 35
    # Draw English notes on left, Arabic notes on right (side by side)
    for i in range(len(notes_english)):
        # English on left
        c.setFont('Helvetica', 9)
        c.drawString(40, ny, notes_english[i])
        # Arabic on right
        if i < len(notes_arabic):
            draw_arabic(notes_arabic[i], width - 40, ny, 9, right_align=True)
        ny -= 16
    
    y = y - 22 - notes_box_height - 10
    
    # ============ TERMS & CONDITIONS (on page 2) ============
    c.setFillColor(section_color)
    c.rect(30, y - 22, width - 60, 22, fill=True, stroke=False)
    c.setFillColor(colors.white)
    c.setFont('Helvetica-Bold', 10)
    c.drawString(40, y - 16, "TERMS & CONDITIONS /")
    # Draw Arabic part with Amiri font
    if arabic_font_available:
        try:
            ar_terms_header = get_display(arabic_reshaper.reshape("الشروط والأحكام"))
            c.setFont('Amiri-Bold', 11)
            c.drawRightString(width - 40, y - 16, ar_terms_header)
        except:
            pass
    
    # Terms content box - fixed height
    terms_box_height = 115
    c.setFillColor(light_bg)
    c.rect(30, y - 22 - terms_box_height, width - 60, terms_box_height, fill=True, stroke=False)
    c.setStrokeColor(colors.HexColor('#d1d5db'))
    c.setLineWidth(0.5)
    c.rect(30, y - 22 - terms_box_height, width - 60, terms_box_height + 22, fill=False, stroke=True)
    
    c.setFillColor(colors.black)
    c.setFont('Helvetica', 8)
    # English terms (left side)
    terms_english = [
        "1. This application is subject to review and approval by BAYAN.",
        "2. All information provided must be accurate and complete.",
        "3. The applicant agrees to comply with all certification requirements.",
        "4. Certification fees are non-refundable once the audit begins.",
        "5. BAYAN reserves the right to conduct surveillance audits.",
    ]
    # Arabic terms (right side) - matching English content
    terms_arabic = [
        "١. هذا الطلب خاضع للمراجعة والموافقة من قبل بيان",
        "٢. يجب أن تكون جميع المعلومات المقدمة دقيقة وكاملة",
        "٣. يوافق المتقدم على الامتثال لجميع متطلبات الاعتماد",
        "٤. رسوم الاعتماد غير قابلة للاسترداد بمجرد بدء التدقيق",
        "٥. تحتفظ بيان بحق إجراء عمليات تدقيق المراقبة",
    ]
    
    ty = y - 35
    for i in range(len(terms_english)):
        c.setFont('Helvetica', 8)
        c.drawString(40, ty, terms_english[i])
        if i < len(terms_arabic):
            draw_arabic(terms_arabic[i], width - 40, ty, 8, right_align=True)
        ty -= 14
    
    y = y - 22 - terms_box_height - 10
    
    # ============ CLIENT DECLARATION SECTION (at the end of page 2) ============
    # Single centered declaration box for client
    sig_box_height = 130
    sig_box_width = width - 60  # Full width box, centered
    
    # ===== CLIENT DECLARATION BOX (Centered) =====
    c.setFillColor(light_bg)
    c.rect(30, y - sig_box_height, sig_box_width, sig_box_height - 10, fill=True, stroke=False)
    c.setStrokeColor(section_color)
    c.setLineWidth(1.5)
    c.rect(30, y - sig_box_height, sig_box_width, sig_box_height - 10, fill=False, stroke=True)
    
    # Header - Client's Declaration
    c.setFillColor(section_color)
    c.setFont('Helvetica-Bold', 9)
    c.drawString(40, y - 20, "CLIENT'S DECLARATION OF THE ACCURACY OF DATA /")
    if arabic_font_available:
        try:
            ar_declaration = get_display(arabic_reshaper.reshape("إقرار العميل بصحة البيانات"))
            c.setFont('Amiri-Bold', 9)
            c.drawRightString(width - 40, y - 20, ar_declaration)
        except: pass
    
    # Client signature fields - arranged in two columns
    c.setFillColor(colors.black)
    col1_x = 50  # Left column
    col2_x = width/2 + 20  # Right column
    line_length = 140
    
    # Name field (left column)
    c.setFont('Helvetica', 8)
    c.drawString(col1_x, y - 45, "Name:")
    c.line(col1_x + 45, y - 47, col1_x + 45 + line_length, y - 47)
    if arabic_font_available:
        try:
            ar_name = get_display(arabic_reshaper.reshape("الاسم:"))
            c.setFont('Amiri', 9)
            c.drawRightString(col1_x + 45 + line_length + 40, y - 45, ar_name)
        except: pass
    
    # Date field (right column)
    c.setFont('Helvetica', 8)
    c.drawString(col2_x, y - 45, "Date:")
    c.line(col2_x + 45, y - 47, col2_x + 45 + line_length, y - 47)
    if arabic_font_available:
        try:
            ar_date = get_display(arabic_reshaper.reshape("التاريخ:"))
            c.setFont('Amiri', 9)
            c.drawRightString(col2_x + 45 + line_length + 40, y - 45, ar_date)
        except: pass
    
    # Signature field (left column)
    c.setFont('Helvetica', 8)
    c.drawString(col1_x, y - 70, "Signature:")
    c.line(col1_x + 55, y - 72, col1_x + 55 + line_length - 10, y - 72)
    if arabic_font_available:
        try:
            ar_sig = get_display(arabic_reshaper.reshape("التوقيع:"))
            c.setFont('Amiri', 9)
            c.drawRightString(col1_x + 45 + line_length + 40, y - 70, ar_sig)
        except: pass
    
    # Stamp field (right column)
    c.setFont('Helvetica', 8)
    c.drawString(col2_x, y - 70, "Stamp:")
    if arabic_font_available:
        try:
            ar_stamp = get_display(arabic_reshaper.reshape("الختم:"))
            c.setFont('Amiri', 9)
            c.drawRightString(col2_x + 45 + line_length + 40, y - 70, ar_stamp)
        except: pass
    
    # Draw stamp placeholder box
    stamp_box_size = 50
    stamp_x = col2_x + 50
    stamp_y = y - sig_box_height + 25
    c.setStrokeColor(colors.grey)
    c.setLineWidth(0.5)
    c.rect(stamp_x, stamp_y, stamp_box_size, stamp_box_size, fill=False, stroke=True)
    
    # Footer for page 2
    c.setFillColor(primary_color)
    c.rect(0, 0, width, 35, fill=True, stroke=False)
    c.setFillColor(colors.white)
    c.setFont('Helvetica', 9)
    c.drawCentredString(width/2 - 60, 18, "BAYAN Auditing & Conformity")
    if arabic_font_available:
        try:
            ar_footer = get_display(arabic_reshaper.reshape("بيان للتحقق والمطابقة"))
            c.setFont('Amiri', 9)
            c.drawString(width/2 + 20, 16, ar_footer)
        except: pass
    c.setFont('Helvetica', 8)
    c.drawCentredString(width/2, 6, "Page 2 of 2")
    
    c.save()
    return str(pdf_path)


@api_router.get("/reports/bilingual_pdf")
async def generate_bilingual_report_pdf(
    credentials: HTTPAuthorizationCredentials = Depends(security),
    report_type: str = "summary",
    start_date: Optional[str] = None,
    end_date: Optional[str] = None
):
    """Generate bilingual (Arabic + English) PDF report"""
    await get_current_user(credentials)
    
    # Get data based on report type
    if report_type == "summary":
        # Get all relevant data
        forms = await db.application_forms.find({}, {"_id": 0}).to_list(1000)
        proposals = await db.proposals.find({}, {"_id": 0}).to_list(1000)
        agreements = await db.certification_agreements.find({}, {"_id": 0}).to_list(1000)
        
        report_data = {
            "total_forms": len(forms),
            "total_proposals": len(proposals),
            "total_agreements": len(agreements),
            "accepted_proposals": len([p for p in proposals if p.get('status') in ['accepted', 'agreement_signed']]),
            "total_revenue": sum(p.get('total_amount', 0) for p in proposals if p.get('status') in ['accepted', 'agreement_signed'])
        }
    else:
        report_data = {}
    
    # Generate the PDF
    pdf_path = await generate_bilingual_report_pdf_file(report_data, report_type)
    
    return FileResponse(
        pdf_path,
        media_type='application/pdf',
        filename=f'report_{report_type}_bilingual.pdf'
    )

async def generate_bilingual_report_pdf_file(data: dict, report_type: str) -> str:
    """Generate a bilingual report PDF"""
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.lib import colors
    import arabic_reshaper
    from bidi.algorithm import get_display
    
    # Register Arabic font
    font_path = ROOT_DIR / "fonts" / "Amiri-Regular.ttf"
    if font_path.exists():
        try:
            pdfmetrics.registerFont(TTFont('Amiri', str(font_path)))
        except:
            pass
    
    pdf_path = CONTRACTS_DIR / f"report_{report_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    c = canvas.Canvas(str(pdf_path), pagesize=A4)
    width, height = A4
    
    def draw_arabic_text(text, x, y, font_size=12):
        try:
            reshaped = arabic_reshaper.reshape(str(text))
            bidi_text = get_display(reshaped)
            c.setFont('Amiri', font_size)
            c.drawRightString(x, y, bidi_text)
        except:
            c.setFont('Helvetica', font_size)
            c.drawRightString(x, y, str(text))
    
    # Header
    c.setFillColor(colors.HexColor('#1e3a5f'))
    c.rect(0, height - 100, width, 100, fill=True, stroke=False)
    
    c.setFillColor(colors.white)
    c.setFont('Helvetica-Bold', 24)
    c.drawCentredString(width/2, height - 50, "SUMMARY REPORT / تقرير ملخص")
    
    y = height - 140
    
    # English Section
    c.setFillColor(colors.black)
    c.setFont('Helvetica-Bold', 14)
    c.drawString(50, y, "ENGLISH")
    y -= 30
    
    c.setFont('Helvetica', 11)
    c.drawString(50, y, f"Total Applications: {data.get('total_forms', 0)}")
    y -= 20
    c.drawString(50, y, f"Total Proposals: {data.get('total_proposals', 0)}")
    y -= 20
    c.drawString(50, y, f"Accepted Proposals: {data.get('accepted_proposals', 0)}")
    y -= 20
    c.drawString(50, y, f"Total Contracts: {data.get('total_agreements', 0)}")
    y -= 20
    c.drawString(50, y, f"Total Revenue: SAR {data.get('total_revenue', 0):,.2f}")
    y -= 20
    c.drawString(50, y, f"Report Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    
    y -= 50
    
    # Divider
    c.setStrokeColor(colors.HexColor('#1e3a5f'))
    c.setLineWidth(2)
    c.line(50, y, width - 50, y)
    y -= 30
    
    # Arabic Section
    c.setFont('Helvetica-Bold', 14)
    draw_arabic_text("العربية", width - 50, y, 14)
    y -= 30
    
    draw_arabic_text(f"إجمالي الطلبات: {data.get('total_forms', 0)}", width - 50, y)
    y -= 20
    draw_arabic_text(f"إجمالي العروض: {data.get('total_proposals', 0)}", width - 50, y)
    y -= 20
    draw_arabic_text(f"العروض المقبولة: {data.get('accepted_proposals', 0)}", width - 50, y)
    y -= 20
    draw_arabic_text(f"إجمالي العقود: {data.get('total_agreements', 0)}", width - 50, y)
    y -= 20
    draw_arabic_text(f"إجمالي الإيرادات: {data.get('total_revenue', 0):,.2f} ريال", width - 50, y)
    
    # Footer
    c.setFillColor(colors.HexColor('#1e3a5f'))
    c.rect(0, 0, width, 50, fill=True, stroke=False)
    c.setFillColor(colors.white)
    c.setFont('Helvetica', 10)
    c.drawCentredString(width/2, 20, "BAYAN Auditing & Conformity | بيان للتحقق والمطابقة")
    
    c.save()
    return str(pdf_path)

# ================= CONTRACT REVIEW / AUDIT PROGRAM ROUTES =================

@api_router.post("/contract-reviews")
async def create_contract_review(data: ContractReviewCreate, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Create a new Contract Review from a signed agreement (Admin only)"""
    await get_current_user(credentials)
    
    # Get the agreement
    agreement = await db.certification_agreements.find_one({"id": data.agreement_id}, {"_id": 0})
    if not agreement:
        raise HTTPException(status_code=404, detail="Agreement not found")
    
    # Get the proposal for additional details
    proposal = await db.proposals.find_one({"id": agreement['proposal_id']}, {"_id": 0})
    if not proposal:
        raise HTTPException(status_code=404, detail="Proposal not found")
    
    # Check if contract review already exists for this agreement
    existing = await db.contract_reviews.find_one({"agreement_id": data.agreement_id})
    if existing:
        raise HTTPException(status_code=400, detail="Contract review already exists for this agreement")
    
    # Create contract review with auto-populated data
    contract_review = ContractReview(
        agreement_id=data.agreement_id,
        organization_name=agreement.get('organization_name', ''),
        standards=agreement.get('selected_standards', []),
        scope_of_services=agreement.get('scope_of_services', ''),
        total_employees=proposal.get('total_employees', 0),
        application_date=datetime.now(timezone.utc).strftime('%Y-%m-%d'),
        client_id=proposal.get('id', '')[:8]
    )
    
    await db.contract_reviews.insert_one(contract_review.dict())
    
    # Create notification
    await create_notification(
        "contract_review_created",
        "Contract Review Created",
        f"Contract review created for {contract_review.organization_name}",
        contract_review.id,
        "contract_review"
    )
    
    return {"id": contract_review.id, "access_token": contract_review.access_token, "message": "Contract review created successfully"}

@api_router.get("/contract-reviews")
async def get_contract_reviews(credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Get all contract reviews (Admin only)"""
    await get_current_user(credentials)
    
    reviews = await db.contract_reviews.find({}, {"_id": 0}).to_list(1000)
    return reviews

@api_router.get("/contract-reviews/{review_id}")
async def get_contract_review(review_id: str, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Get a specific contract review (Admin only)"""
    await get_current_user(credentials)
    
    review = await db.contract_reviews.find_one({"id": review_id}, {"_id": 0})
    if not review:
        raise HTTPException(status_code=404, detail="Contract review not found")
    
    return review

@api_router.put("/contract-reviews/{review_id}/admin")
async def update_contract_review_admin(review_id: str, data: ContractReviewAdminUpdate, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Update contract review with admin/auditor data"""
    await get_current_user(credentials)
    
    review = await db.contract_reviews.find_one({"id": review_id})
    if not review:
        raise HTTPException(status_code=404, detail="Contract review not found")
    
    update_data = data.dict()
    update_data['updated_at'] = datetime.now(timezone.utc)
    
    # Update status based on completion
    if data.reviewed_by_name and data.reviewed_by_date:
        update_data['status'] = 'completed'
    elif data.prepared_by_name:
        update_data['status'] = 'pending_review'
    
    await db.contract_reviews.update_one(
        {"id": review_id},
        {"$set": update_data}
    )
    
    return {"message": "Contract review updated successfully"}

@api_router.delete("/contract-reviews/{review_id}")
async def delete_contract_review(review_id: str, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Delete a contract review (Admin only)"""
    await get_current_user(credentials)
    
    result = await db.contract_reviews.delete_one({"id": review_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Contract review not found")
    
    return {"message": "Contract review deleted successfully"}

# Public Contract Review endpoints (for client access)

@api_router.get("/public/contract-reviews/{access_token}")
async def get_public_contract_review(access_token: str):
    """Get contract review for client (Public access with token)"""
    review = await db.contract_reviews.find_one({"access_token": access_token}, {"_id": 0})
    if not review:
        raise HTTPException(status_code=404, detail="Contract review not found")
    
    # Return only data relevant for client
    return {
        "id": review['id'],
        "organization_name": review.get('organization_name', ''),
        "standards": review.get('standards', []),
        "scope_of_services": review.get('scope_of_services', ''),
        "total_employees": review.get('total_employees', 0),
        "application_date": review.get('application_date', ''),
        "client_id": review.get('client_id', ''),
        # Client-filled data
        "consultant_name": review.get('consultant_name', ''),
        "consultant_affects_impartiality": review.get('consultant_affects_impartiality', False),
        "consultant_impact_explanation": review.get('consultant_impact_explanation', ''),
        "exclusions_justification": review.get('exclusions_justification', ''),
        "client_submitted": review.get('client_submitted', False),
        "status": review.get('status', 'pending_client')
    }

@api_router.put("/public/contract-reviews/{access_token}")
async def submit_contract_review_client(access_token: str, data: ContractReviewClientSubmit):
    """Submit client data for contract review (Public access with token)"""
    review = await db.contract_reviews.find_one({"access_token": access_token})
    if not review:
        raise HTTPException(status_code=404, detail="Contract review not found")
    
    if review.get('client_submitted'):
        raise HTTPException(status_code=400, detail="Client data already submitted")
    
    update_data = data.dict()
    update_data['client_submitted'] = True
    update_data['client_submitted_at'] = datetime.now(timezone.utc)
    update_data['status'] = 'pending_review'
    update_data['updated_at'] = datetime.now(timezone.utc)
    
    await db.contract_reviews.update_one(
        {"access_token": access_token},
        {"$set": update_data}
    )
    
    # Create notification
    await create_notification(
        "contract_review_submitted",
        "Contract Review Data Submitted",
        f"Client submitted data for contract review: {review.get('organization_name', '')}",
        review['id'],
        "contract_review"
    )
    
    return {"message": "Contract review data submitted successfully"}

@api_router.get("/public/contract-reviews/{access_token}/pdf")
async def get_contract_review_pdf(access_token: str):
    """Generate PDF for contract review (Public access)"""
    review = await db.contract_reviews.find_one({"access_token": access_token}, {"_id": 0})
    if not review:
        raise HTTPException(status_code=404, detail="Contract review not found")
    
    try:
        pdf_bytes = generate_contract_review_pdf(review)
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={
                "Content-Disposition": f"attachment; filename=contract_review_{review['id'][:8]}.pdf"
            }
        )
    except Exception as e:
        logging.error(f"Error generating Contract Review PDF: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error generating PDF: {str(e)}")

@api_router.get("/contract-reviews/{review_id}/pdf")
async def get_contract_review_pdf_admin(review_id: str, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Generate PDF for contract review (Admin only)"""
    await get_current_user(credentials)
    
    review = await db.contract_reviews.find_one({"id": review_id}, {"_id": 0})
    if not review:
        raise HTTPException(status_code=404, detail="Contract review not found")
    
    try:
        pdf_bytes = generate_contract_review_pdf(review)
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={
                "Content-Disposition": f"attachment; filename=contract_review_{review_id[:8]}.pdf"
            }
        )
    except Exception as e:
        logging.error(f"Error generating Contract Review PDF: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error generating PDF: {str(e)}")

@api_router.post("/contract-reviews/{review_id}/send-link")
async def send_contract_review_link(review_id: str, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Send contract review link to client (Admin only)"""
    await get_current_user(credentials)
    
    review = await db.contract_reviews.find_one({"id": review_id}, {"_id": 0})
    if not review:
        raise HTTPException(status_code=404, detail="Contract review not found")
    
    # Get agreement and proposal for contact email
    agreement = await db.certification_agreements.find_one({"id": review['agreement_id']}, {"_id": 0})
    if not agreement:
        raise HTTPException(status_code=404, detail="Agreement not found")
    
    proposal = await db.proposals.find_one({"id": agreement['proposal_id']}, {"_id": 0})
    if not proposal:
        raise HTTPException(status_code=404, detail="Proposal not found")
    
    # Create notification with the link
    frontend_url = os.environ.get('FRONTEND_URL', 'https://audit-workflow-pro-1.preview.emergentagent.com')
    link = f"{frontend_url}/contract-review/{review['access_token']}"
    
    await create_notification(
        "contract_review_link_sent",
        "Contract Review Link Sent",
        f"Contract review link sent to {proposal.get('contact_email', '')} for {review.get('organization_name', '')}",
        review['id'],
        "contract_review"
    )
    
    # TODO: Send actual email when SMTP is configured
    return {
        "message": "Contract review link ready to send",
        "link": link,
        "email": proposal.get('contact_email', ''),
        "organization": review.get('organization_name', '')
    }

# ================= AUDIT PROGRAM (BACF6-05) ROUTES =================

@api_router.post("/audit-programs")
async def create_audit_program(data: AuditProgramCreate, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Create a new Audit Program from a Contract Review (Admin only)"""
    await get_current_user(credentials)
    
    # Get the contract review
    review = await db.contract_reviews.find_one({"id": data.contract_review_id}, {"_id": 0})
    if not review:
        raise HTTPException(status_code=404, detail="Contract review not found")
    
    # Check if audit program already exists for this contract review
    existing = await db.audit_programs.find_one({"contract_review_id": data.contract_review_id})
    if existing:
        raise HTTPException(status_code=400, detail="Audit program already exists for this contract review")
    
    # Create audit program with auto-populated data
    audit_program = AuditProgram(
        contract_review_id=data.contract_review_id,
        agreement_id=review.get('agreement_id', ''),
        organization_name=review.get('organization_name', ''),
        standards=review.get('standards', []),
        scope_of_services=review.get('scope_of_services', ''),
        total_employees=review.get('total_employees', 0),
        # Default activities for common audit stages
        activities=[
            {"activity": "Document Review", "audit_type": "Desktop", "stage1": "", "stage2": "", "sur1": "", "sur2": "", "rc": "", "planned_date": ""},
            {"activity": "Opening Meeting", "audit_type": "On-site", "stage1": "", "stage2": "", "sur1": "", "sur2": "", "rc": "", "planned_date": ""},
            {"activity": "Management Review", "audit_type": "On-site", "stage1": "", "stage2": "", "sur1": "", "sur2": "", "rc": "", "planned_date": ""},
            {"activity": "Process Audits", "audit_type": "On-site", "stage1": "", "stage2": "", "sur1": "", "sur2": "", "rc": "", "planned_date": ""},
            {"activity": "Internal Audit Review", "audit_type": "On-site", "stage1": "", "stage2": "", "sur1": "", "sur2": "", "rc": "", "planned_date": ""},
            {"activity": "Closing Meeting", "audit_type": "On-site", "stage1": "", "stage2": "", "sur1": "", "sur2": "", "rc": "", "planned_date": ""},
        ]
    )
    
    await db.audit_programs.insert_one(audit_program.dict())
    
    # Create notification
    await create_notification(
        "audit_program_created",
        "Audit Program Created",
        f"Audit program created for {audit_program.organization_name}",
        audit_program.id,
        "audit_program"
    )
    
    return {"id": audit_program.id, "access_token": audit_program.access_token, "message": "Audit program created successfully"}

@api_router.get("/audit-programs")
async def get_audit_programs(credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Get all audit programs (Admin only)"""
    await get_current_user(credentials)
    
    programs = await db.audit_programs.find({}, {"_id": 0}).to_list(1000)
    return programs

@api_router.get("/audit-programs/{program_id}")
async def get_audit_program(program_id: str, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Get a specific audit program (Admin only)"""
    await get_current_user(credentials)
    
    program = await db.audit_programs.find_one({"id": program_id}, {"_id": 0})
    if not program:
        raise HTTPException(status_code=404, detail="Audit program not found")
    
    return program

@api_router.put("/audit-programs/{program_id}")
async def update_audit_program(program_id: str, data: AuditProgramUpdate, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Update audit program (Admin only)"""
    await get_current_user(credentials)
    
    program = await db.audit_programs.find_one({"id": program_id})
    if not program:
        raise HTTPException(status_code=404, detail="Audit program not found")
    
    update_data = data.dict()
    update_data['updated_at'] = datetime.now(timezone.utc)
    
    # Update status based on approval
    if data.certification_manager and data.approval_date:
        update_data['status'] = 'approved'
    
    await db.audit_programs.update_one(
        {"id": program_id},
        {"$set": update_data}
    )
    
    return {"message": "Audit program updated successfully"}

@api_router.delete("/audit-programs/{program_id}")
async def delete_audit_program(program_id: str, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Delete an audit program (Admin only)"""
    await get_current_user(credentials)
    
    result = await db.audit_programs.delete_one({"id": program_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Audit program not found")
    
    return {"message": "Audit program deleted successfully"}

@api_router.get("/audit-programs/{program_id}/pdf")
async def get_audit_program_pdf(program_id: str, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Generate PDF for audit program (Admin only)"""
    await get_current_user(credentials)
    
    program = await db.audit_programs.find_one({"id": program_id}, {"_id": 0})
    if not program:
        raise HTTPException(status_code=404, detail="Audit program not found")
    
    try:
        pdf_bytes = generate_audit_program_pdf(program)
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={
                "Content-Disposition": f"attachment; filename=audit_program_{program_id[:8]}.pdf"
            }
        )
    except Exception as e:
        logging.error(f"Error generating Audit Program PDF: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error generating PDF: {str(e)}")

@api_router.post("/audit-programs/{program_id}/approve")
async def approve_audit_program(program_id: str, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Approve an audit program (Admin only)"""
    current_user = await get_current_user(credentials)
    
    program = await db.audit_programs.find_one({"id": program_id})
    if not program:
        raise HTTPException(status_code=404, detail="Audit program not found")
    
    await db.audit_programs.update_one(
        {"id": program_id},
        {"$set": {
            "status": "approved",
            "certification_manager": current_user.get('name', ''),
            "approval_date": datetime.now(timezone.utc).strftime('%Y-%m-%d'),
            "updated_at": datetime.now(timezone.utc)
        }}
    )
    
    # Create notification
    await create_notification(
        "audit_program_approved",
        "Audit Program Approved",
        f"Audit program approved for {program.get('organization_name', '')}",
        program_id,
        "audit_program"
    )
    
    return {"message": "Audit program approved successfully"}

# ================= JOB ORDER (BACF6-06) ROUTES =================

@api_router.post("/job-orders")
async def create_job_order(data: JobOrderCreate, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Create a new Job Order from an Audit Program (Admin only)"""
    await get_current_user(credentials)
    
    # Get the audit program
    program = await db.audit_programs.find_one({"id": data.audit_program_id}, {"_id": 0})
    if not program:
        raise HTTPException(status_code=404, detail="Audit program not found")
    
    # Get the auditor
    auditor = await db.auditors.find_one({"id": data.auditor_id}, {"_id": 0})
    if not auditor:
        raise HTTPException(status_code=404, detail="Auditor not found")
    
    # Get contract review for additional details
    contract_review = None
    if program.get('contract_review_id'):
        contract_review = await db.contract_reviews.find_one(
            {"id": program['contract_review_id']}, {"_id": 0}
        )
    
    # Create job order with auto-populated data
    job_order = JobOrder(
        audit_program_id=data.audit_program_id,
        contract_review_id=program.get('contract_review_id', ''),
        agreement_id=program.get('agreement_id', ''),
        # Auditor info
        auditor_id=data.auditor_id,
        auditor_name=auditor.get('name', ''),
        auditor_name_ar=auditor.get('name_ar', ''),
        auditor_email=auditor.get('email', ''),
        position=data.position,
        # Client/Audit details
        organization_name=program.get('organization_name', ''),
        organization_address=contract_review.get('organization_address', '') if contract_review else '',
        total_employees=program.get('total_employees', 0),
        phone=contract_review.get('phone', '') if contract_review else '',
        scope_of_services=program.get('scope_of_services', ''),
        standards=program.get('standards', []),
        audit_type=data.audit_type,
        audit_date=data.audit_date,
        client_ref=contract_review.get('client_ref', '') if contract_review else program.get('id', '')[:8],
    )
    
    await db.job_orders.insert_one(job_order.dict())
    
    # Update auditor's current assignments
    await db.auditors.update_one(
        {"id": data.auditor_id},
        {"$inc": {"current_assignments": 1}}
    )
    
    # Create notification
    await create_notification(
        "job_order_created",
        "Job Order Created",
        f"Job order created for {auditor.get('name', '')} - {program.get('organization_name', '')}",
        job_order.id,
        "job_order"
    )
    
    return {
        "id": job_order.id, 
        "access_token": job_order.access_token, 
        "message": "Job order created successfully"
    }

@api_router.get("/job-orders")
async def get_job_orders(credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Get all job orders (Admin only)"""
    await get_current_user(credentials)
    
    orders = await db.job_orders.find({}, {"_id": 0}).to_list(1000)
    return orders

@api_router.get("/job-orders/{order_id}")
async def get_job_order(order_id: str, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Get a specific job order (Admin only)"""
    await get_current_user(credentials)
    
    order = await db.job_orders.find_one({"id": order_id}, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="Job order not found")
    
    return order

@api_router.put("/job-orders/{order_id}")
async def update_job_order(order_id: str, data: JobOrderUpdate, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Update job order (Admin only)"""
    await get_current_user(credentials)
    
    order = await db.job_orders.find_one({"id": order_id})
    if not order:
        raise HTTPException(status_code=404, detail="Job order not found")
    
    update_data = data.dict()
    update_data['updated_at'] = datetime.now(timezone.utc)
    
    await db.job_orders.update_one(
        {"id": order_id},
        {"$set": update_data}
    )
    
    return {"message": "Job order updated successfully"}

@api_router.delete("/job-orders/{order_id}")
async def delete_job_order(order_id: str, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Delete a job order (Admin only)"""
    await get_current_user(credentials)
    
    order = await db.job_orders.find_one({"id": order_id})
    if not order:
        raise HTTPException(status_code=404, detail="Job order not found")
    
    # Decrease auditor's current assignments
    if order.get('auditor_id'):
        await db.auditors.update_one(
            {"id": order['auditor_id']},
            {"$inc": {"current_assignments": -1}}
        )
    
    await db.job_orders.delete_one({"id": order_id})
    
    return {"message": "Job order deleted successfully"}

@api_router.post("/job-orders/{order_id}/approve")
async def approve_job_order(order_id: str, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Certification Manager approves job order (Admin only)"""
    current_user = await get_current_user(credentials)
    
    order = await db.job_orders.find_one({"id": order_id})
    if not order:
        raise HTTPException(status_code=404, detail="Job order not found")
    
    await db.job_orders.update_one(
        {"id": order_id},
        {"$set": {
            "manager_approved": True,
            "certification_manager": current_user.get('name', 'Admin'),
            "manager_approval_date": datetime.now(timezone.utc).strftime('%Y-%m-%d'),
            "status": "pending_auditor",
            "updated_at": datetime.now(timezone.utc)
        }}
    )
    
    # Create notification
    await create_notification(
        "job_order_approved",
        "Job Order Approved",
        f"Job order approved - {order.get('organization_name', '')}",
        order_id,
        "job_order"
    )
    
    return {"message": "Job order approved successfully"}

@api_router.get("/job-orders/{order_id}/pdf")
async def get_job_order_pdf(order_id: str, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Generate PDF for job order (Admin only)"""
    await get_current_user(credentials)
    
    order = await db.job_orders.find_one({"id": order_id}, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="Job order not found")
    
    try:
        pdf_bytes = generate_job_order_pdf(order)
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={
                "Content-Disposition": f"attachment; filename=job_order_{order_id[:8]}.pdf"
            }
        )
    except Exception as e:
        logging.error(f"Error generating Job Order PDF: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error generating PDF: {str(e)}")

@api_router.post("/job-orders/{order_id}/send-to-auditor")
async def send_job_order_to_auditor(order_id: str, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Send job order confirmation link to auditor (Admin only)"""
    await get_current_user(credentials)
    
    order = await db.job_orders.find_one({"id": order_id}, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="Job order not found")
    
    if not order.get('manager_approved'):
        raise HTTPException(status_code=400, detail="Job order must be approved by manager first")
    
    # Generate confirmation link
    frontend_url = os.environ.get('REACT_APP_FRONTEND_URL', 'https://audit-workflow-pro-1.preview.emergentagent.com')
    link = f"{frontend_url}/job-order-confirm/{order['access_token']}"
    
    # TODO: Send actual email when SMTP is configured
    return {
        "message": "Job order confirmation link ready to send",
        "link": link,
        "auditor_email": order.get('auditor_email', ''),
        "auditor_name": order.get('auditor_name', ''),
        "organization": order.get('organization_name', '')
    }

# ================= PUBLIC JOB ORDER CONFIRMATION (for Auditors) =================

@api_router.get("/public/job-orders/{access_token}")
async def get_public_job_order(access_token: str):
    """Get job order for auditor confirmation (public access with token)"""
    order = await db.job_orders.find_one({"access_token": access_token}, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="Job order not found")
    
    if not order.get('manager_approved'):
        raise HTTPException(status_code=400, detail="Job order is not yet approved by manager")
    
    # Return limited info for auditor
    return {
        "id": order.get('id'),
        "auditor_name": order.get('auditor_name'),
        "auditor_name_ar": order.get('auditor_name_ar'),
        "position": order.get('position'),
        "organization_name": order.get('organization_name'),
        "standards": order.get('standards', []),
        "audit_type": order.get('audit_type'),
        "audit_date": order.get('audit_date'),
        "scope_of_services": order.get('scope_of_services'),
        "certification_manager": order.get('certification_manager'),
        "manager_approval_date": order.get('manager_approval_date'),
        "auditor_confirmed": order.get('auditor_confirmed', False),
        "status": order.get('status')
    }

@api_router.post("/public/job-orders/{access_token}/confirm")
async def confirm_job_order(access_token: str, data: JobOrderAuditorConfirmation):
    """Auditor confirms or rejects job order (public access with token)"""
    order = await db.job_orders.find_one({"access_token": access_token})
    if not order:
        raise HTTPException(status_code=404, detail="Job order not found")
    
    if not order.get('manager_approved'):
        raise HTTPException(status_code=400, detail="Job order is not yet approved by manager")
    
    if order.get('auditor_confirmed'):
        raise HTTPException(status_code=400, detail="Job order already confirmed")
    
    update_data = {
        "auditor_confirmed": data.confirmed,
        "auditor_confirmation_date": datetime.now(timezone.utc).strftime('%Y-%m-%d'),
        "updated_at": datetime.now(timezone.utc)
    }
    
    if data.confirmed:
        update_data["status"] = "confirmed"
    else:
        update_data["status"] = "rejected"
        update_data["unable_reason"] = data.unable_reason
        # Decrease auditor's current assignments if rejected
        if order.get('auditor_id'):
            await db.auditors.update_one(
                {"id": order['auditor_id']},
                {"$inc": {"current_assignments": -1}}
            )
    
    await db.job_orders.update_one(
        {"access_token": access_token},
        {"$set": update_data}
    )
    
    # Create notification
    status_text = "confirmed" if data.confirmed else "rejected"
    await create_notification(
        f"job_order_{status_text}",
        f"Job Order {status_text.title()}",
        f"Auditor {order.get('auditor_name', '')} has {status_text} the job order for {order.get('organization_name', '')}",
        order.get('id'),
        "job_order"
    )
    
    return {"message": f"Job order {status_text} successfully"}

# ================= STAGE 1 AUDIT PLAN (BACF6-07) ROUTES =================

@api_router.post("/stage1-audit-plans")
async def create_stage1_audit_plan(data: Stage1AuditPlanCreate, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Create a new Stage 1 Audit Plan from a confirmed Job Order (Admin only)"""
    await get_current_user(credentials)
    
    # Get the job order
    job_order = await db.job_orders.find_one({"id": data.job_order_id}, {"_id": 0})
    if not job_order:
        raise HTTPException(status_code=404, detail="Job order not found")
    
    if job_order.get('status') != 'confirmed':
        raise HTTPException(status_code=400, detail="Job order must be confirmed by auditor first")
    
    # Check if plan already exists for this job order
    existing = await db.stage1_audit_plans.find_one({"job_order_id": data.job_order_id})
    if existing:
        raise HTTPException(status_code=400, detail="Stage 1 Audit Plan already exists for this job order")
    
    # Get contract review for additional client details
    contract_review = None
    if job_order.get('contract_review_id'):
        contract_review = await db.contract_reviews.find_one(
            {"id": job_order['contract_review_id']}, {"_id": 0}
        )
    
    # Create plan with auto-populated data
    plan = Stage1AuditPlan(
        job_order_id=data.job_order_id,
        audit_program_id=job_order.get('audit_program_id', ''),
        contract_review_id=job_order.get('contract_review_id', ''),
        # Client info
        organization_name=job_order.get('organization_name', ''),
        file_no=job_order.get('client_ref', '') or data.job_order_id[:8],
        address=job_order.get('organization_address', ''),
        plan_date=datetime.now(timezone.utc).strftime('%Y-%m-%d'),
        contact_person=contract_review.get('contact_person', '') if contract_review else '',
        contact_phone=contract_review.get('phone', '') if contract_review else '',
        contact_email=contract_review.get('contact_email', '') if contract_review else '',
        # Audit details
        standards=job_order.get('standards', []),
        audit_type="Stage 1",
        audit_date_from=job_order.get('audit_date', ''),
        scope=job_order.get('scope_of_services', ''),
        # Team leader from job order
        team_leader={
            "auditor_id": job_order.get('auditor_id', ''),
            "name": job_order.get('auditor_name', ''),
            "name_ar": job_order.get('auditor_name_ar', ''),
            "role": job_order.get('position', 'Lead Auditor')
        },
        # Default schedule entries
        schedule_entries=[
            {"date_time": "", "process": "Opening Meeting", "process_owner": "Top Management", "clauses": "4.1, 4.2", "auditor": "Team"},
            {"date_time": "", "process": "Document Review", "process_owner": "", "clauses": "7.5", "auditor": ""},
            {"date_time": "", "process": "Management System Review", "process_owner": "", "clauses": "4.4, 5.0", "auditor": ""},
            {"date_time": "", "process": "Internal Audit Review", "process_owner": "", "clauses": "9.2", "auditor": ""},
            {"date_time": "", "process": "Management Review", "process_owner": "", "clauses": "9.3", "auditor": ""},
            {"date_time": "", "process": "Closing Meeting", "process_owner": "Top Management", "clauses": "N/A", "auditor": "Team"},
        ]
    )
    
    await db.stage1_audit_plans.insert_one(plan.dict())
    
    # Create notification
    await create_notification(
        "stage1_plan_created",
        "Stage 1 Audit Plan Created",
        f"Stage 1 audit plan created for {plan.organization_name}",
        plan.id,
        "stage1_audit_plan"
    )
    
    return {
        "id": plan.id,
        "access_token": plan.access_token,
        "message": "Stage 1 audit plan created successfully"
    }

@api_router.get("/stage1-audit-plans")
async def get_stage1_audit_plans(credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Get all Stage 1 Audit Plans (Admin only)"""
    await get_current_user(credentials)
    
    plans = await db.stage1_audit_plans.find({}, {"_id": 0}).to_list(1000)
    return plans

@api_router.get("/stage1-audit-plans/{plan_id}")
async def get_stage1_audit_plan(plan_id: str, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Get a specific Stage 1 Audit Plan (Admin only)"""
    await get_current_user(credentials)
    
    plan = await db.stage1_audit_plans.find_one({"id": plan_id}, {"_id": 0})
    if not plan:
        raise HTTPException(status_code=404, detail="Stage 1 Audit Plan not found")
    
    return plan

@api_router.put("/stage1-audit-plans/{plan_id}")
async def update_stage1_audit_plan(plan_id: str, data: Stage1AuditPlanUpdate, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Update Stage 1 Audit Plan (Admin only)"""
    await get_current_user(credentials)
    
    plan = await db.stage1_audit_plans.find_one({"id": plan_id})
    if not plan:
        raise HTTPException(status_code=404, detail="Stage 1 Audit Plan not found")
    
    update_data = {}
    for field, value in data.dict().items():
        if value is not None:
            if field == 'team_members':
                update_data[field] = [m.dict() if hasattr(m, 'dict') else m for m in value]
            elif field == 'schedule_entries':
                update_data[field] = [e.dict() if hasattr(e, 'dict') else e for e in value]
            else:
                update_data[field] = value
    
    update_data['updated_at'] = datetime.now(timezone.utc)
    
    await db.stage1_audit_plans.update_one(
        {"id": plan_id},
        {"$set": update_data}
    )
    
    return {"message": "Stage 1 Audit Plan updated successfully"}

@api_router.delete("/stage1-audit-plans/{plan_id}")
async def delete_stage1_audit_plan(plan_id: str, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Delete a Stage 1 Audit Plan (Admin only)"""
    await get_current_user(credentials)
    
    result = await db.stage1_audit_plans.delete_one({"id": plan_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Stage 1 Audit Plan not found")
    
    return {"message": "Stage 1 Audit Plan deleted successfully"}

@api_router.post("/stage1-audit-plans/{plan_id}/manager-approve")
async def manager_approve_stage1_plan(plan_id: str, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Manager internal approval of Stage 1 Audit Plan (Admin only)"""
    current_user = await get_current_user(credentials)
    
    plan = await db.stage1_audit_plans.find_one({"id": plan_id})
    if not plan:
        raise HTTPException(status_code=404, detail="Stage 1 Audit Plan not found")
    
    await db.stage1_audit_plans.update_one(
        {"id": plan_id},
        {"$set": {
            "manager_approved": True,
            "manager_name": current_user.get('name', 'Admin'),
            "manager_approval_date": datetime.now(timezone.utc).strftime('%Y-%m-%d'),
            "status": "manager_approved",
            "updated_at": datetime.now(timezone.utc)
        }}
    )
    
    # Create notification
    await create_notification(
        "stage1_plan_manager_approved",
        "Stage 1 Plan Manager Approved",
        f"Stage 1 audit plan approved internally for {plan.get('organization_name', '')}",
        plan_id,
        "stage1_audit_plan"
    )
    
    return {"message": "Stage 1 Audit Plan approved by manager"}

@api_router.post("/stage1-audit-plans/{plan_id}/send-to-client")
async def send_stage1_plan_to_client(plan_id: str, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Send Stage 1 Audit Plan to client for acceptance (Admin only)"""
    await get_current_user(credentials)
    
    plan = await db.stage1_audit_plans.find_one({"id": plan_id}, {"_id": 0})
    if not plan:
        raise HTTPException(status_code=404, detail="Stage 1 Audit Plan not found")
    
    if not plan.get('manager_approved'):
        raise HTTPException(status_code=400, detail="Plan must be approved by manager first")
    
    # Update status
    await db.stage1_audit_plans.update_one(
        {"id": plan_id},
        {"$set": {
            "sent_to_client": True,
            "status": "pending_client",
            "updated_at": datetime.now(timezone.utc)
        }}
    )
    
    # Generate client review link
    frontend_url = os.environ.get('REACT_APP_FRONTEND_URL', 'https://audit-workflow-pro-1.preview.emergentagent.com')
    link = f"{frontend_url}/stage1-plan-review/{plan['access_token']}"
    
    return {
        "message": "Stage 1 Audit Plan ready to send to client",
        "link": link,
        "client_email": plan.get('contact_email', ''),
        "organization": plan.get('organization_name', '')
    }

@api_router.get("/stage1-audit-plans/{plan_id}/pdf")
async def get_stage1_audit_plan_pdf(plan_id: str, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Generate PDF for Stage 1 Audit Plan (Admin only)"""
    await get_current_user(credentials)
    
    plan = await db.stage1_audit_plans.find_one({"id": plan_id}, {"_id": 0})
    if not plan:
        raise HTTPException(status_code=404, detail="Stage 1 Audit Plan not found")
    
    try:
        pdf_bytes = generate_stage1_audit_plan_pdf(plan)
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={
                "Content-Disposition": f"attachment; filename=stage1_audit_plan_{plan_id[:8]}.pdf"
            }
        )
    except Exception as e:
        logging.error(f"Error generating Stage 1 Audit Plan PDF: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error generating PDF: {str(e)}")

# ================= PUBLIC STAGE 1 AUDIT PLAN REVIEW (for Clients) =================

@api_router.get("/public/stage1-audit-plans/{access_token}")
async def get_public_stage1_audit_plan(access_token: str):
    """Get Stage 1 Audit Plan for client review (public access with token)"""
    plan = await db.stage1_audit_plans.find_one({"access_token": access_token}, {"_id": 0})
    if not plan:
        raise HTTPException(status_code=404, detail="Stage 1 Audit Plan not found")
    
    if not plan.get('sent_to_client'):
        raise HTTPException(status_code=400, detail="Plan has not been sent for client review yet")
    
    # Return plan data for client review
    return {
        "id": plan.get('id'),
        "organization_name": plan.get('organization_name'),
        "plan_date": plan.get('plan_date'),
        "standards": plan.get('standards', []),
        "audit_type": plan.get('audit_type'),
        "audit_date_from": plan.get('audit_date_from'),
        "audit_date_to": plan.get('audit_date_to'),
        "scope": plan.get('scope'),
        "team_leader": plan.get('team_leader', {}),
        "team_members": plan.get('team_members', []),
        "schedule_entries": plan.get('schedule_entries', []),
        "manager_name": plan.get('manager_name'),
        "manager_approval_date": plan.get('manager_approval_date'),
        "client_accepted": plan.get('client_accepted', False),
        "status": plan.get('status')
    }

@api_router.post("/public/stage1-audit-plans/{access_token}/respond")
async def client_respond_to_stage1_plan(access_token: str, data: Stage1AuditPlanClientResponse):
    """Client accepts or requests changes to Stage 1 Audit Plan (public access with token)"""
    plan = await db.stage1_audit_plans.find_one({"access_token": access_token})
    if not plan:
        raise HTTPException(status_code=404, detail="Stage 1 Audit Plan not found")
    
    if not plan.get('sent_to_client'):
        raise HTTPException(status_code=400, detail="Plan has not been sent for client review yet")
    
    if plan.get('client_accepted'):
        raise HTTPException(status_code=400, detail="Client has already accepted this plan")
    
    update_data = {
        "client_acceptance_date": datetime.now(timezone.utc).strftime('%Y-%m-%d'),
        "updated_at": datetime.now(timezone.utc)
    }
    
    if data.accepted:
        update_data["client_accepted"] = True
        update_data["status"] = "client_accepted"
        status_text = "accepted"
    else:
        update_data["client_change_requests"] = data.change_requests
        update_data["status"] = "changes_requested"
        status_text = "requested changes"
    
    await db.stage1_audit_plans.update_one(
        {"access_token": access_token},
        {"$set": update_data}
    )
    
    # Create notification
    await create_notification(
        f"stage1_plan_{status_text.replace(' ', '_')}",
        f"Stage 1 Plan Client {status_text.title()}",
        f"Client has {status_text} the Stage 1 audit plan for {plan.get('organization_name', '')}",
        plan.get('id'),
        "stage1_audit_plan"
    )
    
    return {"message": f"Response recorded: {status_text}"}

# ================= STAGE 2 AUDIT PLAN ENDPOINTS (BACF6-08) =================

@api_router.post("/stage2-audit-plans")
async def create_stage2_audit_plan(data: Stage2AuditPlanCreate, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Create a new Stage 2 Audit Plan (Admin only)
    Can be created from:
    1. An accepted Stage 1 plan (stage1_plan_id)
    2. A confirmed Job Order directly (job_order_id)
    """
    await get_current_user(credentials)
    
    source_data = None
    job_order = None
    stage1_plan = None
    
    # Determine source
    if data.stage1_plan_id:
        # Create from Stage 1 plan
        stage1_plan = await db.stage1_audit_plans.find_one({"id": data.stage1_plan_id}, {"_id": 0})
        if not stage1_plan:
            raise HTTPException(status_code=404, detail="Stage 1 Audit Plan not found")
        if stage1_plan.get('status') != 'client_accepted':
            raise HTTPException(status_code=400, detail="Stage 1 plan must be accepted by client first")
        
        # Check if Stage 2 already exists for this Stage 1
        existing = await db.stage2_audit_plans.find_one({"stage1_plan_id": data.stage1_plan_id})
        if existing:
            raise HTTPException(status_code=400, detail="Stage 2 Audit Plan already exists for this Stage 1 plan")
        
        source_data = stage1_plan
        job_order_id = stage1_plan.get('job_order_id', '')
        
    elif data.job_order_id:
        # Create directly from Job Order
        job_order = await db.job_orders.find_one({"id": data.job_order_id}, {"_id": 0})
        if not job_order:
            raise HTTPException(status_code=404, detail="Job order not found")
        if job_order.get('status') != 'confirmed':
            raise HTTPException(status_code=400, detail="Job order must be confirmed by auditor first")
        
        source_data = job_order
        job_order_id = data.job_order_id
    else:
        raise HTTPException(status_code=400, detail="Either stage1_plan_id or job_order_id is required")
    
    # Get contract review for additional client details
    contract_review = None
    contract_review_id = source_data.get('contract_review_id', '')
    if contract_review_id:
        contract_review = await db.contract_reviews.find_one({"id": contract_review_id}, {"_id": 0})
    
    # Create Stage 2 plan
    plan = Stage2AuditPlan(
        stage1_plan_id=data.stage1_plan_id,
        job_order_id=job_order_id,
        audit_program_id=source_data.get('audit_program_id', ''),
        contract_review_id=contract_review_id,
        # Client info
        organization_name=source_data.get('organization_name', ''),
        file_no=source_data.get('file_no', '') or source_data.get('client_ref', '') or job_order_id[:8],
        address=source_data.get('address', '') or source_data.get('organization_address', ''),
        plan_date=datetime.now(timezone.utc).strftime('%Y-%m-%d'),
        contact_person=source_data.get('contact_person', '') or (contract_review.get('contact_person', '') if contract_review else ''),
        contact_phone=source_data.get('contact_phone', '') or (contract_review.get('phone', '') if contract_review else ''),
        contact_email=source_data.get('contact_email', '') or (contract_review.get('contact_email', '') if contract_review else ''),
        # Audit details
        standards=source_data.get('standards', []),
        audit_type=data.audit_type,
        audit_date_from=source_data.get('audit_date_from', '') or source_data.get('audit_date', ''),
        scope=source_data.get('scope', '') or source_data.get('scope_of_services', ''),
        # Team (inherit from Stage 1 or Job Order)
        team_leader=source_data.get('team_leader', {}) if stage1_plan else {
            "auditor_id": source_data.get('auditor_id', ''),
            "name": source_data.get('auditor_name', ''),
            "name_ar": source_data.get('auditor_name_ar', ''),
            "role": source_data.get('position', 'Lead Auditor')
        },
        team_members=source_data.get('team_members', []) if stage1_plan else [],
        # Default schedule entries for Stage 2
        schedule_entries=[
            {"date_time": "", "process": "Opening Meeting", "process_owner": "Top Management", "clauses": "4.1, 4.2", "auditor": "Team"},
            {"date_time": "", "process": "Context & Leadership Review", "process_owner": "", "clauses": "4.0, 5.0", "auditor": ""},
            {"date_time": "", "process": "Process Audits - Operations", "process_owner": "", "clauses": "8.0", "auditor": ""},
            {"date_time": "", "process": "Process Audits - Support", "process_owner": "", "clauses": "7.0", "auditor": ""},
            {"date_time": "", "process": "Performance Evaluation Review", "process_owner": "", "clauses": "9.0", "auditor": ""},
            {"date_time": "", "process": "Improvement & Corrective Actions", "process_owner": "", "clauses": "10.0", "auditor": ""},
            {"date_time": "", "process": "Closing Meeting / Findings", "process_owner": "Top Management", "clauses": "N/A", "auditor": "Team"},
        ]
    )
    
    await db.stage2_audit_plans.insert_one(plan.dict())
    
    # Create notification
    await create_notification(
        "stage2_plan_created",
        "Stage 2 Audit Plan Created",
        f"Stage 2 audit plan created for {plan.organization_name}",
        plan.id,
        "stage2_audit_plan"
    )
    
    return {
        "id": plan.id,
        "access_token": plan.access_token,
        "message": "Stage 2 audit plan created successfully"
    }

@api_router.get("/stage2-audit-plans")
async def get_stage2_audit_plans(credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Get all Stage 2 Audit Plans (Admin only)"""
    await get_current_user(credentials)
    
    plans = await db.stage2_audit_plans.find({}, {"_id": 0}).to_list(1000)
    return plans

@api_router.get("/stage2-audit-plans/{plan_id}")
async def get_stage2_audit_plan(plan_id: str, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Get a specific Stage 2 Audit Plan (Admin only)"""
    await get_current_user(credentials)
    
    plan = await db.stage2_audit_plans.find_one({"id": plan_id}, {"_id": 0})
    if not plan:
        raise HTTPException(status_code=404, detail="Stage 2 Audit Plan not found")
    
    return plan

@api_router.put("/stage2-audit-plans/{plan_id}")
async def update_stage2_audit_plan(plan_id: str, data: Stage2AuditPlanUpdate, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Update Stage 2 Audit Plan (Admin only)"""
    await get_current_user(credentials)
    
    plan = await db.stage2_audit_plans.find_one({"id": plan_id})
    if not plan:
        raise HTTPException(status_code=404, detail="Stage 2 Audit Plan not found")
    
    update_data = {}
    for field, value in data.dict().items():
        if value is not None:
            if isinstance(value, list):
                update_data[field] = [item.dict() if hasattr(item, 'dict') else item for item in value]
            else:
                update_data[field] = value
    
    update_data["updated_at"] = datetime.now(timezone.utc)
    
    # Set status based on state
    if not plan.get('manager_approved'):
        update_data["status"] = "pending_manager"
    
    await db.stage2_audit_plans.update_one(
        {"id": plan_id},
        {"$set": update_data}
    )
    
    return {"message": "Stage 2 Audit Plan updated successfully"}

@api_router.delete("/stage2-audit-plans/{plan_id}")
async def delete_stage2_audit_plan(plan_id: str, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Delete Stage 2 Audit Plan (Admin only)"""
    await get_current_user(credentials)
    
    result = await db.stage2_audit_plans.delete_one({"id": plan_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Stage 2 Audit Plan not found")
    
    return {"message": "Stage 2 Audit Plan deleted successfully"}

@api_router.post("/stage2-audit-plans/{plan_id}/manager-approve")
async def manager_approve_stage2_plan(plan_id: str, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Manager approval for Stage 2 Audit Plan (Admin only)"""
    user = await get_current_user(credentials)
    
    plan = await db.stage2_audit_plans.find_one({"id": plan_id})
    if not plan:
        raise HTTPException(status_code=404, detail="Stage 2 Audit Plan not found")
    
    await db.stage2_audit_plans.update_one(
        {"id": plan_id},
        {"$set": {
            "manager_approved": True,
            "manager_name": user.get('email', 'Admin'),
            "manager_approval_date": datetime.now(timezone.utc).strftime('%Y-%m-%d'),
            "status": "manager_approved",
            "updated_at": datetime.now(timezone.utc)
        }}
    )
    
    return {"message": "Stage 2 Audit Plan approved by manager"}

@api_router.post("/stage2-audit-plans/{plan_id}/send-to-client")
async def send_stage2_plan_to_client(plan_id: str, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Send Stage 2 Audit Plan to client for acceptance (Admin only)"""
    await get_current_user(credentials)
    
    plan = await db.stage2_audit_plans.find_one({"id": plan_id}, {"_id": 0})
    if not plan:
        raise HTTPException(status_code=404, detail="Stage 2 Audit Plan not found")
    
    if not plan.get('manager_approved'):
        raise HTTPException(status_code=400, detail="Plan must be approved by manager first")
    
    # Update status
    await db.stage2_audit_plans.update_one(
        {"id": plan_id},
        {"$set": {
            "sent_to_client": True,
            "status": "pending_client",
            "updated_at": datetime.now(timezone.utc)
        }}
    )
    
    # Generate client review link
    frontend_url = os.environ.get('REACT_APP_FRONTEND_URL', os.environ.get('FRONTEND_URL', 'http://localhost:3000'))
    link = f"{frontend_url}/stage2-plan-review/{plan['access_token']}"
    
    return {
        "message": "Stage 2 Audit Plan ready to send to client",
        "link": link,
        "client_email": plan.get('contact_email', ''),
        "organization": plan.get('organization_name', '')
    }

@api_router.get("/stage2-audit-plans/{plan_id}/pdf")
async def get_stage2_audit_plan_pdf(plan_id: str, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Generate PDF for Stage 2 Audit Plan (Admin only)"""
    await get_current_user(credentials)
    
    plan = await db.stage2_audit_plans.find_one({"id": plan_id}, {"_id": 0})
    if not plan:
        raise HTTPException(status_code=404, detail="Stage 2 Audit Plan not found")
    
    try:
        pdf_bytes = generate_stage2_audit_plan_pdf(plan)
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={
                "Content-Disposition": f"attachment; filename=stage2_audit_plan_{plan_id[:8]}.pdf"
            }
        )
    except Exception as e:
        logging.error(f"Error generating Stage 2 Audit Plan PDF: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error generating PDF: {str(e)}")

# ================= PUBLIC STAGE 2 AUDIT PLAN REVIEW (for Clients) =================

@api_router.get("/public/stage2-audit-plans/{access_token}")
async def get_public_stage2_audit_plan(access_token: str):
    """Get Stage 2 Audit Plan for client review (public access with token)"""
    plan = await db.stage2_audit_plans.find_one({"access_token": access_token}, {"_id": 0})
    if not plan:
        raise HTTPException(status_code=404, detail="Stage 2 Audit Plan not found")
    
    if not plan.get('sent_to_client'):
        raise HTTPException(status_code=400, detail="Plan has not been sent for client review yet")
    
    # Return plan data for client review
    return {
        "id": plan.get('id'),
        "organization_name": plan.get('organization_name'),
        "plan_date": plan.get('plan_date'),
        "standards": plan.get('standards', []),
        "audit_type": plan.get('audit_type'),
        "audit_date_from": plan.get('audit_date_from'),
        "audit_date_to": plan.get('audit_date_to'),
        "scope": plan.get('scope'),
        "team_leader": plan.get('team_leader', {}),
        "team_members": plan.get('team_members', []),
        "schedule_entries": plan.get('schedule_entries', []),
        "manager_name": plan.get('manager_name'),
        "manager_approval_date": plan.get('manager_approval_date'),
        "client_accepted": plan.get('client_accepted', False),
        "status": plan.get('status')
    }

@api_router.post("/public/stage2-audit-plans/{access_token}/respond")
async def client_respond_to_stage2_plan(access_token: str, data: Stage2AuditPlanClientResponse):
    """Client accepts or requests changes to Stage 2 Audit Plan (public access with token)"""
    plan = await db.stage2_audit_plans.find_one({"access_token": access_token})
    if not plan:
        raise HTTPException(status_code=404, detail="Stage 2 Audit Plan not found")
    
    if not plan.get('sent_to_client'):
        raise HTTPException(status_code=400, detail="Plan has not been sent for client review yet")
    
    if plan.get('client_accepted'):
        raise HTTPException(status_code=400, detail="Client has already accepted this plan")
    
    update_data = {
        "client_acceptance_date": datetime.now(timezone.utc).strftime('%Y-%m-%d'),
        "updated_at": datetime.now(timezone.utc)
    }
    
    if data.accepted:
        update_data["client_accepted"] = True
        update_data["status"] = "client_accepted"
        status_text = "accepted"
    else:
        update_data["client_change_requests"] = data.change_requests
        update_data["status"] = "changes_requested"
        status_text = "requested changes"
    
    await db.stage2_audit_plans.update_one(
        {"access_token": access_token},
        {"$set": update_data}
    )
    
    # Create notification
    await create_notification(
        f"stage2_plan_{status_text.replace(' ', '_')}",
        f"Stage 2 Plan Client {status_text.title()}",
        f"Client has {status_text} the Stage 2 audit plan for {plan.get('organization_name', '')}",
        plan.get('id'),
        "stage2_audit_plan"
    )
    
    return {"message": f"Response recorded: {status_text}"}

# ================= OPENING & CLOSING MEETING ENDPOINTS (BACF6-09) =================

@api_router.post("/opening-closing-meetings")
async def create_opening_closing_meeting(data: OpeningClosingMeetingCreate, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Create Opening & Closing Meeting form - sent after Stage 1 audit (Admin only)"""
    await get_current_user(credentials)
    
    source_data = None
    job_order_id = ""
    
    # Get source data from Stage 1 plan or Job Order
    if data.stage1_plan_id:
        stage1_plan = await db.stage1_audit_plans.find_one({"id": data.stage1_plan_id}, {"_id": 0})
        if not stage1_plan:
            raise HTTPException(status_code=404, detail="Stage 1 Audit Plan not found")
        source_data = stage1_plan
        job_order_id = stage1_plan.get('job_order_id', '')
    elif data.job_order_id:
        job_order = await db.job_orders.find_one({"id": data.job_order_id}, {"_id": 0})
        if not job_order:
            raise HTTPException(status_code=404, detail="Job order not found")
        source_data = job_order
        job_order_id = data.job_order_id
    else:
        raise HTTPException(status_code=400, detail="Either stage1_plan_id or job_order_id is required")
    
    # Check if meeting form already exists
    existing = await db.opening_closing_meetings.find_one({
        "$or": [
            {"stage1_plan_id": data.stage1_plan_id} if data.stage1_plan_id else {"_id": None},
            {"job_order_id": job_order_id}
        ]
    })
    if existing:
        raise HTTPException(status_code=400, detail="Meeting form already exists for this audit")
    
    # Create meeting form
    meeting = OpeningClosingMeeting(
        stage1_plan_id=data.stage1_plan_id,
        job_order_id=job_order_id,
        audit_program_id=source_data.get('audit_program_id', ''),
        organization_name=source_data.get('organization_name', ''),
        file_no=source_data.get('file_no', '') or source_data.get('client_ref', '') or job_order_id[:8],
        address=source_data.get('address', '') or source_data.get('organization_address', ''),
        audit_type=source_data.get('audit_type', 'Stage 1'),
        audit_date=source_data.get('audit_date_from', '') or source_data.get('audit_date', ''),
        standards=source_data.get('standards', []),
        # Default 5 empty attendees
        attendees=[
            {"name": "", "designation": "", "opening_meeting_date": "", "closing_meeting_date": ""}
            for _ in range(5)
        ]
    )
    
    await db.opening_closing_meetings.insert_one(meeting.dict())
    
    # Create notification
    await create_notification(
        "meeting_form_created",
        "Meeting Form Created",
        f"Opening & Closing Meeting form created for {meeting.organization_name}",
        meeting.id,
        "opening_closing_meeting"
    )
    
    return {
        "id": meeting.id,
        "access_token": meeting.access_token,
        "message": "Opening & Closing Meeting form created successfully"
    }

@api_router.get("/opening-closing-meetings")
async def get_opening_closing_meetings(credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Get all Opening & Closing Meeting forms (Admin only)"""
    await get_current_user(credentials)
    
    meetings = await db.opening_closing_meetings.find({}, {"_id": 0}).to_list(1000)
    return meetings

@api_router.get("/opening-closing-meetings/{meeting_id}")
async def get_opening_closing_meeting(meeting_id: str, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Get a specific Opening & Closing Meeting form (Admin only)"""
    await get_current_user(credentials)
    
    meeting = await db.opening_closing_meetings.find_one({"id": meeting_id}, {"_id": 0})
    if not meeting:
        raise HTTPException(status_code=404, detail="Meeting form not found")
    
    return meeting

@api_router.delete("/opening-closing-meetings/{meeting_id}")
async def delete_opening_closing_meeting(meeting_id: str, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Delete Opening & Closing Meeting form (Admin only)"""
    await get_current_user(credentials)
    
    result = await db.opening_closing_meetings.delete_one({"id": meeting_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Meeting form not found")
    
    return {"message": "Meeting form deleted successfully"}

@api_router.post("/opening-closing-meetings/{meeting_id}/send-to-client")
async def send_meeting_to_client(meeting_id: str, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Send Meeting form to client for filling (Admin only)"""
    await get_current_user(credentials)
    
    meeting = await db.opening_closing_meetings.find_one({"id": meeting_id}, {"_id": 0})
    if not meeting:
        raise HTTPException(status_code=404, detail="Meeting form not found")
    
    # Update status
    await db.opening_closing_meetings.update_one(
        {"id": meeting_id},
        {"$set": {
            "sent_to_client": True,
            "updated_at": datetime.now(timezone.utc)
        }}
    )
    
    # Generate client link
    frontend_url = os.environ.get('REACT_APP_FRONTEND_URL', os.environ.get('FRONTEND_URL', 'http://localhost:3000'))
    link = f"{frontend_url}/meeting-form/{meeting['access_token']}"
    
    return {
        "message": "Meeting form ready to send to client",
        "link": link,
        "organization": meeting.get('organization_name', '')
    }

@api_router.get("/opening-closing-meetings/{meeting_id}/pdf")
async def get_meeting_pdf(meeting_id: str, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Generate PDF for Opening & Closing Meeting form (Admin only)"""
    await get_current_user(credentials)
    
    meeting = await db.opening_closing_meetings.find_one({"id": meeting_id}, {"_id": 0})
    if not meeting:
        raise HTTPException(status_code=404, detail="Meeting form not found")
    
    try:
        pdf_bytes = generate_opening_closing_meeting_pdf(meeting)
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={
                "Content-Disposition": f"attachment; filename=opening_closing_meeting_{meeting_id[:8]}.pdf"
            }
        )
    except Exception as e:
        logging.error(f"Error generating Meeting PDF: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error generating PDF: {str(e)}")

# ================= PUBLIC OPENING & CLOSING MEETING (for Clients) =================

@api_router.get("/public/opening-closing-meetings/{access_token}")
async def get_public_meeting(access_token: str):
    """Get Meeting form for client to fill (public access with token)"""
    meeting = await db.opening_closing_meetings.find_one({"access_token": access_token}, {"_id": 0})
    if not meeting:
        raise HTTPException(status_code=404, detail="Meeting form not found")
    
    if not meeting.get('sent_to_client'):
        raise HTTPException(status_code=400, detail="Form has not been sent for client completion yet")
    
    return {
        "id": meeting.get('id'),
        "organization_name": meeting.get('organization_name'),
        "file_no": meeting.get('file_no'),
        "audit_type": meeting.get('audit_type'),
        "audit_date": meeting.get('audit_date'),
        "standards": meeting.get('standards', []),
        "attendees": meeting.get('attendees', []),
        "opening_meeting_notes": meeting.get('opening_meeting_notes', ''),
        "closing_meeting_notes": meeting.get('closing_meeting_notes', ''),
        "status": meeting.get('status')
    }

@api_router.post("/public/opening-closing-meetings/{access_token}/submit")
async def submit_meeting_form(access_token: str, data: OpeningClosingMeetingSubmit):
    """Client submits the meeting attendance form (public access with token)"""
    meeting = await db.opening_closing_meetings.find_one({"access_token": access_token})
    if not meeting:
        raise HTTPException(status_code=404, detail="Meeting form not found")
    
    if not meeting.get('sent_to_client'):
        raise HTTPException(status_code=400, detail="Form has not been sent for client completion yet")
    
    if meeting.get('status') == 'submitted':
        raise HTTPException(status_code=400, detail="Form has already been submitted")
    
    # Update with submitted data
    update_data = {
        "attendees": [att.dict() for att in data.attendees],
        "opening_meeting_notes": data.opening_meeting_notes,
        "closing_meeting_notes": data.closing_meeting_notes,
        "status": "submitted",
        "submitted_date": datetime.now(timezone.utc).strftime('%Y-%m-%d'),
        "updated_at": datetime.now(timezone.utc)
    }
    
    await db.opening_closing_meetings.update_one(
        {"access_token": access_token},
        {"$set": update_data}
    )
    
    # Create notification
    await create_notification(
        "meeting_form_submitted",
        "Meeting Form Submitted",
        f"Opening & Closing Meeting form submitted for {meeting.get('organization_name', '')}",
        meeting.get('id'),
        "opening_closing_meeting"
    )
    
    return {"message": "Meeting form submitted successfully"}

# ================= STAGE 1 AUDIT REPORT ENDPOINTS (BACF6-10) =================

@api_router.post("/stage1-audit-reports")
async def create_stage1_audit_report(data: Stage1AuditReportCreate, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Create Stage 1 Audit Report - generated after Stage 1 audit (Admin only)"""
    await get_current_user(credentials)
    
    source_data = None
    stage1_plan = None
    meeting = None
    job_order_id = ""
    
    # Get source data
    if data.stage1_plan_id:
        stage1_plan = await db.stage1_audit_plans.find_one({"id": data.stage1_plan_id}, {"_id": 0})
        if not stage1_plan:
            raise HTTPException(status_code=404, detail="Stage 1 Audit Plan not found")
        source_data = stage1_plan
        job_order_id = stage1_plan.get('job_order_id', '')
    
    if data.meeting_id:
        meeting = await db.opening_closing_meetings.find_one({"id": data.meeting_id}, {"_id": 0})
    
    if not source_data:
        raise HTTPException(status_code=400, detail="stage1_plan_id is required")
    
    # Check if report already exists
    existing = await db.stage1_audit_reports.find_one({"stage1_plan_id": data.stage1_plan_id})
    if existing:
        raise HTTPException(status_code=400, detail="Audit report already exists for this Stage 1 plan")
    
    # Get contract review for additional details
    contract_review = None
    contract_review_id = source_data.get('contract_review_id', '')
    if contract_review_id:
        contract_review = await db.contract_reviews.find_one({"id": contract_review_id}, {"_id": 0})
    
    # Default checklist items
    default_checklist = [
        {"requirement": "Has client identified Legal and Statutory Requirements?", "status": "C", "comments": ""},
        {"requirement": "Is the Information documented as required as per the standard?", "status": "C", "comments": ""},
        {"requirement": "Is the scope having boundaries and specific to the organization?", "status": "C", "comments": ""},
        {"requirement": "Has the client understanding with all applicable requirements?", "status": "C", "comments": ""},
        {"requirement": "Has the company identified key performance, Process Indicators?", "status": "C", "comments": ""},
        {"requirement": "Has the Client site specific conditions evaluated?", "status": "C", "comments": ""},
        {"requirement": "Is process and Equipment used adequate?", "status": "C", "comments": ""},
        {"requirement": "Is client having Multisite? Level of control identified?", "status": "C", "comments": ""},
        {"requirement": "Is Internal Audit planned, performed and effective?", "status": "C", "comments": ""},
        {"requirement": "Is MRM planned, performed and Effective?", "status": "C", "comments": ""},
        {"requirement": "Has the client identified hazards and risk assessment?", "status": "C", "comments": ""},
        {"requirement": "Has the client identified Environmental aspects/impacts?", "status": "C", "comments": ""},
        {"requirement": "Has the client done emergency preparedness planning?", "status": "C", "comments": ""},
        {"requirement": "Is the resource adequate for Stage 2 audit?", "status": "C", "comments": ""},
    ]
    
    # Create report
    report = Stage1AuditReport(
        stage1_plan_id=data.stage1_plan_id,
        meeting_id=data.meeting_id or "",
        job_order_id=job_order_id,
        audit_program_id=source_data.get('audit_program_id', ''),
        # Organization details
        organization_name=source_data.get('organization_name', ''),
        address=source_data.get('address', '') or (contract_review.get('address', '') if contract_review else ''),
        site_address=contract_review.get('site_address', '') if contract_review else '',
        standards=source_data.get('standards', []),
        num_employees=contract_review.get('num_employees', '') if contract_review else '',
        num_shifts=contract_review.get('num_shifts', '1') if contract_review else '1',
        email=source_data.get('contact_email', '') or (contract_review.get('contact_email', '') if contract_review else ''),
        contact_person=source_data.get('contact_person', '') or (contract_review.get('contact_person', '') if contract_review else ''),
        phone=source_data.get('contact_phone', '') or (contract_review.get('phone', '') if contract_review else ''),
        scope=source_data.get('scope', ''),
        ea_code=contract_review.get('ea_code', '') if contract_review else '',
        exclusions=contract_review.get('exclusions', '') if contract_review else '',
        # Audit team
        audit_team={
            "lead_auditor": source_data.get('team_leader', {}).get('name', ''),
            "auditors": [m.get('name', '') for m in source_data.get('team_members', [])],
            "technical_experts": []
        },
        audit_duration=source_data.get('audit_duration', '1'),
        start_date=source_data.get('audit_date_from', ''),
        end_date=source_data.get('audit_date_to', '') or source_data.get('audit_date_from', ''),
        # Attendees from meeting if available
        attendees=meeting.get('attendees', []) if meeting else [],
        # Default checklist
        checklist_items=default_checklist,
        # Default declarations (all false initially)
        declarations={
            "sampling": False,
            "combined": False,
            "corrective_actions": False,
            "outcomes_effective": False,
            "internal_audit": False,
            "scope_appropriate": False,
            "capability": False,
            "objectives_fulfilled": False
        }
    )
    
    await db.stage1_audit_reports.insert_one(report.dict())
    
    # Create notification
    await create_notification(
        "audit_report_created",
        "Stage 1 Audit Report Created",
        f"Stage 1 Audit Report created for {report.organization_name}",
        report.id,
        "stage1_audit_report"
    )
    
    return {
        "id": report.id,
        "message": "Stage 1 Audit Report created successfully"
    }

@api_router.get("/stage1-audit-reports")
async def get_stage1_audit_reports(credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Get all Stage 1 Audit Reports (Admin only)"""
    await get_current_user(credentials)
    
    reports = await db.stage1_audit_reports.find({}, {"_id": 0}).to_list(1000)
    return reports

@api_router.get("/stage1-audit-reports/{report_id}")
async def get_stage1_audit_report(report_id: str, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Get a specific Stage 1 Audit Report (Admin only)"""
    await get_current_user(credentials)
    
    report = await db.stage1_audit_reports.find_one({"id": report_id}, {"_id": 0})
    if not report:
        raise HTTPException(status_code=404, detail="Audit report not found")
    
    return report

@api_router.put("/stage1-audit-reports/{report_id}")
async def update_stage1_audit_report(report_id: str, data: Stage1AuditReportUpdate, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Update Stage 1 Audit Report (Admin only)"""
    await get_current_user(credentials)
    
    report = await db.stage1_audit_reports.find_one({"id": report_id})
    if not report:
        raise HTTPException(status_code=404, detail="Audit report not found")
    
    update_data = {}
    for field, value in data.dict().items():
        if value is not None:
            update_data[field] = value
    
    update_data["updated_at"] = datetime.now(timezone.utc)
    
    await db.stage1_audit_reports.update_one(
        {"id": report_id},
        {"$set": update_data}
    )
    
    return {"message": "Audit report updated successfully"}

@api_router.delete("/stage1-audit-reports/{report_id}")
async def delete_stage1_audit_report(report_id: str, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Delete Stage 1 Audit Report (Admin only)"""
    await get_current_user(credentials)
    
    result = await db.stage1_audit_reports.delete_one({"id": report_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Audit report not found")
    
    return {"message": "Audit report deleted successfully"}

@api_router.post("/stage1-audit-reports/{report_id}/complete")
async def complete_stage1_audit_report(report_id: str, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Mark Stage 1 Audit Report as completed (Admin only)"""
    await get_current_user(credentials)
    
    report = await db.stage1_audit_reports.find_one({"id": report_id})
    if not report:
        raise HTTPException(status_code=404, detail="Audit report not found")
    
    # Validate report has required fields
    if not report.get('recommendation'):
        raise HTTPException(status_code=400, detail="Recommendation is required to complete the report")
    
    await db.stage1_audit_reports.update_one(
        {"id": report_id},
        {"$set": {
            "status": "completed",
            "completed_date": datetime.now(timezone.utc).strftime('%Y-%m-%d'),
            "updated_at": datetime.now(timezone.utc)
        }}
    )
    
    return {"message": "Audit report marked as completed"}

@api_router.post("/stage1-audit-reports/{report_id}/approve")
async def approve_stage1_audit_report(report_id: str, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Approve Stage 1 Audit Report (Admin only)"""
    user = await get_current_user(credentials)
    
    report = await db.stage1_audit_reports.find_one({"id": report_id})
    if not report:
        raise HTTPException(status_code=404, detail="Audit report not found")
    
    if report.get('status') != 'completed':
        raise HTTPException(status_code=400, detail="Report must be completed before approval")
    
    await db.stage1_audit_reports.update_one(
        {"id": report_id},
        {"$set": {
            "status": "approved",
            "approved_by": user.get('email', 'Admin'),
            "approved_date": datetime.now(timezone.utc).strftime('%Y-%m-%d'),
            "updated_at": datetime.now(timezone.utc)
        }}
    )
    
    return {"message": "Audit report approved"}

@api_router.get("/stage1-audit-reports/{report_id}/pdf")
async def get_stage1_audit_report_pdf(report_id: str, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Generate PDF for Stage 1 Audit Report (Admin only)"""
    await get_current_user(credentials)
    
    report = await db.stage1_audit_reports.find_one({"id": report_id}, {"_id": 0})
    if not report:
        raise HTTPException(status_code=404, detail="Audit report not found")
    
    try:
        pdf_bytes = generate_stage1_audit_report_pdf(report)
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={
                "Content-Disposition": f"attachment; filename=stage1_audit_report_{report_id[:8]}.pdf"
            }
        )
    except Exception as e:
        logging.error(f"Error generating Stage 1 Audit Report PDF: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error generating PDF: {str(e)}")

# ================= STAGE 2 AUDIT REPORT ENDPOINTS (BACF6-11) =================

@api_router.post("/stage2-audit-reports")
async def create_stage2_audit_report(data: Stage2AuditReportCreate, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Create Stage 2 Audit Report - generated after Stage 2 audit (Admin only)"""
    await get_current_user(credentials)
    
    source_data = None
    stage2_plan = None
    stage1_report = None
    job_order_id = ""
    
    # Get source data from Stage 2 plan
    if data.stage2_plan_id:
        stage2_plan = await db.stage2_audit_plans.find_one({"id": data.stage2_plan_id}, {"_id": 0})
        if not stage2_plan:
            raise HTTPException(status_code=404, detail="Stage 2 Audit Plan not found")
        source_data = stage2_plan
        job_order_id = stage2_plan.get('job_order_id', '')
    
    # Get Stage 1 report if provided
    if data.stage1_report_id:
        stage1_report = await db.stage1_audit_reports.find_one({"id": data.stage1_report_id}, {"_id": 0})
    
    if not source_data:
        raise HTTPException(status_code=400, detail="stage2_plan_id is required")
    
    # Check if report already exists
    existing = await db.stage2_audit_reports.find_one({"stage2_plan_id": data.stage2_plan_id})
    if existing:
        raise HTTPException(status_code=400, detail="Audit report already exists for this Stage 2 plan")
    
    # Get contract review for additional details
    contract_review = None
    contract_review_id = source_data.get('contract_review_id', '')
    if contract_review_id:
        contract_review = await db.contract_reviews.find_one({"id": contract_review_id}, {"_id": 0})
    
    # Default checklist items for Stage 2 (ISO clause based)
    default_checklist = [
        {"requirement": "4 Context of the organization", "status": "C", "comments": ""},
        {"requirement": "4.1 Understanding the organization and its context", "status": "C", "comments": ""},
        {"requirement": "4.2 Understanding needs and expectations of interested parties", "status": "C", "comments": ""},
        {"requirement": "4.3 Determining the scope of the management system", "status": "C", "comments": ""},
        {"requirement": "4.4 Management system and its processes", "status": "C", "comments": ""},
        {"requirement": "5 Leadership", "status": "C", "comments": ""},
        {"requirement": "5.1 Leadership and commitment", "status": "C", "comments": ""},
        {"requirement": "5.2 Policy", "status": "C", "comments": ""},
        {"requirement": "5.3 Organizational roles, responsibilities and authorities", "status": "C", "comments": ""},
        {"requirement": "6 Planning", "status": "C", "comments": ""},
        {"requirement": "6.1 Actions to address risks and opportunities", "status": "C", "comments": ""},
        {"requirement": "6.2 Objectives and planning to achieve them", "status": "C", "comments": ""},
        {"requirement": "7 Support", "status": "C", "comments": ""},
        {"requirement": "7.1 Resources", "status": "C", "comments": ""},
        {"requirement": "7.2 Competence", "status": "C", "comments": ""},
        {"requirement": "7.3 Awareness", "status": "C", "comments": ""},
        {"requirement": "7.4 Communication", "status": "C", "comments": ""},
        {"requirement": "7.5 Documented information", "status": "C", "comments": ""},
        {"requirement": "8 Operation", "status": "C", "comments": ""},
        {"requirement": "8.1 Operational planning and control", "status": "C", "comments": ""},
        {"requirement": "9 Performance evaluation", "status": "C", "comments": ""},
        {"requirement": "9.1 Monitoring, measurement, analysis and evaluation", "status": "C", "comments": ""},
        {"requirement": "9.2 Internal audit", "status": "C", "comments": ""},
        {"requirement": "9.3 Management review", "status": "C", "comments": ""},
        {"requirement": "10 Improvement", "status": "C", "comments": ""},
        {"requirement": "10.1 General", "status": "C", "comments": ""},
        {"requirement": "10.2 Nonconformity and corrective action", "status": "C", "comments": ""},
        {"requirement": "10.3 Continual improvement", "status": "C", "comments": ""},
    ]
    
    # Create report
    report = Stage2AuditReport(
        stage2_plan_id=data.stage2_plan_id,
        stage1_report_id=data.stage1_report_id or "",
        job_order_id=job_order_id,
        audit_program_id=source_data.get('audit_program_id', ''),
        # Organization details (from stage 2 plan or stage 1 report)
        organization_name=source_data.get('organization_name', ''),
        address=source_data.get('address', '') or (stage1_report.get('address', '') if stage1_report else ''),
        site_address=stage1_report.get('site_address', '') if stage1_report else '',
        standards=source_data.get('standards', []),
        num_employees=stage1_report.get('num_employees', '') if stage1_report else '',
        num_shifts=stage1_report.get('num_shifts', '1') if stage1_report else '1',
        email=source_data.get('contact_email', '') or (stage1_report.get('email', '') if stage1_report else ''),
        contact_person=source_data.get('contact_person', '') or (stage1_report.get('contact_person', '') if stage1_report else ''),
        phone=source_data.get('contact_phone', '') or (stage1_report.get('phone', '') if stage1_report else ''),
        scope=source_data.get('scope', '') or (stage1_report.get('scope', '') if stage1_report else ''),
        ea_code=stage1_report.get('ea_code', '') if stage1_report else '',
        exclusions=stage1_report.get('exclusions', '') if stage1_report else '',
        # Audit team
        audit_team={
            "lead_auditor": source_data.get('team_leader', {}).get('name', ''),
            "auditors": [m.get('name', '') for m in source_data.get('team_members', [])],
            "technical_experts": []
        },
        audit_duration=source_data.get('audit_duration', '1'),
        start_date=source_data.get('audit_date_from', ''),
        end_date=source_data.get('audit_date_to', '') or source_data.get('audit_date_from', ''),
        # Default checklist
        checklist_items=default_checklist,
        # Default certification recommendations (all false)
        certification_recommendation={
            "issue_certificate": False,
            "use_logo": False,
            "refuse_certificate": False,
            "post_audit": False,
            "modify_certificate": False,
            "other": False
        }
    )
    
    await db.stage2_audit_reports.insert_one(report.dict())
    
    # Create notification
    await create_notification(
        "stage2_audit_report_created",
        "Stage 2 Audit Report Created",
        f"Stage 2 Audit Report created for {report.organization_name}",
        report.id,
        "stage2_audit_report"
    )
    
    return {
        "id": report.id,
        "message": "Stage 2 Audit Report created successfully"
    }

@api_router.get("/stage2-audit-reports")
async def get_stage2_audit_reports(credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Get all Stage 2 Audit Reports (Admin only)"""
    await get_current_user(credentials)
    
    reports = await db.stage2_audit_reports.find({}, {"_id": 0}).to_list(1000)
    return reports

@api_router.get("/stage2-audit-reports/{report_id}")
async def get_stage2_audit_report(report_id: str, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Get a specific Stage 2 Audit Report (Admin only)"""
    await get_current_user(credentials)
    
    report = await db.stage2_audit_reports.find_one({"id": report_id}, {"_id": 0})
    if not report:
        raise HTTPException(status_code=404, detail="Audit report not found")
    
    return report

@api_router.put("/stage2-audit-reports/{report_id}")
async def update_stage2_audit_report(report_id: str, data: Stage2AuditReportUpdate, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Update Stage 2 Audit Report (Admin only)"""
    await get_current_user(credentials)
    
    report = await db.stage2_audit_reports.find_one({"id": report_id})
    if not report:
        raise HTTPException(status_code=404, detail="Audit report not found")
    
    update_data = {}
    for field, value in data.dict().items():
        if value is not None:
            update_data[field] = value
    
    update_data["updated_at"] = datetime.now(timezone.utc)
    
    await db.stage2_audit_reports.update_one(
        {"id": report_id},
        {"$set": update_data}
    )
    
    return {"message": "Audit report updated successfully"}

@api_router.delete("/stage2-audit-reports/{report_id}")
async def delete_stage2_audit_report(report_id: str, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Delete Stage 2 Audit Report (Admin only)"""
    await get_current_user(credentials)
    
    result = await db.stage2_audit_reports.delete_one({"id": report_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Audit report not found")
    
    return {"message": "Audit report deleted successfully"}

@api_router.post("/stage2-audit-reports/{report_id}/complete")
async def complete_stage2_audit_report(report_id: str, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Mark Stage 2 Audit Report as completed (Admin only)"""
    await get_current_user(credentials)
    
    report = await db.stage2_audit_reports.find_one({"id": report_id})
    if not report:
        raise HTTPException(status_code=404, detail="Audit report not found")
    
    if not report.get('overall_recommendation'):
        raise HTTPException(status_code=400, detail="Overall recommendation is required to complete the report")
    
    await db.stage2_audit_reports.update_one(
        {"id": report_id},
        {"$set": {
            "status": "completed",
            "completed_date": datetime.now(timezone.utc).strftime('%Y-%m-%d'),
            "updated_at": datetime.now(timezone.utc)
        }}
    )
    
    return {"message": "Audit report marked as completed"}

@api_router.post("/stage2-audit-reports/{report_id}/approve")
async def approve_stage2_audit_report(report_id: str, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Approve Stage 2 Audit Report (Admin only)"""
    user = await get_current_user(credentials)
    
    report = await db.stage2_audit_reports.find_one({"id": report_id})
    if not report:
        raise HTTPException(status_code=404, detail="Audit report not found")
    
    if report.get('status') != 'completed':
        raise HTTPException(status_code=400, detail="Report must be completed before approval")
    
    await db.stage2_audit_reports.update_one(
        {"id": report_id},
        {"$set": {
            "status": "approved",
            "approved_by": user.get('email', 'Admin'),
            "approved_date": datetime.now(timezone.utc).strftime('%Y-%m-%d'),
            "updated_at": datetime.now(timezone.utc)
        }}
    )
    
    return {"message": "Audit report approved"}

@api_router.get("/stage2-audit-reports/{report_id}/pdf")
async def get_stage2_audit_report_pdf(report_id: str, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Generate PDF for Stage 2 Audit Report (Admin only)"""
    await get_current_user(credentials)
    
    report = await db.stage2_audit_reports.find_one({"id": report_id}, {"_id": 0})
    if not report:
        raise HTTPException(status_code=404, detail="Audit report not found")
    
    try:
        pdf_bytes = generate_stage2_audit_report_pdf(report)
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={
                "Content-Disposition": f"attachment; filename=stage2_audit_report_{report_id[:8]}.pdf"
            }
        )
    except Exception as e:
        logging.error(f"Error generating Stage 2 Audit Report PDF: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error generating PDF: {str(e)}")

# ================= AUDITOR NOTES ENDPOINTS (BACF6-12) =================

@api_router.post("/auditor-notes")
async def create_auditor_notes(
    data: AuditorNotesCreate,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Create new Auditor Notes - can be from Stage 2 report or independently"""
    await get_current_user(credentials)
    
    # Initialize auditor notes
    notes = AuditorNotes()
    
    # If creating from Stage 2 report, pull data from there
    if data.stage2_report_id:
        report = await db.stage2_audit_reports.find_one({"id": data.stage2_report_id}, {"_id": 0})
        if not report:
            raise HTTPException(status_code=404, detail="Stage 2 Audit Report not found")
        
        notes.stage2_report_id = data.stage2_report_id
        notes.stage1_report_id = report.get("stage1_report_id", "")
        notes.job_order_id = report.get("job_order_id", "")
        notes.audit_program_id = report.get("audit_program_id", "")
        notes.client_name = report.get("organization_name", "")
        notes.location = report.get("address", "") or report.get("site_address", "")
        notes.standards = report.get("standards", [])
        notes.audit_type = "Stage 2"
        notes.audit_date = report.get("end_date", "") or report.get("start_date", "")
        
        # Get auditor info from audit team
        audit_team = report.get("audit_team", {})
        if audit_team:
            notes.auditor_name = audit_team.get("lead_auditor", "")
            notes.auditor_id = audit_team.get("lead_auditor_id", "")
    else:
        # Use provided data
        notes.client_name = data.client_name
        notes.location = data.location
        notes.standards = data.standards
        notes.auditor_name = data.auditor_name
        notes.audit_type = data.audit_type
        notes.audit_date = data.audit_date
        notes.department = data.department
    
    # Save to database
    await db.auditor_notes.insert_one(notes.model_dump())
    
    return notes.model_dump()

@api_router.get("/auditor-notes")
async def get_auditor_notes_list(
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Get all auditor notes"""
    await get_current_user(credentials)
    
    notes_list = await db.auditor_notes.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return notes_list

@api_router.get("/auditor-notes/{notes_id}")
async def get_auditor_notes(
    notes_id: str,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Get specific auditor notes by ID"""
    await get_current_user(credentials)
    
    notes = await db.auditor_notes.find_one({"id": notes_id}, {"_id": 0})
    if not notes:
        raise HTTPException(status_code=404, detail="Auditor Notes not found")
    
    return notes

@api_router.put("/auditor-notes/{notes_id}")
async def update_auditor_notes(
    notes_id: str,
    data: AuditorNotesUpdate,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Update auditor notes"""
    await get_current_user(credentials)
    
    existing = await db.auditor_notes.find_one({"id": notes_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Auditor Notes not found")
    
    # Build update dict with only provided fields
    update_data = {"updated_at": datetime.now(timezone.utc)}
    
    if data.client_name is not None:
        update_data["client_name"] = data.client_name
    if data.location is not None:
        update_data["location"] = data.location
    if data.standards is not None:
        update_data["standards"] = data.standards
    if data.auditor_name is not None:
        update_data["auditor_name"] = data.auditor_name
    if data.audit_type is not None:
        update_data["audit_type"] = data.audit_type
    if data.audit_date is not None:
        update_data["audit_date"] = data.audit_date
    if data.department is not None:
        update_data["department"] = data.department
    if data.notes is not None:
        update_data["notes"] = data.notes
    if data.notes_ar is not None:
        update_data["notes_ar"] = data.notes_ar
    
    await db.auditor_notes.update_one(
        {"id": notes_id},
        {"$set": update_data}
    )
    
    return await db.auditor_notes.find_one({"id": notes_id}, {"_id": 0})

@api_router.delete("/auditor-notes/{notes_id}")
async def delete_auditor_notes(
    notes_id: str,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Delete auditor notes"""
    await get_current_user(credentials)
    
    result = await db.auditor_notes.delete_one({"id": notes_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Auditor Notes not found")
    
    return {"message": "Auditor Notes deleted"}

@api_router.post("/auditor-notes/{notes_id}/complete")
async def complete_auditor_notes(
    notes_id: str,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Mark auditor notes as completed"""
    await get_current_user(credentials)
    
    existing = await db.auditor_notes.find_one({"id": notes_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Auditor Notes not found")
    
    await db.auditor_notes.update_one(
        {"id": notes_id},
        {"$set": {
            "status": "completed",
            "completed_at": datetime.now(timezone.utc),
            "updated_at": datetime.now(timezone.utc)
        }}
    )
    
    return {"message": "Auditor Notes marked as completed"}

@api_router.get("/auditor-notes/{notes_id}/pdf")
async def get_auditor_notes_pdf(
    notes_id: str,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Generate PDF for Auditor Notes"""
    await get_current_user(credentials)
    
    notes = await db.auditor_notes.find_one({"id": notes_id}, {"_id": 0})
    if not notes:
        raise HTTPException(status_code=404, detail="Auditor Notes not found")
    
    try:
        pdf_path = generate_auditor_notes_pdf(notes)
        return FileResponse(
            pdf_path,
            media_type="application/pdf",
            filename=f"auditor_notes_{notes_id[:8]}.pdf"
        )
    except Exception as e:
        logging.error(f"Error generating Auditor Notes PDF: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error generating PDF: {str(e)}")

# ================= NONCONFORMITY REPORT ENDPOINTS (BACF6-13) =================

@api_router.post("/nonconformity-reports")
async def create_nonconformity_report(
    data: NonconformityReportCreate,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Create new Nonconformity Report - from Stage 2 report or independently"""
    await get_current_user(credentials)
    
    report = NonconformityReport()
    
    # If creating from Stage 2 report, pull data from there
    if data.stage2_report_id:
        stage2_report = await db.stage2_audit_reports.find_one({"id": data.stage2_report_id}, {"_id": 0})
        if not stage2_report:
            raise HTTPException(status_code=404, detail="Stage 2 Audit Report not found")
        
        report.stage2_report_id = data.stage2_report_id
        report.job_order_id = stage2_report.get("job_order_id", "")
        report.audit_program_id = stage2_report.get("audit_program_id", "")
        report.client_name = stage2_report.get("organization_name", "")
        report.standards = stage2_report.get("standards", [])
        report.audit_type = stage2_report.get("audit_type", "Stage 2")
        report.audit_date = stage2_report.get("audit_date_to", "") or stage2_report.get("end_date", "")
        
        # Get lead auditor from audit team
        team_leader = stage2_report.get("team_leader", {})
        report.lead_auditor = team_leader.get("name", "")
        
        # Import nonconformities from Stage 2 report if any
        ncs = stage2_report.get("nonconformities", [])
        for nc in ncs:
            nc_item = {
                "id": str(uuid.uuid4()),
                "standard_clause": nc.get("clause", ""),
                "description": nc.get("description", ""),
                "description_ar": nc.get("description_ar", ""),
                "nc_type": "major" if nc.get("rating", 1) == 2 else "minor",
                "status": "open"
            }
            report.nonconformities.append(nc_item)
        
        # Calculate counts
        report.total_major = len([n for n in report.nonconformities if n.get("nc_type") == "major"])
        report.total_minor = len([n for n in report.nonconformities if n.get("nc_type") == "minor"])
    else:
        # Use provided data
        report.client_name = data.client_name
        report.certificate_no = data.certificate_no
        report.standards = data.standards
        report.audit_type = data.audit_type
        report.audit_date = data.audit_date
        report.lead_auditor = data.lead_auditor
        report.management_representative = data.management_representative
    
    # Initialize verification options
    report.verification_options = {
        "corrections_appropriate": False,
        "corrections_verified": False,
        "verify_next_audit": False,
        "re_audit_performed": False
    }
    
    await db.nonconformity_reports.insert_one(report.model_dump())
    
    return report.model_dump()

@api_router.get("/nonconformity-reports")
async def get_nonconformity_reports(
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Get all nonconformity reports"""
    await get_current_user(credentials)
    
    reports = await db.nonconformity_reports.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return reports

@api_router.get("/nonconformity-reports/{report_id}")
async def get_nonconformity_report(
    report_id: str,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Get specific nonconformity report by ID"""
    await get_current_user(credentials)
    
    report = await db.nonconformity_reports.find_one({"id": report_id}, {"_id": 0})
    if not report:
        raise HTTPException(status_code=404, detail="Nonconformity Report not found")
    
    return report

@api_router.put("/nonconformity-reports/{report_id}")
async def update_nonconformity_report(
    report_id: str,
    data: NonconformityReportUpdate,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Update nonconformity report"""
    await get_current_user(credentials)
    
    existing = await db.nonconformity_reports.find_one({"id": report_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Nonconformity Report not found")
    
    update_data = {"updated_at": datetime.now(timezone.utc)}
    
    if data.client_name is not None:
        update_data["client_name"] = data.client_name
    if data.certificate_no is not None:
        update_data["certificate_no"] = data.certificate_no
    if data.standards is not None:
        update_data["standards"] = data.standards
    if data.audit_type is not None:
        update_data["audit_type"] = data.audit_type
    if data.audit_date is not None:
        update_data["audit_date"] = data.audit_date
    if data.lead_auditor is not None:
        update_data["lead_auditor"] = data.lead_auditor
    if data.management_representative is not None:
        update_data["management_representative"] = data.management_representative
    if data.nonconformities is not None:
        update_data["nonconformities"] = data.nonconformities
        # Recalculate counts
        update_data["total_major"] = len([n for n in data.nonconformities if n.get("nc_type") == "major"])
        update_data["total_minor"] = len([n for n in data.nonconformities if n.get("nc_type") == "minor"])
        update_data["closed_count"] = len([n for n in data.nonconformities if n.get("status") == "closed"])
    if data.submission_deadline is not None:
        update_data["submission_deadline"] = data.submission_deadline
    if data.verification_options is not None:
        update_data["verification_options"] = data.verification_options
    if data.management_rep_date is not None:
        update_data["management_rep_date"] = data.management_rep_date
    if data.audit_team_leader_date is not None:
        update_data["audit_team_leader_date"] = data.audit_team_leader_date
    if data.evidence_submission_date is not None:
        update_data["evidence_submission_date"] = data.evidence_submission_date
    if data.final_date is not None:
        update_data["final_date"] = data.final_date
    
    await db.nonconformity_reports.update_one(
        {"id": report_id},
        {"$set": update_data}
    )
    
    return await db.nonconformity_reports.find_one({"id": report_id}, {"_id": 0})

@api_router.delete("/nonconformity-reports/{report_id}")
async def delete_nonconformity_report(
    report_id: str,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Delete nonconformity report"""
    await get_current_user(credentials)
    
    result = await db.nonconformity_reports.delete_one({"id": report_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Nonconformity Report not found")
    
    return {"message": "Nonconformity Report deleted"}

@api_router.post("/nonconformity-reports/{report_id}/add-nc")
async def add_nonconformity(
    report_id: str,
    nc_data: NonconformityItem,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Add a new nonconformity to the report"""
    await get_current_user(credentials)
    
    existing = await db.nonconformity_reports.find_one({"id": report_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Nonconformity Report not found")
    
    nc_dict = nc_data.model_dump()
    ncs = existing.get("nonconformities", [])
    ncs.append(nc_dict)
    
    # Recalculate counts
    total_major = len([n for n in ncs if n.get("nc_type") == "major"])
    total_minor = len([n for n in ncs if n.get("nc_type") == "minor"])
    closed_count = len([n for n in ncs if n.get("status") == "closed"])
    
    await db.nonconformity_reports.update_one(
        {"id": report_id},
        {"$set": {
            "nonconformities": ncs,
            "total_major": total_major,
            "total_minor": total_minor,
            "closed_count": closed_count,
            "updated_at": datetime.now(timezone.utc)
        }}
    )
    
    return {"message": "Nonconformity added", "nc_id": nc_dict["id"]}

@api_router.put("/nonconformity-reports/{report_id}/nc/{nc_id}")
async def update_nonconformity(
    report_id: str,
    nc_id: str,
    nc_data: NonconformityItem,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Update a specific nonconformity in the report"""
    await get_current_user(credentials)
    
    existing = await db.nonconformity_reports.find_one({"id": report_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Nonconformity Report not found")
    
    ncs = existing.get("nonconformities", [])
    updated = False
    
    for i, nc in enumerate(ncs):
        if nc.get("id") == nc_id:
            nc_dict = nc_data.model_dump()
            nc_dict["id"] = nc_id  # Preserve original ID
            ncs[i] = nc_dict
            updated = True
            break
    
    if not updated:
        raise HTTPException(status_code=404, detail="Nonconformity not found")
    
    # Recalculate counts
    total_major = len([n for n in ncs if n.get("nc_type") == "major"])
    total_minor = len([n for n in ncs if n.get("nc_type") == "minor"])
    closed_count = len([n for n in ncs if n.get("status") == "closed"])
    
    await db.nonconformity_reports.update_one(
        {"id": report_id},
        {"$set": {
            "nonconformities": ncs,
            "total_major": total_major,
            "total_minor": total_minor,
            "closed_count": closed_count,
            "updated_at": datetime.now(timezone.utc)
        }}
    )
    
    return {"message": "Nonconformity updated"}

@api_router.delete("/nonconformity-reports/{report_id}/nc/{nc_id}")
async def delete_nonconformity(
    report_id: str,
    nc_id: str,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Delete a specific nonconformity from the report"""
    await get_current_user(credentials)
    
    existing = await db.nonconformity_reports.find_one({"id": report_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Nonconformity Report not found")
    
    ncs = existing.get("nonconformities", [])
    original_len = len(ncs)
    ncs = [nc for nc in ncs if nc.get("id") != nc_id]
    
    if len(ncs) == original_len:
        raise HTTPException(status_code=404, detail="Nonconformity not found")
    
    # Recalculate counts
    total_major = len([n for n in ncs if n.get("nc_type") == "major"])
    total_minor = len([n for n in ncs if n.get("nc_type") == "minor"])
    closed_count = len([n for n in ncs if n.get("status") == "closed"])
    
    await db.nonconformity_reports.update_one(
        {"id": report_id},
        {"$set": {
            "nonconformities": ncs,
            "total_major": total_major,
            "total_minor": total_minor,
            "closed_count": closed_count,
            "updated_at": datetime.now(timezone.utc)
        }}
    )
    
    return {"message": "Nonconformity deleted"}

@api_router.post("/nonconformity-reports/{report_id}/close")
async def close_nonconformity_report(
    report_id: str,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Close the nonconformity report"""
    await get_current_user(credentials)
    
    existing = await db.nonconformity_reports.find_one({"id": report_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Nonconformity Report not found")
    
    await db.nonconformity_reports.update_one(
        {"id": report_id},
        {"$set": {
            "status": "closed",
            "final_date": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
            "updated_at": datetime.now(timezone.utc)
        }}
    )
    
    return {"message": "Nonconformity Report closed"}

@api_router.get("/nonconformity-reports/{report_id}/pdf")
async def get_nonconformity_report_pdf(
    report_id: str,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Generate PDF for Nonconformity Report"""
    await get_current_user(credentials)
    
    report = await db.nonconformity_reports.find_one({"id": report_id}, {"_id": 0})
    if not report:
        raise HTTPException(status_code=404, detail="Nonconformity Report not found")
    
    try:
        pdf_bytes = generate_nonconformity_report_pdf(report)
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={
                "Content-Disposition": f"attachment; filename=nonconformity_report_{report_id[:8]}.pdf"
            }
        )
    except Exception as e:
        logging.error(f"Error generating Nonconformity Report PDF: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error generating PDF: {str(e)}")

# ================= CERTIFICATE DATA ENDPOINTS (BACF6-14) =================

@api_router.post("/certificate-data")
async def create_certificate_data(
    data: CertificateDataCreate,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Create new Certificate Data - from NC report, Stage 2 report, or independently"""
    await get_current_user(credentials)
    
    cert_data = CertificateData()
    
    # If creating from NC report, pull data from there
    if data.nc_report_id:
        nc_report = await db.nonconformity_reports.find_one({"id": data.nc_report_id}, {"_id": 0})
        if not nc_report:
            raise HTTPException(status_code=404, detail="NC Report not found")
        
        cert_data.nc_report_id = data.nc_report_id
        cert_data.stage2_report_id = nc_report.get("stage2_report_id", "")
        cert_data.client_name = nc_report.get("client_name", "")
        cert_data.standards = nc_report.get("standards", [])
        cert_data.lead_auditor = nc_report.get("lead_auditor", "")
        cert_data.audit_type = nc_report.get("audit_type", "")
        cert_data.audit_date = nc_report.get("audit_date", "")
    
    # If creating from Stage 2 report
    elif data.stage2_report_id:
        report = await db.stage2_audit_reports.find_one({"id": data.stage2_report_id}, {"_id": 0})
        if not report:
            raise HTTPException(status_code=404, detail="Stage 2 Report not found")
        
        cert_data.stage2_report_id = data.stage2_report_id
        cert_data.client_name = report.get("organization_name", "")
        cert_data.standards = report.get("standards", [])
        team_leader = report.get("team_leader", {})
        cert_data.lead_auditor = team_leader.get("name", "")
        cert_data.audit_type = "CA - Certification Audit"
        cert_data.audit_date = report.get("audit_date_to", "") or report.get("end_date", "")
        
        # Try to get scope from related agreement
        if report.get("audit_program_id"):
            program = await db.audit_programs.find_one({"id": report.get("audit_program_id")}, {"_id": 0})
            if program:
                cert_data.agreed_certification_scope = program.get("scope_of_services", "")
    else:
        # Use provided data
        cert_data.client_name = data.client_name
        cert_data.standards = data.standards
        cert_data.lead_auditor = data.lead_auditor
        cert_data.audit_type = data.audit_type
        cert_data.audit_date = data.audit_date
    
    await db.certificate_data.insert_one(cert_data.model_dump())
    
    return cert_data.model_dump()

@api_router.get("/certificate-data")
async def get_certificate_data_list(
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Get all certificate data records"""
    await get_current_user(credentials)
    
    records = await db.certificate_data.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return records

@api_router.get("/certificate-data/{record_id}")
async def get_certificate_data(
    record_id: str,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Get specific certificate data by ID"""
    await get_current_user(credentials)
    
    record = await db.certificate_data.find_one({"id": record_id}, {"_id": 0})
    if not record:
        raise HTTPException(status_code=404, detail="Certificate Data not found")
    
    return record

@api_router.put("/certificate-data/{record_id}")
async def update_certificate_data(
    record_id: str,
    data: CertificateDataUpdate,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Update certificate data"""
    await get_current_user(credentials)
    
    existing = await db.certificate_data.find_one({"id": record_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Certificate Data not found")
    
    update_data = {"updated_at": datetime.now(timezone.utc)}
    
    for field in ["client_name", "standards", "lead_auditor", "audit_type", "audit_date",
                  "agreed_certification_scope", "agreed_certification_scope_ar", "ea_code",
                  "technical_category", "company_data_local", "certification_scope_local",
                  "company_data_english", "certification_scope_english"]:
        value = getattr(data, field, None)
        if value is not None:
            update_data[field] = value
    
    await db.certificate_data.update_one(
        {"id": record_id},
        {"$set": update_data}
    )
    
    return await db.certificate_data.find_one({"id": record_id}, {"_id": 0})

@api_router.delete("/certificate-data/{record_id}")
async def delete_certificate_data(
    record_id: str,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Delete certificate data"""
    await get_current_user(credentials)
    
    result = await db.certificate_data.delete_one({"id": record_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Certificate Data not found")
    
    return {"message": "Certificate Data deleted"}

@api_router.post("/certificate-data/{record_id}/send-to-client")
async def send_certificate_data_to_client(
    record_id: str,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Send certificate data form to client for confirmation"""
    await get_current_user(credentials)
    
    existing = await db.certificate_data.find_one({"id": record_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Certificate Data not found")
    
    await db.certificate_data.update_one(
        {"id": record_id},
        {"$set": {
            "status": "sent_to_client",
            "updated_at": datetime.now(timezone.utc)
        }}
    )
    
    # Generate public link
    base_url = os.environ.get("REACT_APP_BACKEND_URL", "")
    public_url = f"{base_url}/public/certificate-data/{existing['access_token']}"
    
    return {"message": "Sent to client", "public_url": public_url}

# Public endpoint for client confirmation
@api_router.get("/public/certificate-data/{access_token}")
async def get_public_certificate_data(access_token: str):
    """Get certificate data via public access token"""
    record = await db.certificate_data.find_one(
        {"access_token": access_token},
        {"_id": 0, "client_signature": 0, "client_stamp": 0}  # Exclude large data
    )
    
    if not record:
        raise HTTPException(status_code=404, detail="Certificate Data not found")
    
    if record.get("status") not in ["sent_to_client", "client_confirmed"]:
        raise HTTPException(status_code=400, detail="Form not available for confirmation")
    
    return record

@api_router.post("/public/certificate-data/{access_token}/confirm")
async def confirm_certificate_data(
    access_token: str,
    data: CertificateDataClientConfirm
):
    """Client confirms certificate data"""
    existing = await db.certificate_data.find_one({"access_token": access_token})
    
    if not existing:
        raise HTTPException(status_code=404, detail="Certificate Data not found")
    
    if existing.get("status") != "sent_to_client":
        raise HTTPException(status_code=400, detail="Form already confirmed or not available")
    
    await db.certificate_data.update_one(
        {"access_token": access_token},
        {"$set": {
            "client_signature": data.client_signature,
            "client_stamp": data.client_stamp,
            "client_signature_date": data.signature_date,
            "client_confirmed": True,
            "client_confirmed_at": datetime.now(timezone.utc),
            "status": "client_confirmed",
            "updated_at": datetime.now(timezone.utc)
        }}
    )
    
    return {"message": "Certificate data confirmed successfully"}

@api_router.post("/certificate-data/{record_id}/issue-certificate")
async def issue_certificate_from_data(
    record_id: str,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Generate certificate from confirmed certificate data"""
    await get_current_user(credentials)
    
    record = await db.certificate_data.find_one({"id": record_id}, {"_id": 0})
    if not record:
        raise HTTPException(status_code=404, detail="Certificate Data not found")
    
    if not record.get("client_confirmed"):
        raise HTTPException(status_code=400, detail="Certificate data not yet confirmed by client")
    
    # Generate certificate number
    cert_number = await generate_certificate_number()
    issue_date = datetime.now().strftime("%Y-%m-%d")
    expiry_date = (datetime.now() + timedelta(days=3*365)).strftime("%Y-%m-%d")  # 3 year validity
    
    # Update certificate data record
    await db.certificate_data.update_one(
        {"id": record_id},
        {"$set": {
            "certificate_number": cert_number,
            "issue_date": issue_date,
            "expiry_date": expiry_date,
            "certificate_generated": True,
            "status": "certificate_issued",
            "updated_at": datetime.now(timezone.utc)
        }}
    )
    
    # Create certificate record
    certificate = {
        "id": str(uuid.uuid4()),
        "certificate_number": cert_number,
        "certificate_data_id": record_id,
        "organization_name": record.get("client_name", ""),
        "organization_name_ar": record.get("client_name_ar", ""),
        "standards": record.get("standards", []),
        "scope": record.get("certification_scope_english", "") or record.get("agreed_certification_scope", ""),
        "scope_ar": record.get("certification_scope_local", "") or record.get("agreed_certification_scope_ar", ""),
        "issue_date": issue_date,
        "expiry_date": expiry_date,
        "status": "active",
        "lead_auditor": record.get("lead_auditor", ""),
        "created_at": datetime.now(timezone.utc)
    }
    
    await db.certificates.insert_one(certificate)
    
    return {
        "message": "Certificate issued successfully",
        "certificate_number": cert_number,
        "certificate_id": certificate["id"]
    }

@api_router.get("/certificate-data/{record_id}/pdf")
async def get_certificate_data_pdf(
    record_id: str,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Generate PDF for Certificate Data"""
    await get_current_user(credentials)
    
    record = await db.certificate_data.find_one({"id": record_id}, {"_id": 0})
    if not record:
        raise HTTPException(status_code=404, detail="Certificate Data not found")
    
    try:
        pdf_bytes = generate_certificate_data_pdf(record)
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={
                "Content-Disposition": f"attachment; filename=certificate_data_{record_id[:8]}.pdf"
            }
        )
    except Exception as e:
        logging.error(f"Error generating Certificate Data PDF: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error generating PDF: {str(e)}")

# ================= TECHNICAL REVIEW ENDPOINTS (BACF6-15) =================

@api_router.get("/technical-reviews")
async def get_technical_reviews(
    status: Optional[str] = None,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Get all technical reviews"""
    await get_current_user(credentials)
    
    query = {}
    if status and status != 'all':
        query['status'] = status
    
    reviews = await db.technical_reviews.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return reviews

@api_router.post("/technical-reviews")
async def create_technical_review(
    data: TechnicalReviewCreate,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Create a new technical review - from Stage 2 report or independently"""
    await get_current_user(credentials)
    
    # Initialize with default checklist
    checklist_items = [item.copy() for item in DEFAULT_TECHNICAL_REVIEW_CHECKLIST]
    
    review = TechnicalReview(
        stage2_report_id=data.stage2_report_id,
        client_name=data.client_name,
        client_name_ar=data.client_name_ar,
        location=data.location,
        scope=data.scope,
        ea_code=data.ea_code,
        standards=data.standards,
        audit_type=data.audit_type,
        audit_dates=data.audit_dates,
        audit_team_members=data.audit_team_members,
        technical_expert=data.technical_expert,
        checklist_items=checklist_items,
        status="draft"
    )
    
    # If linked to Stage 2 report, auto-fill from linked data
    if data.stage2_report_id:
        stage2_report = await db.stage2_audit_reports.find_one({"id": data.stage2_report_id}, {"_id": 0})
        if stage2_report:
            review.job_order_id = stage2_report.get('job_order_id', '')
            review.audit_program_id = stage2_report.get('audit_program_id', '')
            review.contract_review_id = stage2_report.get('contract_review_id', '')
            review.agreement_id = stage2_report.get('agreement_id', '')
            
            # Fill from stage2 report if not provided
            if not review.client_name:
                review.client_name = stage2_report.get('client_name', '')
            if not review.client_name_ar:
                review.client_name_ar = stage2_report.get('client_name_ar', '')
            if not review.scope:
                review.scope = stage2_report.get('scope', '')
            if not review.standards:
                review.standards = stage2_report.get('standards', [])
            if not review.audit_team_members:
                team = []
                if stage2_report.get('team_leader'):
                    team.append(stage2_report.get('team_leader'))
                team.extend(stage2_report.get('team_members', []))
                review.audit_team_members = team
    
    review_doc = review.model_dump()
    review_doc['created_at'] = review_doc['created_at'].isoformat()
    
    await db.technical_reviews.insert_one(review_doc)
    
    await create_notification(
        notification_type="technical_review_created",
        title="Technical Review Created",
        message=f"Technical review created for {review.client_name}",
        related_id=review.id,
        related_type="technical_review"
    )
    
    return {"message": "Technical review created", "id": review.id}

@api_router.get("/technical-reviews/{review_id}")
async def get_technical_review(
    review_id: str,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Get a specific technical review"""
    await get_current_user(credentials)
    
    review = await db.technical_reviews.find_one({"id": review_id}, {"_id": 0})
    if not review:
        raise HTTPException(status_code=404, detail="Technical review not found")
    
    return review

@api_router.put("/technical-reviews/{review_id}")
async def update_technical_review(
    review_id: str,
    data: TechnicalReviewUpdate,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Update a technical review"""
    await get_current_user(credentials)
    
    existing = await db.technical_reviews.find_one({"id": review_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Technical review not found")
    
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    update_data['updated_at'] = datetime.now(timezone.utc).isoformat()
    
    await db.technical_reviews.update_one(
        {"id": review_id},
        {"$set": update_data}
    )
    
    return {"message": "Technical review updated"}

@api_router.delete("/technical-reviews/{review_id}")
async def delete_technical_review(
    review_id: str,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Delete a technical review"""
    await get_current_user(credentials)
    
    existing = await db.technical_reviews.find_one({"id": review_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Technical review not found")
    
    await db.technical_reviews.delete_one({"id": review_id})
    return {"message": "Technical review deleted"}

@api_router.post("/technical-reviews/{review_id}/submit-review")
async def submit_technical_review(
    review_id: str,
    data: dict,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Submit technical review (add technical reviewer info)"""
    await get_current_user(credentials)
    
    existing = await db.technical_reviews.find_one({"id": review_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Technical review not found")
    
    await db.technical_reviews.update_one(
        {"id": review_id},
        {"$set": {
            "technical_reviewer": data.get('technical_reviewer', ''),
            "review_date": data.get('review_date', datetime.now().strftime("%Y-%m-%d")),
            "review_comments": data.get('review_comments', ''),
            "status": "under_review",
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"message": "Technical review submitted"}

@api_router.post("/technical-reviews/{review_id}/make-decision")
async def make_certification_decision(
    review_id: str,
    data: dict,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Make certification decision"""
    await get_current_user(credentials)
    
    existing = await db.technical_reviews.find_one({"id": review_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Technical review not found")
    
    decision = data.get('certification_decision', '')
    if decision not in ['issue_certificate', 'reject_certificate', 'needs_review']:
        raise HTTPException(status_code=400, detail="Invalid certification decision")
    
    update_data = {
        "certification_decision": decision,
        "decision_comments": data.get('decision_comments', ''),
        "status": "decision_made",
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.technical_reviews.update_one({"id": review_id}, {"$set": update_data})
    
    await create_notification(
        notification_type="certification_decision",
        title="Certification Decision Made",
        message=f"Decision for {existing.get('client_name', '')}: {decision.replace('_', ' ').title()}",
        related_id=review_id,
        related_type="technical_review"
    )
    
    return {"message": f"Certification decision recorded: {decision}"}

@api_router.post("/technical-reviews/{review_id}/approve")
async def approve_technical_review(
    review_id: str,
    data: dict,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Approve technical review and optionally issue certificate"""
    await get_current_user(credentials)
    
    existing = await db.technical_reviews.find_one({"id": review_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Technical review not found")
    
    update_data = {
        "approved_by": data.get('approved_by', ''),
        "approval_date": data.get('approval_date', datetime.now().strftime("%Y-%m-%d")),
        "status": "approved",
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    certificate_result = None
    
    # If decision is to issue certificate, auto-generate it
    if existing.get('certification_decision') == 'issue_certificate':
        # Generate certificate number
        cert_number = await generate_certificate_number()
        
        # Calculate dates
        issue_date = datetime.now().strftime("%Y-%m-%d")
        expiry_date = (datetime.now() + timedelta(days=365*3)).strftime("%Y-%m-%d")
        
        # Generate verification URL and QR
        base_url = os.environ.get('FRONTEND_URL', 'https://audit-workflow-pro-1.preview.emergentagent.com')
        verification_url = f"{base_url}/verify/{cert_number}"
        qr_base64 = get_qr_code_base64(verification_url)
        
        # Create certificate
        certificate = Certificate(
            certificate_number=cert_number,
            contract_id=existing.get('agreement_id', ''),
            audit_id=existing.get('stage2_report_id', ''),
            organization_name=existing.get('client_name', ''),
            organization_name_ar=existing.get('client_name_ar', ''),
            standards=existing.get('standards', []),
            scope=existing.get('scope', ''),
            issue_date=issue_date,
            expiry_date=expiry_date,
            status="active",
            verification_url=verification_url,
            qr_code_data=qr_base64,
            lead_auditor=existing.get('audit_team_members', [''])[0] if existing.get('audit_team_members') else '',
            audit_team=existing.get('audit_team_members', [])
        )
        
        cert_doc = certificate.model_dump()
        cert_doc['created_at'] = cert_doc['created_at'].isoformat()
        
        await db.certificates.insert_one(cert_doc)
        
        # Update technical review with certificate info
        update_data['certificate_id'] = certificate.id
        update_data['certificate_number'] = cert_number
        update_data['status'] = 'certificate_issued'
        
        await create_notification(
            notification_type="certificate_issued",
            title="Certificate Issued",
            message=f"Certificate {cert_number} issued for {existing.get('client_name', '')}",
            related_id=certificate.id,
            related_type="certificate"
        )
        
        certificate_result = {
            "certificate_id": certificate.id,
            "certificate_number": cert_number
        }
    
    await db.technical_reviews.update_one({"id": review_id}, {"$set": update_data})
    
    response = {"message": "Technical review approved", "status": update_data['status']}
    if certificate_result:
        response["certificate"] = certificate_result
    
    return response

@api_router.get("/technical-reviews/{review_id}/pdf")
async def get_technical_review_pdf(
    review_id: str,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Generate PDF for Technical Review"""
    await get_current_user(credentials)
    
    review = await db.technical_reviews.find_one({"id": review_id}, {"_id": 0})
    if not review:
        raise HTTPException(status_code=404, detail="Technical review not found")
    
    try:
        contracts_dir = Path(__file__).parent / "contracts"
        contracts_dir.mkdir(exist_ok=True)
        pdf_path = str(contracts_dir / f"technical_review_{review_id[:8]}.pdf")
        
        generate_technical_review_pdf(review, pdf_path)
        
        return FileResponse(
            pdf_path,
            media_type="application/pdf",
            filename=f"technical_review_{review_id[:8]}.pdf"
        )
    except Exception as e:
        logging.error(f"Error generating Technical Review PDF: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error generating PDF: {str(e)}")

# ================= CUSTOMER FEEDBACK ENDPOINTS (BACF6-16) =================

@api_router.get("/customer-feedback")
async def get_customer_feedback_list(
    status: Optional[str] = None,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Get all customer feedback forms"""
    await get_current_user(credentials)
    
    query = {}
    if status and status != 'all':
        query['status'] = status
    
    feedbacks = await db.customer_feedback.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return feedbacks

@api_router.post("/customer-feedback")
async def create_customer_feedback(
    data: CustomerFeedbackCreate,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Create a new customer feedback form"""
    await get_current_user(credentials)
    
    # Initialize with default questions
    questions = [q.copy() for q in DEFAULT_FEEDBACK_QUESTIONS]
    
    feedback = CustomerFeedback(
        certificate_id=data.certificate_id,
        audit_id=data.audit_id,
        organization_name=data.organization_name,
        organization_name_ar=data.organization_name_ar,
        audit_type=data.audit_type,
        standards=data.standards,
        audit_date=data.audit_date,
        lead_auditor=data.lead_auditor,
        auditor=data.auditor,
        questions=questions,
        status="pending"
    )
    
    feedback_doc = feedback.model_dump()
    feedback_doc['created_at'] = feedback_doc['created_at'].isoformat()
    
    await db.customer_feedback.insert_one(feedback_doc)
    
    await create_notification(
        notification_type="feedback_created",
        title="Customer Feedback Created",
        message=f"Feedback form created for {feedback.organization_name}",
        related_id=feedback.id,
        related_type="customer_feedback"
    )
    
    return {
        "message": "Customer feedback form created",
        "id": feedback.id,
        "access_token": feedback.access_token,
        "public_url": f"/feedback/{feedback.access_token}"
    }

@api_router.get("/customer-feedback/{feedback_id}")
async def get_customer_feedback(
    feedback_id: str,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Get a specific customer feedback"""
    await get_current_user(credentials)
    
    feedback = await db.customer_feedback.find_one({"id": feedback_id}, {"_id": 0})
    if not feedback:
        raise HTTPException(status_code=404, detail="Customer feedback not found")
    
    return feedback

@api_router.put("/customer-feedback/{feedback_id}")
async def update_customer_feedback(
    feedback_id: str,
    data: dict,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Update a customer feedback (admin only)"""
    await get_current_user(credentials)
    
    existing = await db.customer_feedback.find_one({"id": feedback_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Customer feedback not found")
    
    # Recalculate score if questions are updated
    if 'questions' in data:
        score, evaluation = calculate_feedback_score(data['questions'])
        data['overall_score'] = score
        data['evaluation_result'] = evaluation
    
    await db.customer_feedback.update_one(
        {"id": feedback_id},
        {"$set": data}
    )
    
    return {"message": "Customer feedback updated"}

@api_router.delete("/customer-feedback/{feedback_id}")
async def delete_customer_feedback(
    feedback_id: str,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Delete a customer feedback"""
    await get_current_user(credentials)
    
    existing = await db.customer_feedback.find_one({"id": feedback_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Customer feedback not found")
    
    await db.customer_feedback.delete_one({"id": feedback_id})
    return {"message": "Customer feedback deleted"}

@api_router.post("/customer-feedback/{feedback_id}/review")
async def review_customer_feedback(
    feedback_id: str,
    data: dict,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Mark feedback as reviewed by admin"""
    await get_current_user(credentials)
    
    existing = await db.customer_feedback.find_one({"id": feedback_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Customer feedback not found")
    
    await db.customer_feedback.update_one(
        {"id": feedback_id},
        {"$set": {
            "reviewed_by": data.get('reviewed_by', ''),
            "review_date": data.get('review_date', datetime.now().strftime("%Y-%m-%d")),
            "review_comments": data.get('review_comments', ''),
            "status": "reviewed"
        }}
    )
    
    return {"message": "Feedback marked as reviewed"}

@api_router.get("/customer-feedback/{feedback_id}/pdf")
async def get_customer_feedback_pdf(
    feedback_id: str,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Generate PDF for Customer Feedback"""
    await get_current_user(credentials)
    
    feedback = await db.customer_feedback.find_one({"id": feedback_id}, {"_id": 0})
    if not feedback:
        raise HTTPException(status_code=404, detail="Customer feedback not found")
    
    try:
        contracts_dir = Path(__file__).parent / "contracts"
        contracts_dir.mkdir(exist_ok=True)
        pdf_path = str(contracts_dir / f"customer_feedback_{feedback_id[:8]}.pdf")
        
        generate_customer_feedback_pdf(feedback, pdf_path)
        
        return FileResponse(
            pdf_path,
            media_type="application/pdf",
            filename=f"customer_feedback_{feedback_id[:8]}.pdf"
        )
    except Exception as e:
        logging.error(f"Error generating Customer Feedback PDF: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error generating PDF: {str(e)}")

# Public endpoint for client to submit feedback
@api_router.get("/public/feedback/{access_token}")
async def get_public_feedback_form(access_token: str):
    """Get feedback form for client to fill (public, no auth required)"""
    feedback = await db.customer_feedback.find_one(
        {"access_token": access_token},
        {"_id": 0}
    )
    
    if not feedback:
        raise HTTPException(status_code=404, detail="Feedback form not found")
    
    if feedback.get('status') == 'submitted':
        raise HTTPException(status_code=400, detail="This feedback has already been submitted")
    
    # Return only necessary fields for public view
    return {
        "id": feedback['id'],
        "organization_name": feedback.get('organization_name', ''),
        "organization_name_ar": feedback.get('organization_name_ar', ''),
        "audit_type": feedback.get('audit_type', ''),
        "standards": feedback.get('standards', []),
        "audit_date": feedback.get('audit_date', ''),
        "lead_auditor": feedback.get('lead_auditor', ''),
        "auditor": feedback.get('auditor', ''),
        "questions": feedback.get('questions', []),
        "status": feedback.get('status', 'pending')
    }

@api_router.post("/public/feedback/{access_token}/submit")
async def submit_public_feedback(access_token: str, data: CustomerFeedbackSubmit):
    """Submit feedback from client (public, no auth required)"""
    feedback = await db.customer_feedback.find_one({"access_token": access_token})
    
    if not feedback:
        raise HTTPException(status_code=404, detail="Feedback form not found")
    
    if feedback.get('status') == 'submitted':
        raise HTTPException(status_code=400, detail="This feedback has already been submitted")
    
    # Update questions with ratings
    questions = feedback.get('questions', [])
    for submitted_q in data.questions:
        q_idx = submitted_q.get('index', -1)
        if 0 <= q_idx < len(questions):
            questions[q_idx]['rating'] = submitted_q.get('rating')
    
    # Calculate score
    score, evaluation = calculate_feedback_score(questions)
    
    # Update feedback
    await db.customer_feedback.update_one(
        {"access_token": access_token},
        {"$set": {
            "questions": questions,
            "want_same_team": data.want_same_team,
            "suggestions": data.suggestions,
            "respondent_name": data.respondent_name,
            "respondent_designation": data.respondent_designation,
            "submission_date": datetime.now().strftime("%Y-%m-%d"),
            "overall_score": score,
            "evaluation_result": evaluation,
            "status": "submitted",
            "submitted_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    await create_notification(
        notification_type="feedback_submitted",
        title="Customer Feedback Submitted",
        message=f"Feedback received from {feedback.get('organization_name', '')} - Score: {score}%",
        related_id=feedback['id'],
        related_type="customer_feedback"
    )
    
    return {
        "message": "Feedback submitted successfully",
        "overall_score": score,
        "evaluation_result": evaluation
    }

# ================= PRE-TRANSFER REVIEW ENDPOINTS (BACF6-17) =================

@api_router.get("/pre-transfer-reviews")
async def get_pre_transfer_reviews(
    status: Optional[str] = None,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Get all pre-transfer reviews"""
    await get_current_user(credentials)
    
    query = {}
    if status and status != 'all':
        query['status'] = status
    
    reviews = await db.pre_transfer_reviews.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return reviews

@api_router.post("/pre-transfer-reviews")
async def create_pre_transfer_review(
    data: PreTransferReviewCreate,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Create a new pre-transfer review"""
    await get_current_user(credentials)
    
    # Initialize with default checklist
    checklist = DEFAULT_PRETRANSFER_CHECKLIST.copy()
    
    review = PreTransferReview(
        client_name=data.client_name,
        client_name_ar=data.client_name_ar,
        client_address=data.client_address,
        client_phone=data.client_phone,
        enquiry_reference=data.enquiry_reference,
        transfer_reason=data.transfer_reason,
        existing_cb=data.existing_cb,
        certificate_number=data.certificate_number,
        validity=data.validity,
        scope=data.scope,
        sites=data.sites,
        eac_code=data.eac_code,
        standards=data.standards,
        checklist=checklist,
        status="draft"
    )
    
    review_doc = review.model_dump()
    review_doc['created_at'] = review_doc['created_at'].isoformat()
    
    await db.pre_transfer_reviews.insert_one(review_doc)
    
    await create_notification(
        notification_type="pre_transfer_created",
        title="Pre-Transfer Review Created",
        message=f"Pre-transfer review created for {review.client_name}",
        related_id=review.id,
        related_type="pre_transfer_review"
    )
    
    return {"message": "Pre-transfer review created", "id": review.id}

@api_router.get("/pre-transfer-reviews/{review_id}")
async def get_pre_transfer_review(
    review_id: str,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Get a specific pre-transfer review"""
    await get_current_user(credentials)
    
    review = await db.pre_transfer_reviews.find_one({"id": review_id}, {"_id": 0})
    if not review:
        raise HTTPException(status_code=404, detail="Pre-transfer review not found")
    
    return review

@api_router.put("/pre-transfer-reviews/{review_id}")
async def update_pre_transfer_review(
    review_id: str,
    data: PreTransferReviewUpdate,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Update a pre-transfer review"""
    await get_current_user(credentials)
    
    existing = await db.pre_transfer_reviews.find_one({"id": review_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Pre-transfer review not found")
    
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    update_data['updated_at'] = datetime.now(timezone.utc).isoformat()
    
    await db.pre_transfer_reviews.update_one(
        {"id": review_id},
        {"$set": update_data}
    )
    
    return {"message": "Pre-transfer review updated"}

@api_router.delete("/pre-transfer-reviews/{review_id}")
async def delete_pre_transfer_review(
    review_id: str,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Delete a pre-transfer review"""
    await get_current_user(credentials)
    
    existing = await db.pre_transfer_reviews.find_one({"id": review_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Pre-transfer review not found")
    
    await db.pre_transfer_reviews.delete_one({"id": review_id})
    return {"message": "Pre-transfer review deleted"}

@api_router.post("/pre-transfer-reviews/{review_id}/make-decision")
async def make_transfer_decision(
    review_id: str,
    data: dict,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Make transfer decision (approve/reject)"""
    await get_current_user(credentials)
    
    existing = await db.pre_transfer_reviews.find_one({"id": review_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Pre-transfer review not found")
    
    decision = data.get('transfer_decision', '')
    if decision not in ['approved', 'rejected']:
        raise HTTPException(status_code=400, detail="Invalid decision. Use 'approved' or 'rejected'")
    
    update_data = {
        "transfer_decision": decision,
        "decision_reason": data.get('decision_reason', ''),
        "reviewed_by": data.get('reviewed_by', ''),
        "review_date": data.get('review_date', datetime.now().strftime("%Y-%m-%d")),
        "approved_by": data.get('approved_by', ''),
        "approval_date": data.get('approval_date', datetime.now().strftime("%Y-%m-%d")),
        "status": "decision_made",
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.pre_transfer_reviews.update_one({"id": review_id}, {"$set": update_data})
    
    await create_notification(
        notification_type="transfer_decision",
        title="Transfer Decision Made",
        message=f"Transfer for {existing.get('client_name', '')} has been {decision}",
        related_id=review_id,
        related_type="pre_transfer_review"
    )
    
    return {"message": f"Transfer decision recorded: {decision}"}

@api_router.get("/pre-transfer-reviews/{review_id}/pdf")
async def get_pre_transfer_review_pdf(
    review_id: str,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Generate PDF for Pre-Transfer Review"""
    await get_current_user(credentials)
    
    review = await db.pre_transfer_reviews.find_one({"id": review_id}, {"_id": 0})
    if not review:
        raise HTTPException(status_code=404, detail="Pre-transfer review not found")
    
    try:
        contracts_dir = Path(__file__).parent / "contracts"
        contracts_dir.mkdir(exist_ok=True)
        pdf_path = str(contracts_dir / f"pre_transfer_review_{review_id[:8]}.pdf")
        
        generate_pre_transfer_review_pdf(review, pdf_path)
        
        return FileResponse(
            pdf_path,
            media_type="application/pdf",
            filename=f"pre_transfer_review_{review_id[:8]}.pdf"
        )
    except Exception as e:
        logging.error(f"Error generating Pre-Transfer Review PDF: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error generating PDF: {str(e)}")

# ================= CERTIFIED CLIENTS REGISTRY ENDPOINTS (BAC-F6-19) =================

async def get_next_serial_number():
    """Get next serial number for certified client registry"""
    last_record = await db.certified_clients.find_one(
        {},
        sort=[("serial_number", -1)]
    )
    if last_record:
        return last_record.get('serial_number', 0) + 1
    return 1

@api_router.get("/certified-clients")
async def get_certified_clients(
    status: str = None,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Get all certified client records"""
    await get_current_user(credentials)
    
    query = {}
    if status and status != 'all':
        query['status'] = status
    
    clients = await db.certified_clients.find(query, {"_id": 0}).sort("serial_number", 1).to_list(1000)
    
    # Check for expired certificates and update status
    today = datetime.now().strftime("%Y-%m-%d")
    for client in clients:
        if client.get('status') == 'active' and client.get('expiry_date') and client.get('expiry_date') < today:
            await db.certified_clients.update_one(
                {"id": client['id']},
                {"$set": {"status": "expired", "updated_at": datetime.now(timezone.utc).isoformat()}}
            )
            client['status'] = 'expired'
    
    return clients

@api_router.post("/certified-clients")
async def create_certified_client(
    data: CertifiedClientCreate,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Create a new certified client record"""
    await get_current_user(credentials)
    
    serial_number = await get_next_serial_number()
    
    client = CertifiedClient(
        serial_number=serial_number,
        client_name=data.client_name,
        client_name_ar=data.client_name_ar,
        address=data.address,
        address_ar=data.address_ar,
        contact_person=data.contact_person,
        contact_number=data.contact_number,
        scope=data.scope,
        scope_ar=data.scope_ar,
        accreditation=data.accreditation,
        ea_code=data.ea_code,
        certificate_number=data.certificate_number,
        issue_date=data.issue_date,
        expiry_date=data.expiry_date,
        surveillance_1_date=data.surveillance_1_date,
        surveillance_2_date=data.surveillance_2_date,
        recertification_date=data.recertification_date,
        linked_certificate_id=data.linked_certificate_id
    )
    
    client_doc = client.model_dump()
    client_doc['created_at'] = datetime.now(timezone.utc).isoformat()
    
    await db.certified_clients.insert_one(client_doc)
    
    await create_notification(
        notification_type="certified_client_added",
        title="New Certified Client Added",
        message=f"New client '{data.client_name}' added to registry with certificate {data.certificate_number}",
        related_id=client.id,
        related_type="certified_client"
    )
    
    return {k: v for k, v in client_doc.items() if k != '_id'}

@api_router.get("/certified-clients/{client_id}")
async def get_certified_client(
    client_id: str,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Get a specific certified client record"""
    await get_current_user(credentials)
    
    client = await db.certified_clients.find_one({"id": client_id}, {"_id": 0})
    if not client:
        raise HTTPException(status_code=404, detail="Certified client not found")
    
    return client

@api_router.put("/certified-clients/{client_id}")
async def update_certified_client(
    client_id: str,
    data: CertifiedClientUpdate,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Update a certified client record"""
    await get_current_user(credentials)
    
    existing = await db.certified_clients.find_one({"id": client_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Certified client not found")
    
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    update_data['updated_at'] = datetime.now(timezone.utc).isoformat()
    
    await db.certified_clients.update_one(
        {"id": client_id},
        {"$set": update_data}
    )
    
    return {"message": "Certified client updated successfully"}

@api_router.delete("/certified-clients/{client_id}")
async def delete_certified_client(
    client_id: str,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Delete a certified client record"""
    await get_current_user(credentials)
    
    existing = await db.certified_clients.find_one({"id": client_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Certified client not found")
    
    await db.certified_clients.delete_one({"id": client_id})
    return {"message": "Certified client deleted successfully"}

@api_router.get("/certified-clients/stats/overview")
async def get_certified_clients_stats(
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Get statistics for certified clients registry"""
    await get_current_user(credentials)
    
    clients = await db.certified_clients.find({}, {"_id": 0}).to_list(1000)
    
    today = datetime.now().strftime("%Y-%m-%d")
    thirty_days_later = (datetime.now() + timedelta(days=30)).strftime("%Y-%m-%d")
    
    active = len([c for c in clients if c.get('status') == 'active'])
    suspended = len([c for c in clients if c.get('status') == 'suspended'])
    expired = len([c for c in clients if c.get('status') == 'expired'])
    expiring_soon = len([c for c in clients if c.get('status') == 'active' and c.get('expiry_date') and today < c.get('expiry_date') <= thirty_days_later])
    
    # Upcoming surveillance audits
    surveillance_1_upcoming = len([c for c in clients if c.get('surveillance_1_date') and today <= c.get('surveillance_1_date') <= thirty_days_later])
    surveillance_2_upcoming = len([c for c in clients if c.get('surveillance_2_date') and today <= c.get('surveillance_2_date') <= thirty_days_later])
    
    return {
        "total": len(clients),
        "active": active,
        "suspended": suspended,
        "expired": expired,
        "expiring_soon": expiring_soon,
        "surveillance_1_upcoming": surveillance_1_upcoming,
        "surveillance_2_upcoming": surveillance_2_upcoming
    }

@api_router.post("/certified-clients/sync-from-certificates")
async def sync_from_certificates(
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Sync certified clients registry from existing certificates in the system"""
    await get_current_user(credentials)
    
    # Get all certificates not already in the registry
    certificates = await db.certificates.find({}, {"_id": 0}).to_list(1000)
    existing_links = await db.certified_clients.distinct("linked_certificate_id")
    
    added_count = 0
    for cert in certificates:
        if cert.get('id') not in existing_links and cert.get('certificate_number'):
            serial_number = await get_next_serial_number()
            
            client = CertifiedClient(
                serial_number=serial_number,
                client_name=cert.get('organization_name', ''),
                client_name_ar=cert.get('organization_name_ar', ''),
                scope=cert.get('scope', ''),
                scope_ar=cert.get('scope_ar', ''),
                accreditation=cert.get('standards', []),
                certificate_number=cert.get('certificate_number', ''),
                issue_date=cert.get('issue_date', ''),
                expiry_date=cert.get('expiry_date', ''),
                status=cert.get('status', 'active'),
                linked_certificate_id=cert.get('id', '')
            )
            
            client_doc = client.model_dump()
            client_doc['created_at'] = datetime.now(timezone.utc).isoformat()
            
            await db.certified_clients.insert_one(client_doc)
            added_count += 1
    
    return {"message": f"Synced {added_count} certificates to registry"}

@api_router.get("/certified-clients/export/excel")
async def export_certified_clients_excel(
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Export certified clients registry to Excel"""
    await get_current_user(credentials)
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    
    clients = await db.certified_clients.find({}, {"_id": 0}).sort("serial_number", 1).to_list(1000)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Certified Clients"
    
    # Header styling
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="1a4d7c", end_color="1a4d7c", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title row
    ws.merge_cells('A1:N1')
    ws['A1'] = "BAYAN AUDITING & CONFORMITY - List of Certified Clients (BAC-F6-19)"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal="center")
    
    # Headers (Row 3)
    headers = [
        "S.No.", "Client Name", "Client Name (AR)", "Address", "Contact Person",
        "Contact No.", "Scope", "Accreditation", "EA Code", "Certificate No.",
        "Issue Date", "Expiry Date", "Surveillance 1", "Surveillance 2", "Recertification", "Status"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Data rows
    for row_idx, client in enumerate(clients, 4):
        ws.cell(row=row_idx, column=1).value = client.get('serial_number', row_idx - 3)
        ws.cell(row=row_idx, column=2).value = client.get('client_name', '')
        ws.cell(row=row_idx, column=3).value = client.get('client_name_ar', '')
        ws.cell(row=row_idx, column=4).value = client.get('address', '')
        ws.cell(row=row_idx, column=5).value = client.get('contact_person', '')
        ws.cell(row=row_idx, column=6).value = client.get('contact_number', '')
        ws.cell(row=row_idx, column=7).value = client.get('scope', '')
        ws.cell(row=row_idx, column=8).value = ', '.join(client.get('accreditation', []))
        ws.cell(row=row_idx, column=9).value = client.get('ea_code', '')
        ws.cell(row=row_idx, column=10).value = client.get('certificate_number', '')
        ws.cell(row=row_idx, column=11).value = client.get('issue_date', '')
        ws.cell(row=row_idx, column=12).value = client.get('expiry_date', '')
        ws.cell(row=row_idx, column=13).value = client.get('surveillance_1_date', '')
        ws.cell(row=row_idx, column=14).value = client.get('surveillance_2_date', '')
        ws.cell(row=row_idx, column=15).value = client.get('recertification_date', '')
        ws.cell(row=row_idx, column=16).value = client.get('status', 'active').upper()
        
        for col in range(1, 17):
            ws.cell(row=row_idx, column=col).border = thin_border
    
    # Adjust column widths
    column_widths = [8, 25, 25, 30, 20, 15, 30, 25, 12, 18, 12, 12, 12, 12, 12, 12]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i) if i < 27 else chr(64 + (i-1)//26) + chr(65 + (i-1)%26)].width = width
    
    # Save to bytes
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=certified_clients_registry.xlsx"}
    )

@api_router.get("/certified-clients/{client_id}/pdf")
async def get_certified_client_pdf(
    client_id: str,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Generate PDF for a single certified client record"""
    await get_current_user(credentials)
    
    client = await db.certified_clients.find_one({"id": client_id}, {"_id": 0})
    if not client:
        raise HTTPException(status_code=404, detail="Certified client not found")
    
    try:
        from certified_clients_generator import generate_certified_client_pdf
        
        contracts_dir = Path(__file__).parent / "contracts"
        contracts_dir.mkdir(exist_ok=True)
        pdf_path = str(contracts_dir / f"certified_client_{client_id[:8]}.pdf")
        
        generate_certified_client_pdf(client, pdf_path)
        
        return FileResponse(
            pdf_path,
            media_type="application/pdf",
            filename=f"certified_client_{client.get('certificate_number', client_id[:8])}.pdf"
        )
    except Exception as e:
        logging.error(f"Error generating Certified Client PDF: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error generating PDF: {str(e)}")

# ================= SUSPENDED CLIENTS REGISTRY ENDPOINTS (BAC-F6-20) =================

async def get_next_suspended_serial_number():
    """Get next serial number for suspended client registry"""
    last_record = await db.suspended_clients.find_one(
        {},
        sort=[("serial_number", -1)]
    )
    if last_record:
        return last_record.get('serial_number', 0) + 1
    return 1

@api_router.get("/suspended-clients")
async def get_suspended_clients(
    status: str = None,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Get all suspended client records"""
    await get_current_user(credentials)
    
    query = {}
    if status and status != 'all':
        query['status'] = status
    
    clients = await db.suspended_clients.find(query, {"_id": 0}).sort("serial_number", 1).to_list(1000)
    return clients

@api_router.post("/suspended-clients")
async def create_suspended_client(
    data: SuspendedClientCreate,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Create a new suspended client record"""
    await get_current_user(credentials)
    
    serial_number = await get_next_suspended_serial_number()
    
    client = SuspendedClient(
        serial_number=serial_number,
        client_id=data.client_id,
        client_name=data.client_name,
        client_name_ar=data.client_name_ar,
        address=data.address,
        address_ar=data.address_ar,
        registration_date=data.registration_date,
        suspended_on=data.suspended_on,
        reason_for_suspension=data.reason_for_suspension,
        reason_for_suspension_ar=data.reason_for_suspension_ar,
        future_action=data.future_action,
        remarks=data.remarks,
        linked_certified_client_id=data.linked_certified_client_id
    )
    
    client_doc = client.model_dump()
    client_doc['created_at'] = datetime.now(timezone.utc).isoformat()
    
    await db.suspended_clients.insert_one(client_doc)
    
    # Update linked certified client status if linked
    if data.linked_certified_client_id:
        await db.certified_clients.update_one(
            {"id": data.linked_certified_client_id},
            {"$set": {"status": "suspended", "updated_at": datetime.now(timezone.utc).isoformat()}}
        )
    
    await create_notification(
        notification_type="client_suspended",
        title="Client Suspended",
        message=f"Client '{data.client_name}' has been suspended. Reason: {data.reason_for_suspension[:50]}...",
        related_id=client.id,
        related_type="suspended_client"
    )
    
    return {k: v for k, v in client_doc.items() if k != '_id'}

@api_router.get("/suspended-clients/{client_id}")
async def get_suspended_client(
    client_id: str,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Get a specific suspended client record"""
    await get_current_user(credentials)
    
    client = await db.suspended_clients.find_one({"id": client_id}, {"_id": 0})
    if not client:
        raise HTTPException(status_code=404, detail="Suspended client record not found")
    
    return client

@api_router.put("/suspended-clients/{client_id}")
async def update_suspended_client(
    client_id: str,
    data: SuspendedClientUpdate,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Update a suspended client record"""
    await get_current_user(credentials)
    
    existing = await db.suspended_clients.find_one({"id": client_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Suspended client record not found")
    
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    update_data['updated_at'] = datetime.now(timezone.utc).isoformat()
    
    await db.suspended_clients.update_one(
        {"id": client_id},
        {"$set": update_data}
    )
    
    return {"message": "Suspended client record updated successfully"}

@api_router.delete("/suspended-clients/{client_id}")
async def delete_suspended_client(
    client_id: str,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Delete a suspended client record"""
    await get_current_user(credentials)
    
    existing = await db.suspended_clients.find_one({"id": client_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Suspended client record not found")
    
    await db.suspended_clients.delete_one({"id": client_id})
    return {"message": "Suspended client record deleted successfully"}

@api_router.get("/suspended-clients/stats/overview")
async def get_suspended_clients_stats(
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Get statistics for suspended clients registry"""
    await get_current_user(credentials)
    
    clients = await db.suspended_clients.find({}, {"_id": 0}).to_list(1000)
    
    suspended = len([c for c in clients if c.get('status') == 'suspended'])
    reinstated = len([c for c in clients if c.get('status') == 'reinstated'])
    withdrawn = len([c for c in clients if c.get('status') == 'withdrawn'])
    
    # Group by future action
    pending_reinstatement = len([c for c in clients if c.get('future_action') == 'reinstate' and c.get('status') == 'suspended'])
    pending_withdrawal = len([c for c in clients if c.get('future_action') == 'withdraw' and c.get('status') == 'suspended'])
    under_review = len([c for c in clients if c.get('future_action') == 'under_review' and c.get('status') == 'suspended'])
    
    return {
        "total": len(clients),
        "suspended": suspended,
        "reinstated": reinstated,
        "withdrawn": withdrawn,
        "pending_reinstatement": pending_reinstatement,
        "pending_withdrawal": pending_withdrawal,
        "under_review": under_review
    }

@api_router.post("/suspended-clients/{client_id}/lift-suspension")
async def lift_suspension(
    client_id: str,
    action: str = "reinstate",  # reinstate or withdraw
    reason: str = "",
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Lift suspension - either reinstate or withdraw certification"""
    await get_current_user(credentials)
    
    existing = await db.suspended_clients.find_one({"id": client_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Suspended client record not found")
    
    if existing.get('status') != 'suspended':
        raise HTTPException(status_code=400, detail="Client is not currently suspended")
    
    new_status = "reinstated" if action == "reinstate" else "withdrawn"
    today = datetime.now().strftime("%Y-%m-%d")
    
    await db.suspended_clients.update_one(
        {"id": client_id},
        {"$set": {
            "status": new_status,
            "lifted_on": today,
            "lifted_reason": reason,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    # Update linked certified client if exists
    if existing.get('linked_certified_client_id'):
        cert_status = "active" if action == "reinstate" else "withdrawn"
        await db.certified_clients.update_one(
            {"id": existing['linked_certified_client_id']},
            {"$set": {"status": cert_status, "updated_at": datetime.now(timezone.utc).isoformat()}}
        )
    
    await create_notification(
        notification_type="suspension_lifted",
        title=f"Suspension {new_status.title()}",
        message=f"Client '{existing.get('client_name', '')}' suspension has been lifted. Action: {new_status}",
        related_id=client_id,
        related_type="suspended_client"
    )
    
    return {"message": f"Suspension lifted. Client status changed to: {new_status}"}

@api_router.post("/suspended-clients/sync-from-certified")
async def sync_suspended_from_certified(
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Sync suspended clients from certified clients registry"""
    await get_current_user(credentials)
    
    # Get all certified clients with suspended status
    certified = await db.certified_clients.find({"status": "suspended"}, {"_id": 0}).to_list(1000)
    existing_links = await db.suspended_clients.distinct("linked_certified_client_id")
    
    added_count = 0
    for cert in certified:
        if cert.get('id') and cert.get('id') not in existing_links:
            serial_number = await get_next_suspended_serial_number()
            
            client = SuspendedClient(
                serial_number=serial_number,
                client_id=cert.get('certificate_number', ''),
                client_name=cert.get('client_name', ''),
                client_name_ar=cert.get('client_name_ar', ''),
                address=cert.get('address', ''),
                registration_date=cert.get('issue_date', ''),
                suspended_on=datetime.now().strftime("%Y-%m-%d"),
                reason_for_suspension="Synced from certified clients registry",
                future_action="under_review",
                linked_certified_client_id=cert.get('id', '')
            )
            
            client_doc = client.model_dump()
            client_doc['created_at'] = datetime.now(timezone.utc).isoformat()
            
            await db.suspended_clients.insert_one(client_doc)
            added_count += 1
    
    return {"message": f"Synced {added_count} suspended clients from certified clients registry"}

@api_router.get("/suspended-clients/export/excel")
async def export_suspended_clients_excel(
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Export suspended clients registry to Excel"""
    await get_current_user(credentials)
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    
    clients = await db.suspended_clients.find({}, {"_id": 0}).sort("serial_number", 1).to_list(1000)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Suspended Clients"
    
    # Header styling
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="c0392b", end_color="c0392b", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title row
    ws.merge_cells('A1:J1')
    ws['A1'] = "BAYAN AUDITING & CONFORMITY - List of Suspended Clients (BAC-F6-20)"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal="center")
    
    # Headers (Row 3)
    headers = [
        "Sr. No.", "Client ID", "Client Name", "Address", "Registration Date",
        "Suspended On", "Reason for Suspension", "Future Action", "Status", "Remarks"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Data rows
    for row_idx, client in enumerate(clients, 4):
        ws.cell(row=row_idx, column=1).value = client.get('serial_number', row_idx - 3)
        ws.cell(row=row_idx, column=2).value = client.get('client_id', '')
        ws.cell(row=row_idx, column=3).value = client.get('client_name', '')
        ws.cell(row=row_idx, column=4).value = client.get('address', '')
        ws.cell(row=row_idx, column=5).value = client.get('registration_date', '')
        ws.cell(row=row_idx, column=6).value = client.get('suspended_on', '')
        ws.cell(row=row_idx, column=7).value = client.get('reason_for_suspension', '')
        ws.cell(row=row_idx, column=8).value = client.get('future_action', '').replace('_', ' ').title()
        ws.cell(row=row_idx, column=9).value = client.get('status', 'suspended').upper()
        ws.cell(row=row_idx, column=10).value = client.get('remarks', '')
        
        for col in range(1, 11):
            ws.cell(row=row_idx, column=col).border = thin_border
    
    # Adjust column widths
    column_widths = [8, 15, 25, 30, 15, 15, 30, 15, 12, 25]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # Save to bytes
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=suspended_clients_registry.xlsx"}
    )

@api_router.get("/suspended-clients/{client_id}/pdf")
async def get_suspended_client_pdf(
    client_id: str,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Generate PDF for a suspended client record"""
    await get_current_user(credentials)
    
    client = await db.suspended_clients.find_one({"id": client_id}, {"_id": 0})
    if not client:
        raise HTTPException(status_code=404, detail="Suspended client record not found")
    
    try:
        from suspended_clients_generator import generate_suspended_client_pdf
        
        contracts_dir = Path(__file__).parent / "contracts"
        contracts_dir.mkdir(exist_ok=True)
        pdf_path = str(contracts_dir / f"suspended_client_{client_id[:8]}.pdf")
        
        generate_suspended_client_pdf(client, pdf_path)
        
        return FileResponse(
            pdf_path,
            media_type="application/pdf",
            filename=f"suspended_client_{client.get('client_id', client_id[:8])}.pdf"
        )
    except Exception as e:
        logging.error(f"Error generating Suspended Client PDF: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error generating PDF: {str(e)}")

# ================= CERTIFICATE ENDPOINTS =================

async def generate_certificate_number():
    """Generate unique certificate number: CERT-YYYY-XXXX"""
    year = datetime.now().year
    last_cert = await db.certificates.find_one(
        {"certificate_number": {"$regex": f"^CERT-{year}-"}},
        sort=[("certificate_number", -1)]
    )
    if last_cert:
        last_num = int(last_cert['certificate_number'].split('-')[-1])
        new_num = last_num + 1
    else:
        new_num = 1
    return f"CERT-{year}-{new_num:04d}"

@api_router.get("/certificates")
async def get_certificates(
    status: str = None,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Get all certificates with optional status filtering"""
    await get_current_user(credentials)
    
    query = {}
    if status and status != 'all':
        query['status'] = status
    
    certificates = await db.certificates.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    
    # Check for expired certificates and update status
    today = datetime.now().strftime("%Y-%m-%d")
    for cert in certificates:
        if cert.get('status') == 'active' and cert.get('expiry_date') and cert.get('expiry_date') < today:
            await db.certificates.update_one(
                {"id": cert['id']},
                {"$set": {"status": "expired"}}
            )
            cert['status'] = 'expired'
    
    return certificates

@api_router.get("/certificates/stats")
async def get_certificate_stats(credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Get certificate statistics"""
    await get_current_user(credentials)
    
    certificates = await db.certificates.find({}, {"_id": 0}).to_list(1000)
    today = datetime.now().strftime("%Y-%m-%d")
    
    # Calculate expiring soon (within 90 days)
    expiring_date = (datetime.now() + timedelta(days=90)).strftime("%Y-%m-%d")
    
    return {
        "total_certificates": len(certificates),
        "active_count": len([c for c in certificates if c.get('status') == 'active']),
        "expired_count": len([c for c in certificates if c.get('status') == 'expired' or (c.get('status') == 'active' and c.get('expiry_date', '') < today)]),
        "suspended_count": len([c for c in certificates if c.get('status') == 'suspended']),
        "expiring_soon_count": len([c for c in certificates if c.get('status') == 'active' and today < c.get('expiry_date', '') <= expiring_date])
    }

@api_router.post("/certificates")
async def create_certificate(cert_data: CertificateCreate, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Create a new certificate after successful audit"""
    await get_current_user(credentials)
    
    # Get contract/agreement data
    agreement = await db.certification_agreements.find_one({"id": cert_data.contract_id}, {"_id": 0})
    if not agreement:
        raise HTTPException(status_code=404, detail="Agreement not found")
    
    # Get proposal data for organization info
    proposal = await db.proposals.find_one({"id": agreement.get('proposal_id')}, {"_id": 0})
    if not proposal:
        raise HTTPException(status_code=404, detail="Proposal not found")
    
    # Get form data for detailed info
    form = await db.forms.find_one({"id": proposal.get('form_id')}, {"_id": 0})
    
    # Generate certificate number
    cert_number = await generate_certificate_number()
    
    # Calculate dates
    issue_date = datetime.now().strftime("%Y-%m-%d")
    expiry_date = (datetime.now() + timedelta(days=365*3)).strftime("%Y-%m-%d")  # 3 years validity
    
    # Generate verification URL
    base_url = os.environ.get('FRONTEND_URL', 'https://audit-workflow-pro-1.preview.emergentagent.com')
    verification_url = f"{base_url}/verify/{cert_number}"
    
    # Generate QR code
    qr_base64 = get_qr_code_base64(verification_url)
    
    # Get organization names
    org_name = form.get('responses', {}).get('organizationName', '') if form else agreement.get('organizationName', '')
    org_name_ar = form.get('responses', {}).get('organizationNameAr', '') if form else ''
    
    # Get standards
    standards = cert_data.standards if cert_data.standards else proposal.get('certification_standards', [])
    
    # Create certificate
    certificate = Certificate(
        certificate_number=cert_number,
        contract_id=cert_data.contract_id,
        audit_id=cert_data.audit_id,
        organization_name=org_name,
        organization_name_ar=org_name_ar,
        standards=standards,
        scope=cert_data.scope or proposal.get('scope', ''),
        scope_ar=cert_data.scope_ar or '',
        issue_date=issue_date,
        expiry_date=expiry_date,
        status="active",
        verification_url=verification_url,
        qr_code_data=qr_base64,
        lead_auditor=cert_data.lead_auditor,
        audit_team=cert_data.audit_team
    )
    
    cert_doc = certificate.model_dump()
    cert_doc['created_at'] = cert_doc['created_at'].isoformat()
    
    await db.certificates.insert_one(cert_doc)
    
    # Create notification
    await create_notification(
        notification_type="certificate_issued",
        title="Certificate Issued",
        message=f"Certificate {cert_number} has been issued for {org_name}",
        related_id=certificate.id,
        related_type="certificate"
    )
    
    return {"message": "Certificate created", "certificate_id": certificate.id, "certificate_number": cert_number}

@api_router.get("/certificates/{certificate_id}")
async def get_certificate(certificate_id: str, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Get certificate details"""
    await get_current_user(credentials)
    
    certificate = await db.certificates.find_one({"id": certificate_id}, {"_id": 0})
    if not certificate:
        raise HTTPException(status_code=404, detail="Certificate not found")
    
    return certificate

@api_router.put("/certificates/{certificate_id}/status")
async def update_certificate_status(certificate_id: str, status_data: dict, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Update certificate status (suspend, withdraw, reactivate)"""
    await get_current_user(credentials)
    
    certificate = await db.certificates.find_one({"id": certificate_id})
    if not certificate:
        raise HTTPException(status_code=404, detail="Certificate not found")
    
    new_status = status_data.get('status')
    valid_statuses = ['active', 'suspended', 'withdrawn', 'expired']
    if new_status not in valid_statuses:
        raise HTTPException(status_code=400, detail=f"Invalid status. Must be one of {valid_statuses}")
    
    await db.certificates.update_one(
        {"id": certificate_id},
        {"$set": {
            "status": new_status,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"message": f"Certificate status updated to {new_status}"}

@api_router.get("/certificates/{certificate_id}/pdf")
async def download_certificate_pdf(certificate_id: str, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Generate and download certificate PDF"""
    await get_current_user(credentials)
    
    certificate = await db.certificates.find_one({"id": certificate_id}, {"_id": 0})
    if not certificate:
        raise HTTPException(status_code=404, detail="Certificate not found")
    
    # Generate PDF
    CERTIFICATES_DIR = ROOT_DIR / "certificates"
    CERTIFICATES_DIR.mkdir(exist_ok=True)
    
    pdf_filename = f"certificate_{certificate['certificate_number'].replace('-', '_')}.pdf"
    pdf_path = CERTIFICATES_DIR / pdf_filename
    
    generate_certificate_pdf(certificate, str(pdf_path))
    
    return FileResponse(
        str(pdf_path),
        media_type="application/pdf",
        filename=pdf_filename
    )

@api_router.get("/public/verify/{certificate_number}")
async def verify_certificate_public(certificate_number: str):
    """Public endpoint to verify a certificate via QR code scan"""
    certificate = await db.certificates.find_one(
        {"certificate_number": certificate_number},
        {"_id": 0, "qr_code_data": 0}  # Exclude QR code data for public
    )
    
    if not certificate:
        raise HTTPException(status_code=404, detail="Certificate not found")
    
    # Check if expired
    today = datetime.now().strftime("%Y-%m-%d")
    if certificate.get('expiry_date') and certificate.get('expiry_date') < today:
        certificate['status'] = 'expired'
    
    return {
        "valid": certificate.get('status') == 'active',
        "certificate_number": certificate.get('certificate_number'),
        "organization_name": certificate.get('organization_name'),
        "organization_name_ar": certificate.get('organization_name_ar'),
        "standards": certificate.get('standards', []),
        "scope": certificate.get('scope'),
        "issue_date": certificate.get('issue_date'),
        "expiry_date": certificate.get('expiry_date'),
        "status": certificate.get('status'),
        "lead_auditor": certificate.get('lead_auditor')
    }

# ================= EXPIRATION ALERTS ENDPOINTS =================

@api_router.get("/alerts/expiring")
async def get_expiring_items(
    days: int = 90,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Get all items expiring within specified days (certificates, contracts)"""
    await get_current_user(credentials)
    
    today = datetime.now()
    expiry_date = (today + timedelta(days=days)).strftime("%Y-%m-%d")
    today_str = today.strftime("%Y-%m-%d")
    
    # Get expiring certificates
    expiring_certs = await db.certificates.find({
        "status": "active",
        "expiry_date": {"$gte": today_str, "$lte": expiry_date}
    }, {"_id": 0}).to_list(100)
    
    # Categorize by urgency
    alerts = {
        "critical": [],  # Expiring within 30 days
        "warning": [],   # Expiring within 60 days
        "info": []       # Expiring within 90 days
    }
    
    for cert in expiring_certs:
        exp_date = datetime.strptime(cert['expiry_date'], "%Y-%m-%d")
        days_until = (exp_date - today).days
        
        alert_item = {
            "type": "certificate",
            "id": cert['id'],
            "reference": cert['certificate_number'],
            "organization": cert['organization_name'],
            "expiry_date": cert['expiry_date'],
            "days_until_expiry": days_until,
            "standards": cert.get('standards', [])
        }
        
        if days_until <= 30:
            alerts["critical"].append(alert_item)
        elif days_until <= 60:
            alerts["warning"].append(alert_item)
        else:
            alerts["info"].append(alert_item)
    
    # Get upcoming surveillance audits
    upcoming_audits = await db.audit_schedules.find({
        "status": {"$ne": "completed"},
        "scheduled_date": {"$gte": today_str, "$lte": expiry_date}
    }, {"_id": 0}).to_list(100)
    
    for audit in upcoming_audits:
        audit_date = datetime.strptime(audit['scheduled_date'], "%Y-%m-%d")
        days_until = (audit_date - today).days
        
        alert_item = {
            "type": "audit",
            "id": audit['id'],
            "reference": audit.get('audit_type', 'Audit'),
            "organization": audit.get('organization_name', ''),
            "expiry_date": audit['scheduled_date'],
            "days_until_expiry": days_until,
            "audit_type": audit.get('audit_type', '')
        }
        
        if days_until <= 7:
            alerts["critical"].append(alert_item)
        elif days_until <= 14:
            alerts["warning"].append(alert_item)
        else:
            alerts["info"].append(alert_item)
    
    return {
        "summary": {
            "total_alerts": len(alerts["critical"]) + len(alerts["warning"]) + len(alerts["info"]),
            "critical_count": len(alerts["critical"]),
            "warning_count": len(alerts["warning"]),
            "info_count": len(alerts["info"])
        },
        "alerts": alerts
    }

@api_router.get("/dashboard/analytics")
async def get_dashboard_analytics(credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Get comprehensive analytics for the dashboard"""
    await get_current_user(credentials)
    
    # Get all data
    forms = await db.forms.find({}, {"_id": 0}).to_list(1000)
    proposals = await db.proposals.find({}, {"_id": 0}).to_list(1000)
    agreements = await db.certification_agreements.find({}, {"_id": 0}).to_list(1000)
    invoices = await db.invoices.find({}, {"_id": 0}).to_list(1000)
    certificates = await db.certificates.find({}, {"_id": 0}).to_list(1000)
    audits = await db.audit_schedules.find({}, {"_id": 0}).to_list(1000)
    
    # Calculate conversion rates
    total_forms = len(forms)
    submitted_forms = len([f for f in forms if f.get('status') in ['submitted', 'proposal_sent', 'proposal_accepted', 'agreement_signed']])
    proposals_sent = len(proposals)
    proposals_accepted = len([p for p in proposals if p.get('status') == 'agreement_signed'])
    contracts_signed = len(agreements)
    
    form_to_proposal_rate = (proposals_sent / total_forms * 100) if total_forms > 0 else 0
    proposal_to_contract_rate = (contracts_signed / proposals_sent * 100) if proposals_sent > 0 else 0
    overall_conversion_rate = (contracts_signed / total_forms * 100) if total_forms > 0 else 0
    
    # Revenue analytics
    total_quoted = sum(p.get('total_fees', 0) for p in proposals)
    total_accepted = sum(p.get('total_fees', 0) for p in proposals if p.get('status') == 'agreement_signed')
    total_invoiced = sum(i.get('total_amount', 0) for i in invoices)
    total_paid = sum(i.get('paid_amount', 0) for i in invoices if i.get('status') == 'paid')
    
    # Monthly data (last 6 months)
    monthly_data = []
    for i in range(5, -1, -1):
        month_date = datetime.now() - timedelta(days=30*i)
        month_str = month_date.strftime("%Y-%m")
        
        month_forms = len([f for f in forms if f.get('created_at', '')[:7] == month_str])
        month_proposals = len([p for p in proposals if p.get('created_at', '')[:7] == month_str])
        month_contracts = len([a for a in agreements if a.get('created_at', '')[:7] == month_str])
        month_revenue = sum(p.get('total_fees', 0) for p in proposals if p.get('created_at', '')[:7] == month_str and p.get('status') == 'agreement_signed')
        
        monthly_data.append({
            "month": month_date.strftime("%b %Y"),
            "forms": month_forms,
            "proposals": month_proposals,
            "contracts": month_contracts,
            "revenue": month_revenue
        })
    
    # Standards breakdown
    standards_count = {}
    for p in proposals:
        for std in p.get('certification_standards', []):
            standards_count[std] = standards_count.get(std, 0) + 1
    
    # Audit status
    completed_audits = len([a for a in audits if a.get('status') == 'completed'])
    scheduled_audits = len([a for a in audits if a.get('status') == 'scheduled'])
    pending_audits = len([a for a in audits if a.get('status') == 'pending'])
    
    return {
        "overview": {
            "total_forms": total_forms,
            "total_proposals": proposals_sent,
            "total_contracts": contracts_signed,
            "total_certificates": len(certificates),
            "active_certificates": len([c for c in certificates if c.get('status') == 'active'])
        },
        "conversion_rates": {
            "form_to_proposal": round(form_to_proposal_rate, 1),
            "proposal_to_contract": round(proposal_to_contract_rate, 1),
            "overall": round(overall_conversion_rate, 1)
        },
        "revenue": {
            "total_quoted": total_quoted,
            "total_accepted": total_accepted,
            "total_invoiced": total_invoiced,
            "total_collected": total_paid,
            "collection_rate": round((total_paid / total_invoiced * 100) if total_invoiced > 0 else 0, 1)
        },
        "monthly_trends": monthly_data,
        "standards_breakdown": standards_count,
        "audits": {
            "total": len(audits),
            "completed": completed_audits,
            "scheduled": scheduled_audits,
            "pending": pending_audits
        }
    }

@app.on_event("startup")
async def startup_event():
    """Run startup tasks"""
    await seed_default_templates()

# Include the main api_router in the app
app.include_router(api_router)

# Include modular routers
app.include_router(auth_router, prefix="/api")
app.include_router(notifications_router, prefix="/api")
app.include_router(sites_router, prefix="/api")
app.include_router(contacts_router, prefix="/api")
app.include_router(documents_router, prefix="/api")

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
